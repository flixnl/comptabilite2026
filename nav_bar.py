#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""dashboard/nav_bar.py
=========================================================
Bandeau de navigation interne pour le cockpit comptable.

Écrit un menu horizontal en ligne 1 avec des hyperliens internes
vers les feuilles principales. La feuille active est mise en
évidence (fond bleu, texte blanc).

Compatible openpyxl — pas de macros, pas de formes.
"""
from __future__ import annotations
from typing import Optional
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

# ── Configuration ────────────────────────────────────────────────────────────

NAV_ITEMS: list[tuple[str, str]] = [
    ('📊 Dashboard',      'DASHBOARD V2'),
    ('💰 Obligations',    'OBLIGATIONS'),
    ('📅 Acomptes',       'ACOMPTES'),
    ('🧮 Estimation',     'ESTIMATION FISCALE'),
    ('🧾 Taxes',          'TAXES MARS'),
]

NAV_ROW        = 1
NAV_HEIGHT     = 24.0
NAV_START_COL  = 2     # col B

# ── Palette ──────────────────────────────────────────────────────────────────

_BG_NORMAL  = 'F3F4F6'   # gris clair
_FG_NORMAL  = '374151'   # gris foncé
_BG_ACTIVE  = '1D4ED8'   # bleu
_FG_ACTIVE  = 'FFFFFF'   # blanc
_BG_EDGE    = 'F3F4F6'   # cellules hors-menu (même fond pour uniformité)
_BORDER_BOT = Border(bottom=Side(style='thin', color='D1D5DB'))
_NO_BORDER  = Border()

_FONT_NORMAL = Font(name='Calibri', size=9, bold=True, color=_FG_NORMAL)
_FONT_ACTIVE = Font(name='Calibri', size=9, bold=True, color=_FG_ACTIVE)
_FILL_NORMAL = PatternFill('solid', fgColor=_BG_NORMAL)
_FILL_ACTIVE = PatternFill('solid', fgColor=_BG_ACTIVE)
_FILL_EDGE   = PatternFill('solid', fgColor=_BG_EDGE)
_ALIGN       = Alignment(horizontal='center', vertical='center')


# ── API publique ─────────────────────────────────────────────────────────────

def write_nav_bar(ws: Worksheet, active_sheet: str,
                  total_cols: int = 12) -> None:
    """Écrit le bandeau de navigation en ligne 1.

    Parameters
    ----------
    ws            : Feuille openpyxl cible (doit exister).
    active_sheet  : Nom exact de la feuille active (pour la mise en évidence).
    total_cols    : Nombre de colonnes à couvrir pour le fond uniforme.
    """
    ws.row_dimensions[NAV_ROW].height = NAV_HEIGHT

    # Fond uniforme sur toute la largeur
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=NAV_ROW, column=c)
        cell.fill = _FILL_EDGE
        cell.border = _BORDER_BOT
        cell.font = Font(name='Calibri', size=2, color=_BG_EDGE)  # invisible
        cell.alignment = _ALIGN

    # Écrire chaque item de navigation
    for i, (label, sheet_name) in enumerate(NAV_ITEMS):
        col = NAV_START_COL + i
        cell = ws.cell(row=NAV_ROW, column=col)

        is_active = (sheet_name == active_sheet)
        cell.value = label
        cell.font = _FONT_ACTIVE if is_active else _FONT_NORMAL
        cell.fill = _FILL_ACTIVE if is_active else _FILL_NORMAL
        cell.alignment = _ALIGN
        cell.border = _BORDER_BOT

        # Hyperlien interne (sauf feuille active → pas besoin)
        # On utilise Hyperlink(location=...) pour éviter de créer un rId
        # (les hyperlinks internes ne nécessitent PAS de relationship).
        if not is_active:
            hl = Hyperlink(ref=cell.coordinate,
                           location=f"'{sheet_name}'!A1")
            hl.display = label
            cell.hyperlink = hl
            cell.value = label          # re-set après hyperlink
            cell.font = _FONT_NORMAL    # supprimer le soulignement bleu


def freeze_below_nav(ws: Worksheet) -> None:
    """Place le freeze pane juste sous le bandeau nav (ligne 2)."""
    ws.freeze_panes = 'A2'
