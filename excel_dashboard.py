#!/usr/bin/env python3
"""
dashboard/excel_dashboard.py — Écriture de l'onglet COCKPIT dans VENTES_2026.xlsx.

Structure de l'onglet COCKPIT
──────────────────────────────
  A. EN-TÊTE          — titre + date de génération
  B. REVENUS          — encaissé / non-encaissé / total / ratio
  C. COMPTES À RECEVOIR (AR) — total, count, aging (0-30j / 30-60j / 60+j)
  D. FISCAL           — TPS/TVQ collectée, impôt estimé, total obligations
  E. ACOMPTES         — calendrier + prochain acompte
  F. ALERTES          — liste priorisée HIGH → MEDIUM → LOW
  G. SCORE GLOBAL     — label + emoji
  H. MAINTENANCE AUTOPILOT — dernier run, anomalies, corrections, actions manuelles

Usage
─────
    from dashboard.excel_dashboard import write_cockpit_tab
    ok = write_cockpit_tab('/path/to/VENTES_2026.xlsx', ctx)
    # Avec résultat autopilot optionnel :
    ok = write_cockpit_tab('/path/to/VENTES_2026.xlsx', ctx, autopilot_result=result)
"""
from __future__ import annotations

import os
import shutil
from datetime import date, datetime
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from dashboard.context import CockpitContext

# Nom de l'onglet COCKPIT
COCKPIT_SHEET = 'COCKPIT'


# ─── Helpers de mise en forme ─────────────────────────────────────────────────

def _money(v: float | None, decimals: int = 2) -> str:
    if v is None:
        return '—'
    fmt = f'{{:,.{decimals}f}} $'
    return fmt.format(v)


def _pct(v: float | None) -> str:
    if v is None:
        return '—'
    return f'{v * 100:.1f} %'


# ─── Écriture dans openpyxl ───────────────────────────────────────────────────

def _write_rows(ws, rows: list[tuple], start_row: int = 1):
    """Écrit une liste de (col_A, col_B) dans la feuille."""
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            ws.cell(row=start_row + i, column=j + 1, value=val)
    return start_row + len(rows)


def _section_header(ws, row: int, title: str):
    """Écrit un séparateur de section avec le titre en A."""
    ws.cell(row=row, column=1, value=f'── {title} ──')
    return row + 1


# ─── Constructeur des données ─────────────────────────────────────────────────

def _build_cockpit_rows(ctx: 'CockpitContext', today: date, autopilot_result=None) -> list[tuple]:
    """Construit toutes les lignes du cockpit comme liste de (label, valeur)."""
    rows: list[tuple] = []
    ac = ctx.alert_ctx

    # EN-TÊTE
    rows.append(('COCKPIT COMPTABILITÉ 2026', ''))
    rows.append((f'Généré le {today.strftime("%Y-%m-%d %H:%M")}', ''))
    rows.append(('', ''))

    # ── A. REVENUS ────────────────────────────────────────────────────────────
    rows.append(('── A. REVENUS ──', ''))
    if ac.metrics:
        m = ac.metrics
        rows.append(('Revenus encaissés (ST)',    _money(m.encaisse_st)))
        rows.append(('Revenus encaissés (TTC)',   _money(m.encaisse_ttc)))
        rows.append(('Dont AQTIS (ST)',           _money(m.encaisse_aqtis_st)))
        rows.append(('Non encaissé (ST)',         _money(m.non_encaisse_st)))
        rows.append(('Non encaissé (TTC)',        _money(m.non_encaisse_ttc)))
        rows.append(('Total facturé (ST)',        _money(m.total_st)))
        rows.append(('Ratio encaissement',        _pct(m.ratio_encaisse)))
        if m.plus_grosse_non_encaisse:
            e = m.plus_grosse_non_encaisse
            rows.append(('Plus grosse non encaissée', f'{e.ref} — {_money(e.st)}'))
    else:
        rows.append(('Données revenus', 'indisponibles'))
    rows.append(('', ''))

    # ── B. COMPTES À RECEVOIR ─────────────────────────────────────────────────
    rows.append(('── B. COMPTES À RECEVOIR (AR) ──', ''))
    if ac.ar:
        ar = ac.ar
        rows.append(('AR total (ST)',    _money(ar.total_st)))
        rows.append(('AR total (TTC)',   _money(ar.total_ttc)))
        rows.append(('Nombre factures', str(ar.count)))
        rows.append(('', ''))
        rows.append(('Aging — 0-30 j',  _money(ar.aging['0-30j']['st'])))
        rows.append(('Aging — 30-60 j', _money(ar.aging['30-60j']['st'])))
        rows.append(('Aging — 60+ j',   _money(ar.aging['60+ j']['st'])))
        rows.append(('', ''))
        # Top 5 AR
        rows.append(('Top factures non encaissées', ''))
        top5 = sorted(ar.entries, key=lambda e: -(e.st or 0))[:5]
        for e in top5:
            age_str = f'{e.age_days(today)}j' if e.age_days(today) is not None else '?j'
            rows.append((f'  {e.ref} ({e.client[:20]})', f'{_money(e.st)}  [{age_str}]'))
    else:
        rows.append(('Données AR', 'indisponibles'))
    rows.append(('', ''))

    # ── C. FISCAL ─────────────────────────────────────────────────────────────
    rows.append(('── C. FISCAL ──', ''))
    if ac.fiscal:
        f = ac.fiscal
        rows.append(('TPS collectée (encaissée)',  _money(f.tps_collectee)))
        rows.append(('TVQ collectée (encaissée)',  _money(f.tvq_collectee)))
        rows.append(('TPS à recevoir (AR)',        _money(f.tps_a_recevoir)))
        rows.append(('TVQ à recevoir (AR)',        _money(f.tvq_a_recevoir)))
        rows.append(('Revenu brut YTD',            _money(f.revenu_brut_ytd)))
        rows.append(('Revenu net YTD',             _money(f.revenu_net_ytd)))
        rows.append(('Déductions AQTIS YTD',       _money(f.aqtis_deductions_ytd)))

    # ── Position TPS/TVQ (source de vérité) ───────────────────────────────────
    if ac.taxes_position is not None:
        pos = ac.taxes_position
        rows.append(('', ''))
        rows.append(('Position TPS / TVQ', ''))
        rows.append(('TPS collectées (total)',   _money(pos.tps_collectees)))
        rows.append(('TVQ collectées (total)',   _money(pos.tvq_collectees)))
        rows.append(('CTI (crédits TPS)',        _money(pos.cti)))
        rows.append(('RTI (crédits TVQ)',        _money(pos.rti)))
        rows.append(('TPS nette due',            _money(pos.tps_nette_due)))
        rows.append(('TVQ nette due',            _money(pos.tvq_nette_due)))
        rows.append(('TPS payées au gouv.',      _money(pos.tps_payees)))
        rows.append(('TVQ payées au gouv.',      _money(pos.tvq_payees)))
        rows.append(('TPS RESTANTES',            _money(pos.tps_restantes)))
        rows.append(('TVQ RESTANTES',            _money(pos.tvq_restantes)))
        rows.append(('TOTAL TAXES RESTANTES',    _money(pos.total_restantes)))
        if pos.is_overpaid:
            rows.append(('⚠️ Trop-payé',         _money(abs(pos.tps_surplus + pos.tvq_surplus))))
        if pos.plan_payments_total > 0:
            rows.append(('Plan de paiement (versements)', _money(pos.plan_payments_total)))
        if pos.installments_total > 0:
            rows.append(('Acomptes provisionnels versés', _money(pos.installments_total)))

    if ac.tax_estimate:
        te = ac.tax_estimate
        rows.append(('', ''))
        rows.append(('Estimation fiscale (impôts)', ''))
        rows.append(('Revenu imposable',        _money(te.revenu_imposable)))
        rows.append(('Impôt fédéral',           _money(te.impot_federal)))
        rows.append(('Impôt Québec',            _money(te.impot_qc)))
        rows.append(('Cotisations autonomes',   _money(te.cotisations_autonomes)))
        rows.append(('Total obligations',       _money(te.total_obligations)))
        rows.append(('Taux effectif global',    _pct(te.taux_effectif_global)))
    if not ac.fiscal and not ac.tax_estimate:
        rows.append(('Données fiscales', 'indisponibles'))
    rows.append(('', ''))

    # ── D. ACOMPTES ───────────────────────────────────────────────────────────
    rows.append(('── D. ACOMPTES PROVISIONNELS ──', ''))
    if ac.installments:
        sched = ac.installments
        rows.append(('Acomptes requis', 'Oui' if sched.required else 'Non'))
        if sched.required:
            rows.append(('Total acomptes annuels', _money(sched.total)))
            for p in sched.periods:
                status = p.status(today)
                icon   = {'EN RETARD': '🔴', 'IMMINENT': '🟡', 'À VENIR': '📅'}.get(status, '')
                rows.append((f'  {p.label} — {p.due_date}',
                             f'{_money(p.amount)}  {icon} {status}'))
        if ac.next_installment:
            nxt = ac.next_installment
            rows.append(('', ''))
            rows.append(('Prochain acompte',
                         f'{nxt.period.due_date}  ({nxt.status})'))
    else:
        rows.append(('Données acomptes', 'indisponibles'))
    rows.append(('', ''))

    # ── E-bis. PLAN NANTEL ────────────────────────────────────────────────────
    try:
        from fiscal.payment_plan import nantel_plan, build_plan_summary
        _plan   = nantel_plan()
        _summ   = build_plan_summary(_plan, today)
        rows.append(('── E-bis. PLAN NANTEL ──', ''))
        rows.append(('Montant mensuel',          f'{_plan.amount:,.2f} $  /  {_plan.day_of_month} du mois'))
        rows.append(('Début du plan',             str(_plan.start_date)))
        rows.append(('Total payé (all-time)',      _money(_summ.total_paid_all_time)))
        rows.append(('Total attendu YTD',          _money(_summ.total_expected_ytd)))
        rows.append(('Total payé YTD',             _money(_summ.total_paid_ytd)))
        n_late   = len(_summ.late_payments)
        n_missed = len(_summ.missed_payments)
        if n_missed:
            rows.append(('⛔ Paiements manqués',   str(n_missed)))
        if n_late:
            rows.append(('🔴 Paiements en retard', str(n_late)))
        if _summ.is_on_track:
            rows.append(('Statut',                 '✅ À jour'))
        nxt = _summ.next_payment
        if nxt:
            rows.append(('Prochain versement',     f'{nxt.due_date}  {_money(nxt.amount)}'))
        rows.append(('', ''))
    except Exception:
        pass   # Plan Nantel indisponible — section omise silencieusement

    # ── E. ALERTES ────────────────────────────────────────────────────────────
    rows.append(('── E. ALERTES ──', ''))
    if ctx.alerts:
        for a in ctx.alerts:
            val_str = f'  [{_money(a.value)}]' if a.value is not None else ''
            rows.append((f'{a.icon} [{a.severity}] {a.message}{val_str}', ''))
    else:
        rows.append(('✅ Aucune alerte', ''))
    rows.append(('', ''))

    # ── F. SCORE GLOBAL ───────────────────────────────────────────────────────
    rows.append(('── F. SCORE GLOBAL ──', ''))
    rows.append(('Score',
                 f'{ctx.score_emoji}  {ctx.score_label}'))
    n_high   = sum(1 for a in ctx.alerts if a.severity == 'HIGH')
    n_medium = sum(1 for a in ctx.alerts if a.severity == 'MEDIUM')
    n_low    = sum(1 for a in ctx.alerts if a.severity == 'LOW')
    rows.append(('Alertes HIGH',   str(n_high)))
    rows.append(('Alertes MEDIUM', str(n_medium)))
    rows.append(('Alertes LOW',    str(n_low)))

    if ctx.error:
        rows.append(('', ''))
        rows.append(('⚠️ Erreur partielle', ctx.error[:120]))

    # ── H. MAINTENANCE AUTOPILOT ──────────────────────────────────────────────
    rows.append(('', ''))
    rows.append(('── H. MAINTENANCE AUTOPILOT ──', ''))
    if autopilot_result is not None:
        ap = autopilot_result
        run_str = (ap.run_date.strftime('%Y-%m-%d')
                   if hasattr(ap.run_date, 'strftime') else str(ap.run_date))
        mode_str = '🔍 Dry-run' if ap.dry_run else '✅ Apply'
        rows.append(('Dernier run',           f'{run_str}  —  {mode_str}'))
        rows.append(('Score système',         f'{ap.score_emoji}  {ap.score_label}'))
        rows.append(('Durée',                 f'{ap.duration_secs:.1f} s'))
        rows.append(('Anomalies HIGH',        str(ap.n_high)))
        rows.append(('Anomalies MEDIUM',      str(ap.n_medium)))
        rows.append(('Anomalies LOW',         str(ap.n_low)))
        rows.append(('Total anomalies',       str(len(ap.anomalies))))
        rows.append(('Corrections auto dispo',str(ap.n_safe_fixes)))
        rows.append(('Corrections appliquées',str(len(ap.corrections_applied))))
        rows.append(('Actions manuelles',     str(ap.n_manual)))
        if ap.corrections_applied:
            rows.append(('', ''))
            rows.append(('Corrections effectuées', ''))
            for c in ap.corrections_applied[:6]:
                rows.append((f'  {c}', ''))
            if len(ap.corrections_applied) > 6:
                rows.append((f'  … +{len(ap.corrections_applied) - 6} autres', ''))
        if ap.manual_actions:
            rows.append(('', ''))
            rows.append(('Actions manuelles requises', ''))
            for i, action in enumerate(ap.manual_actions[:5], 1):
                rows.append((f'  {i}. {action[:80]}', ''))
            if len(ap.manual_actions) > 5:
                rows.append((f'  … +{len(ap.manual_actions) - 5} autres', ''))
        if ap.report_path:
            rows.append(('', ''))
            rows.append(('Rapport',           os.path.basename(ap.report_path)))
        if ap.error:
            rows.append(('⚠️ Erreur run',     ap.error[:100]))
    else:
        rows.append(('Autopilot',             'Jamais exécuté'))
        rows.append(('',                      'Lancer via : python3 autopilot/run_autopilot.py'))

    return rows


# ─── write_cockpit_tab ────────────────────────────────────────────────────────

def write_cockpit_tab(
    ventes_excel_path: str,
    ctx: 'CockpitContext',
    backup: bool = True,
    autopilot_result=None,
) -> bool:
    """Écrit (ou recrée) l'onglet COCKPIT dans VENTES_2026.xlsx.

    Ne modifie AUCUN autre onglet.

    Parameters
    ----------
    ventes_excel_path : chemin absolu vers VENTES_AAAA.xlsx
    ctx               : CockpitContext (depuis build_cockpit_context)
    backup            : si True, crée une copie de sauvegarde avant écriture
    autopilot_result  : AutopilotResult optionnel pour section H

    Returns
    -------
    True si succès, False si erreur (l'exception est loguée dans ctx.error).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    if not os.path.isfile(ventes_excel_path):
        return False

    today = ctx.reference_date or date.today()

    # ── Backup ────────────────────────────────────────────────────────────────
    if backup:
        ts      = datetime.now().strftime('%Y%m%d_%H%M%S')
        base    = os.path.splitext(ventes_excel_path)[0]
        bck     = f'{base}_backup_{ts}_COCKPIT.xlsx'
        try:
            shutil.copy2(ventes_excel_path, bck)
        except Exception:
            pass   # non-fatal

    # ── Chargement ────────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(ventes_excel_path)
    except Exception:
        return False

    # ── Suppression + recréation ──────────────────────────────────────────────
    if COCKPIT_SHEET in wb.sheetnames:
        del wb[COCKPIT_SHEET]
    ws = wb.create_sheet(COCKPIT_SHEET)

    # ── Mise en page basique ──────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 30

    # ── Données ───────────────────────────────────────────────────────────────
    rows = _build_cockpit_rows(ctx, today, autopilot_result=autopilot_result)

    HEADER_FILL  = PatternFill('solid', fgColor='2F75B6')
    SECTION_FILL = PatternFill('solid', fgColor='D9E1F2')
    HEADER_FONT  = Font(bold=True, color='FFFFFF', size=13)
    SECTION_FONT = Font(bold=True, color='1F3864')

    for i, (label, value) in enumerate(rows, start=1):
        cell_a = ws.cell(row=i, column=1, value=label)
        cell_b = ws.cell(row=i, column=2, value=value)

        if i == 1:
            # Titre principal
            cell_a.font = HEADER_FONT
            cell_a.fill = HEADER_FILL
        elif isinstance(label, str) and label.startswith('── '):
            # Séparateur de section
            cell_a.font = SECTION_FONT
            cell_a.fill = SECTION_FILL
            cell_b.fill = SECTION_FILL

    # Figer la première ligne
    ws.freeze_panes = 'A2'

    # ── Forcer recalcul à l'ouverture (H10 et autres formules DÉPENSES) ──────
    # openpyxl écrit les formules mais ne les évalue pas.
    # fullCalcOnLoad garantit qu'Excel recalcule tout dès l'ouverture du fichier.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass   # non-fatal si la propriété n'est pas supportée

    # ── Sauvegarde ────────────────────────────────────────────────────────────
    try:
        wb.save(ventes_excel_path)
    except Exception:
        return False

    return True
