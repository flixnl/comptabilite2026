#!/usr/bin/env python3
"""
dashboard/context.py — Agrégateur central du cockpit de pilotage.

Construit un CockpitContext complet à partir des chemins VENTES + DÉPENSES.
Délègue à : ar.calculator, analytics.alerts, analytics.ventes_metrics,
            analytics.anomalies, fiscal.estimator, fiscal.installments.

Usage
─────
    from dashboard.context import build_cockpit_context
    ctx = build_cockpit_context(ventes_path, depenses_path)

    # Accès direct aux sous-contextes
    ctx.alert_ctx.ar.total_st
    ctx.alerts       # list[Alert] triée HIGH→MEDIUM→LOW
    ctx.score_label  # 'OK' | 'Attention' | 'Risque'
    ctx.score_emoji  # '🟢' | '🟡' | '🔴'
    ctx.obligations  # list[dict] obligations actives depuis DÉPENSES.xlsm
"""
from __future__ import annotations

import glob as _glob_mod
import json as _json
import os
from dataclasses import dataclass, field
from datetime import date
from typing import Optional, TYPE_CHECKING

if TYPE_CHECKING:
    from analytics.alerts import AlertContext, Alert


# ─── _read_obligations ────────────────────────────────────────────────────────

def _read_obligations(wb) -> list[dict]:
    """Lit les obligations actives depuis la feuille OBLIGATIONS.

    Retourne une liste de dicts :
      {'label', 'paid', 'total', 'pct', 'due_date', 'source', 'row'}

    Source canonique : dashboard/context.py
    Importée par     : dashboard/write_tableau_de_bord.py
    """
    if 'OBLIGATIONS' not in wb.sheetnames:
        return []

    ws_obl = wb['OBLIGATIONS']
    result = []

    for row in range(2, 20):  # jusqu'à 18 obligations max
        obl_id = ws_obl.cell(row=row, column=1).value
        if not obl_id:
            break
        status = ws_obl.cell(row=row, column=15).value  # O = STATUT
        if status and str(status).lower() != 'actif':
            continue

        paid    = float(ws_obl.cell(row=row, column=11).value or 0)  # K
        total   = float(ws_obl.cell(row=row, column=17).value or 0)  # Q
        pct     = (paid / total) if total > 0 else 0.0
        due_raw = ws_obl.cell(row=row, column=7).value  # G = DATE_FIN

        # Parse due_date (handles datetime, date, or Excel serial number)
        due_dt = None
        if due_raw is not None:
            if hasattr(due_raw, 'date'):
                due_dt = due_raw.date() if callable(due_raw.date) else due_raw
            elif hasattr(due_raw, 'year'):
                due_dt = due_raw
            elif isinstance(due_raw, (int, float)) and 40000 < due_raw < 60000:
                # Excel serial number → convert to date
                from datetime import datetime as _dt, timedelta as _td
                due_dt = (_dt(1899, 12, 30) + _td(days=int(due_raw))).date()

        label = ws_obl.cell(row=row, column=2).value or str(obl_id)  # B
        source = ws_obl.cell(row=row, column=20).value or ''  # T = SOURCE

        result.append({
            'label':    str(label),
            'paid':     paid,
            'total':    total,
            'pct':      pct,
            'due_date': due_dt,
            'source':   str(source),
            'row':      row,
        })

    return result


# ─── _build_monthly_revenue (STEP 1.2) ──────────────────────────────────────

def _build_monthly_revenue(entries, annee: int) -> list[dict]:
    """Construit les revenus mensuels depuis les VenteEntry déjà chargées.

    Retourne une liste triée par mois :
      [{'mois': 202601, 'freelance': ..., 'salaire': ...}, ...]

    Logique miroir de generate_dashboard.py::read_ventes() :
    - freelance = payé + tps_flag == 'Oui'
    - salaire   = payé + tps_flag != 'Oui'
    """
    rev: dict[int, dict] = {}  # cle_mois_int → {freelance, salaire}

    for e in entries:
        if e.is_doublon:
            continue
        cm = e.cle_mois
        # Normaliser cle_mois en int (202601) — peut être str '2026-01' ou int
        if isinstance(cm, str) and '-' in cm:
            parts = cm.split('-')
            cm = int(parts[0]) * 100 + int(parts[1])
        elif cm is not None:
            cm = int(cm)
        if cm is None and e.date_facture:
            cm = e.date_facture.year * 100 + e.date_facture.month
        if cm is None:
            continue
        # Filtrer uniquement l'année en cours
        if cm // 100 != annee:
            continue

        if cm not in rev:
            rev[cm] = {'freelance': 0.0, 'salaire': 0.0}

        if not e.is_paid:
            continue

        if e.tps_flag == 'Oui':
            rev[cm]['freelance'] += e.st
        else:
            rev[cm]['salaire'] += e.st

    result = []
    for cm in sorted(rev):
        result.append({
            'mois':      cm,
            'freelance': round(rev[cm]['freelance'], 2),
            'salaire':   round(rev[cm]['salaire'], 2),
        })
    return result


# ─── _build_expenses_by_category (STEP 1.2) ─────────────────────────────────

_MOIS_ORDER = ['JANVIER', 'FÉVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN',
               'JUILLET', 'AOÛT', 'SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DÉCEMBRE']


def _build_expenses_by_category(wb, annee: int) -> dict[str, float]:
    """Agrège les dépenses par catégorie depuis les feuilles mensuelles.

    Retourne {catégorie: montant_total}.
    Logique miroir de generate_dashboard.py::read_depenses() (section catégories).
    """
    categories: dict[str, float] = {}
    for mois in _MOIS_ORDER:
        sheet_name = f'{mois} {annee}'
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            cat = row[0]
            prix = row[4] if len(row) > 4 else None
            if cat and prix is not None and isinstance(prix, (int, float)):
                categories[str(cat)] = categories.get(str(cat), 0.0) + prix
    # Arrondir
    return {k: round(v, 2) for k, v in categories.items()}


# ─── _build_expenses_by_month (DASHBOARD V2) ────────────────────────────────

def _build_expenses_by_month(wb, annee: int) -> dict[int, float]:
    """Agrège les dépenses totales par mois depuis les feuilles mensuelles.

    Retourne {1: montant_jan, 2: montant_fev, ...}.
    Même source que _build_expenses_by_category() mais agrégé par mois.
    """
    result: dict[int, float] = {}
    for idx, mois in enumerate(_MOIS_ORDER, 1):
        sheet_name = f'{mois} {annee}'
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        total = 0.0
        for row in ws.iter_rows(min_row=2, values_only=True):
            prix = row[4] if len(row) > 4 else None
            if prix is not None and isinstance(prix, (int, float)):
                total += prix
        if total != 0:
            result[idx] = round(total, 2)
    return result


# ─── _read_taxes_current_month (DASHBOARD V2) ────────────────────────────────

def _compute_taxes_from_raw(wb, mois_name: str) -> Optional[float]:
    """Calcule le total taxes depuis les cellules brutes (non-formule).

    Fallback quand les cached values des formules sont détruites.
    Recompose : (TPS_collectée - CTI) + (TVQ_collectée - RTI)
    en cherchant les lignes par leur code formulaire (103, 104, 107,
    203, 204, 207) dans la colonne A. Résiste aux décalages de lignes
    causés par l'insertion de la nav bar.
    """
    taxes_sheet = f'TAXES {mois_name}'
    if taxes_sheet not in wb.sheetnames:
        return None

    ws = wb[taxes_sheet]

    # Construire un index {code_formulaire: row} en scannant col A
    code_row: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is not None:
            sa = str(a).strip()
            if sa in ('103', '104', '106', '107', '203', '204', '206', '207'):
                code_row[sa] = r

    def _raw(row: int) -> float:
        v = ws.cell(row=row, column=3).value
        if v is not None and isinstance(v, (int, float)):
            return float(v)
        return 0.0

    tps_collectee = _raw(code_row['103']) if '103' in code_row else 0.0
    tps_redress   = _raw(code_row['104']) if '104' in code_row else 0.0
    cti_direct    = _raw(code_row['106']) if '106' in code_row else 0.0
    cti_redress   = _raw(code_row['107']) if '107' in code_row else 0.0
    tvq_collectee = _raw(code_row['203']) if '203' in code_row else 0.0
    tvq_redress   = _raw(code_row['204']) if '204' in code_row else 0.0
    rti_direct    = _raw(code_row['206']) if '206' in code_row else 0.0
    rti_redress   = _raw(code_row['207']) if '207' in code_row else 0.0

    if tps_collectee == 0 and tvq_collectee == 0:
        return None  # Pas de données → mois sans activité

    # CTI et RTI : préférer codes 106/206 de TAXES sheet si disponibles
    if cti_direct > 0 or rti_direct > 0:
        cti = cti_direct
        rti = rti_direct
        # TPS nette = (TPS_collectée + redress) - (CTI + redress_CTI)
        tps_net = (tps_collectee + tps_redress) - (cti + cti_redress)
        # TVQ nette = (TVQ_collectée + redress) - (RTI + redress_RTI)
        tvq_net = (tvq_collectee + tvq_redress) - (rti + rti_redress)
        total = tps_net + tvq_net
        return max(0.0, round(total, 2))

    # Fallback : CTI/RTI depuis la feuille mensuelle des dépenses
    # CTI = INDEX('MOIS AAAA'!F:F, MATCH("À RÉCLAMER", 'MOIS AAAA'!D:D, 0))
    # RTI = INDEX('MOIS AAAA'!G:G, MATCH("À RÉCLAMER", 'MOIS AAAA'!D:D, 0))
    annee = None
    for sn in wb.sheetnames:
        if sn.startswith(mois_name + ' '):
            annee = sn.split()[-1]
            break

    cti = 0.0
    rti = 0.0
    mois_sheet = f'{mois_name} {annee}' if annee else None
    if mois_sheet and mois_sheet in wb.sheetnames:
        ws_mois = wb[mois_sheet]
        # Chercher "À RÉCLAMER" dans colonne D
        for r in range(1, ws_mois.max_row + 1):
            d = ws_mois.cell(r, 4).value
            if isinstance(d, str) and 'RÉCLAMER' in d.upper():
                f_val = ws_mois.cell(r, 6).value  # CTI
                g_val = ws_mois.cell(r, 7).value  # RTI
                # Ces cellules sont des formules → cached value peut être None
                # Si None, calculer manuellement depuis les ajustements
                if f_val is not None and isinstance(f_val, (int, float)):
                    cti = float(f_val)
                if g_val is not None and isinstance(g_val, (int, float)):
                    rti = float(g_val)
                break

    if cti == 0 and rti == 0 and mois_sheet and mois_sheet in wb.sheetnames:
        # Les cached values CTI/RTI sont aussi détruites → recalculer depuis E, K, L
        # Formule Excel par ligne : F{r} = IF(K{r}=TRUE(), 0, ROUND(E{r}*5%, 2))
        #                           G{r} = IF(L{r}=TRUE(), 0, ROUND(E{r}*9.975%, 2))
        ws_mois = wb[mois_sheet]
        sum_f, sum_g = 0.0, 0.0

        # Phase 1 : dépenses individuelles (lignes avant TOTAL)
        for r in range(2, ws_mois.max_row + 1):
            d_val = ws_mois.cell(r, 4).value
            if isinstance(d_val, str) and d_val.strip().upper() in ('TOTAL', 'À RÉCLAMER'):
                break

            # D'abord essayer les cached values F/G (au cas où certaines existent)
            f_v = ws_mois.cell(r, 6).value
            g_v = ws_mois.cell(r, 7).value
            f_ok = f_v is not None and isinstance(f_v, (int, float))
            g_ok = g_v is not None and isinstance(g_v, (int, float))

            if f_ok and g_ok:
                sum_f += float(f_v)
                sum_g += float(g_v)
            else:
                # Recalculer depuis E, I, J
                # K = NOT(I), L = NOT(J) — K/L sont des formules détruites
                # Lire I/J directement (valeurs brutes boolean)
                # I=TRUE → dépense déductible TPS (K=FALSE → CTI s'applique)
                # I=FALSE → exempt TPS (K=TRUE → CTI = 0)
                e_v = ws_mois.cell(r, 5).value
                if e_v is not None and isinstance(e_v, (int, float)):
                    i_v = ws_mois.cell(r, 9).value   # déductible TPS?
                    j_v = ws_mois.cell(r, 10).value  # déductible TVQ?
                    # I=TRUE ou None → déductible (CTI)
                    i_deductible = (i_v is not False)
                    j_deductible = (j_v is not False)
                    if i_deductible:
                        sum_f += round(float(e_v) * 0.05, 2)
                    if j_deductible:
                        sum_g += round(float(e_v) * 0.09975, 2)

        # Phase 2 : ajustements (lignes entre TOTAL et À RÉCLAMER)
        in_adjustments = False
        for r in range(2, ws_mois.max_row + 1):
            d_val = ws_mois.cell(r, 4).value
            if isinstance(d_val, str) and d_val.strip().upper() == 'TOTAL':
                in_adjustments = True
                continue
            if isinstance(d_val, str) and 'RÉCLAMER' in d_val.upper():
                break
            if in_adjustments:
                f_v = ws_mois.cell(r, 6).value
                g_v = ws_mois.cell(r, 7).value
                if f_v is not None and isinstance(f_v, (int, float)):
                    sum_f += float(f_v)
                if g_v is not None and isinstance(g_v, (int, float)):
                    sum_g += float(g_v)

        cti = sum_f
        rti = sum_g

    # TPS nette = (TPS_collectée + redress) - (CTI + redress_CTI)
    tps_net = (tps_collectee + tps_redress) - (cti + cti_redress)
    # TVQ nette = (TVQ_collectée + redress) - (RTI + redress_RTI)
    tvq_net = (tvq_collectee + tvq_redress) - (rti + rti_redress)
    total = tps_net + tvq_net

    return max(0.0, round(total, 2))


def _read_taxes_current_month(wb, today: Optional[date] = None) -> tuple[float, str]:
    """Lit le total taxes à remettre pour le mois courant (TAXES {MOIS}!C23).

    Retourne (montant, nom_mois_fr).  MAX(0, C23) pour cohérence avec H10.
    Si le mois courant n'a pas de données, remonte au dernier mois non-vide.

    Stratégie par mois (du plus récent au plus ancien) :
      1. Essayer C23 (cached value)
      2. Si C23 vide → fallback calcul brut depuis codes 103/106/203/206
      3. Si aucun résultat → passer au mois précédent
    """
    today = today or date.today()

    for m in range(today.month, 0, -1):
        mois_name = _MOIS_ORDER[m - 1]
        sheet_name = f'TAXES {mois_name}'
        if sheet_name not in wb.sheetnames:
            continue

        # Essayer C23 d'abord (cached formula value)
        ws = wb[sheet_name]
        c23 = ws.cell(row=23, column=3).value
        if c23 is not None and isinstance(c23, (int, float)):
            return max(0.0, round(float(c23), 2)), mois_name.capitalize()

        # Fallback : calculer depuis les cellules brutes
        computed = _compute_taxes_from_raw(wb, mois_name)
        if computed is not None:
            return computed, mois_name.capitalize()

    return 0.0, _MOIS_ORDER[today.month - 1].capitalize()


# ─── _build_ar_par_mois (à recevoir par mois) ────────────────────────────────

def _build_ar_par_mois(entries, annee: int) -> dict[int, float]:
    """Construit {mois_int: montant_a_recevoir} depuis les VenteEntry non payées.

    Retourne ex: {202603: 7150.0} pour mars 2026.
    """
    result: dict[int, float] = {}
    for e in entries:
        if e.is_doublon or e.is_paid:
            continue
        cm = e.cle_mois
        if isinstance(cm, str) and '-' in cm:
            parts = cm.split('-')
            cm = int(parts[0]) * 100 + int(parts[1])
        elif cm is not None:
            cm = int(cm)
        if cm is None and e.date_facture:
            cm = e.date_facture.year * 100 + e.date_facture.month
        if cm is None:
            continue
        if cm // 100 != annee:
            continue
        result[cm] = result.get(cm, 0.0) + (e.st or 0.0)
    return {k: round(v, 2) for k, v in result.items()}


# ─── _read_fiscal_2025_summary (STEP 1.2) ───────────────────────────────────

def _read_fiscal_2025_summary(depenses_file: str) -> dict:
    """Lit le résumé fiscal 2025 depuis le cache permanent .fiscal_2025_cache.json.

    Le cache est écrit par generate_dashboard.py::read_depenses_2025() ou
    init_fiscal_cache_2025.py. 2025 étant terminée, les chiffres ne changent plus.

    Retourne un dict avec au minimum :
      {'annee', 'rev_freelance', 'rev_salarie', 'depenses', 'total_charges',
       'deja_paye', 'solde_du', 'reer_paie', 'celiapp'}
    ou {} si le cache est absent.
    """
    cache_path = os.path.join(os.path.dirname(depenses_file), '.fiscal_2025_cache.json')
    # Aussi chercher dans _OUTILS/ (emplacement historique)
    if not os.path.isfile(cache_path):
        alt_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '.fiscal_2025_cache.json')
        if os.path.isfile(alt_path):
            cache_path = alt_path
    try:
        with open(cache_path) as f:
            data = _json.load(f)
        if data.get('annee') == 2025 and data.get('total_charges', 0) > 0:
            return data
    except (FileNotFoundError, _json.JSONDecodeError, KeyError):
        pass
    return {}


# ─── CockpitContext ───────────────────────────────────────────────────────────

@dataclass
class FiscalForecast:
    """Prévision fiscale annualisée (Pipeline V3 — Phase 5)."""
    projected_income:        float = 0.0   # Revenu projeté sur 12 mois
    projected_tax:           float = 0.0   # Impôt projeté total (fédéral + QC + cotisations)
    remaining_installments:  float = 0.0   # Acomptes restants à verser
    fiscal_risk:             str   = ''    # 'low' | 'medium' | 'high'
    months_elapsed:          int   = 0     # Mois écoulés pour la projection
    note:                    str   = ''


@dataclass
class CockpitContext:
    """Contexte complet du cockpit de pilotage financier.

    Attributes
    ----------
    alert_ctx    : AlertContext complet (AR, métriques, fiscal, acomptes, anomalies)
    alerts       : Liste d'alertes triée HIGH → MEDIUM → LOW
    score_label  : 'OK' | 'Attention' | 'Risque'
    score_emoji  : '🟢' | '🟡' | '🔴'
    reference_date : Date de référence utilisée pour les calculs
    obligations  : Liste des obligations actives (payé/total/%) depuis DÉPENSES.xlsm
    error        : Message d'erreur si le build a échoué partiellement ('' = succès)

    Pipeline V3 additions:
    current_treasury     : Trésorerie courante (revenus - dépenses - paiements)
    treasury_after_tax   : Trésorerie après obligations restantes
    fiscal_forecast      : Prévision fiscale annualisée
    """
    alert_ctx:             'AlertContext'
    alerts:                list['Alert']    = field(default_factory=list)
    score_label:           str              = 'OK'
    score_emoji:           str              = '🟢'
    reference_date:        Optional[date]  = None
    obligations:           list[dict]       = field(default_factory=list)
    monthly_revenue:       list[dict]       = field(default_factory=list)
    expenses_by_category:  dict             = field(default_factory=dict)
    expenses_by_month:     dict             = field(default_factory=dict)
    taxes_current_month:   float            = 0.0
    taxes_current_label:   str              = ''
    ar_par_mois:           dict             = field(default_factory=dict)
    fiscal_2025_summary:   dict             = field(default_factory=dict)
    ytd_2025:              float            = 0.0
    # Pipeline V3 — Phase 4 : Trésorerie
    current_treasury:      float            = 0.0
    treasury_after_tax:    float            = 0.0
    # Pipeline V3 — Phase 5 : Prévision fiscale
    fiscal_forecast:       Optional[FiscalForecast] = None
    error:                 str              = ''


# ─── build_cockpit_context ────────────────────────────────────────────────────

def build_cockpit_context(
    ventes_path:   str,
    depenses_path: str,
    today:         Optional[date] = None,
) -> CockpitContext:
    """Construit un CockpitContext complet à partir des chemins bruts.

    Paramètres
    ----------
    ventes_path   : Chemin vers le dossier VENTES/ (contient VENTES_AAAA.xlsx)
    depenses_path : Chemin vers le dossier DÉPENSES/ (contient DÉPENSES_AAAA.xlsm)
    today         : Date de référence (défaut : aujourd'hui)

    Returns
    -------
    CockpitContext — jamais de raise : les erreurs sont loguées dans .error.
    """
    from analytics.alerts import build_context, generate_alerts, compute_score
    from ar.calculator import load_journal_entries

    today = today or date.today()
    error = ''

    # Objet paths minimal compatible avec build_context()
    class _Paths:
        def __init__(self, v, d):
            self.ventes   = v
            self.depenses = d

    # Résoudre depenses_path : si c'est un répertoire, trouver le .xlsm de l'année courante.
    # Sans cette résolution, detect_tax_payments() reçoit un répertoire → os.path.isfile()
    # retourne False → (cti, rti, tps_payees, tvq_payees) = (0,0,0,0) → taxes gonflées.
    if os.path.isdir(depenses_path):
        xlsm_hits = sorted(
            f for f in _glob_mod.glob(os.path.join(depenses_path, '*PENSES*.xlsm'))
            if 'backup' not in os.path.basename(f).lower()
        )
        depenses_file = xlsm_hits[0] if xlsm_hits else depenses_path
    else:
        depenses_file = depenses_path

    paths = _Paths(ventes_path, depenses_file)

    # Déterminer l'année depuis le nom du fichier ou la date
    annee = today.year

    # Lire obligations + dépenses par catégorie + par mois depuis DÉPENSES.xlsm
    obligations: list[dict] = []
    expenses_by_category: dict = {}
    expenses_by_month: dict = {}
    taxes_current_month: float = 0.0
    taxes_current_label: str = ''
    try:
        import openpyxl as _openpyxl
        _wb_dep = _openpyxl.load_workbook(depenses_file, keep_vba=True, read_only=False, data_only=True)
        obligations = _read_obligations(_wb_dep)
        expenses_by_category = _build_expenses_by_category(_wb_dep, annee)
        expenses_by_month = _build_expenses_by_month(_wb_dep, annee)
        taxes_current_month, taxes_current_label = _read_taxes_current_month(_wb_dep, today)
        _wb_dep.close()
    except Exception:
        pass  # Pas critique — le cockpit reste fonctionnel sans ces données

    # Lire le résumé fiscal 2025
    fiscal_2025_summary = _read_fiscal_2025_summary(depenses_file)

    # Total versé pour l'AF 2025 = acomptes provisionnels + retenues à la source
    ytd_2025 = 0.0
    try:
        from fiscal.installments_ledger import ytd_total as _ytd_total
        ytd_2025 = _ytd_total(2025)
    except Exception:
        pass  # Pas critique

    # ── Montants exacts déjà payés pour AF2025 ──
    # Source : ESTIMATION FISCALE 2025 (DÉPENSES 2025.xlsm) + feuillets T4/RL-1
    #
    # FÉDÉRAL (ARC) :
    #   T4 case 22 retenues source : BRP-F1 1,379.61 + HOMMESENOR XII 215.87 = 1,595.48
    #   Trop-payé 2024 (row 52)    : 1,826.99
    #   Acomptes 2025 (ledger)     : 9,000.00 (déjà dans ytd via installments_ledger)
    #   Total ARC                  : 12,422.47
    #
    # PROVINCIAL (Revenu Québec) :
    #   RL-1 case E retenues source: BRP-F1 1,649.24 + HOMMESENOR XII 295.99 = 1,945.23
    #   Trop-payé 2024 (row 53)    : 3,212.96
    #   Total RQ                   : 5,158.19
    #
    _RETENUES_2025_CAN = 1595.48    # T4 case 22
    _RETENUES_2025_QC  = 1945.23    # RL-1 case E
    _TROPPAYE_2024_CAN = 1826.99    # ESTIMATION FISCALE row 52
    _TROPPAYE_2024_QC  = 3212.96    # ESTIMATION FISCALE row 53

    _deja_paye_map = {
        'Agence du revenu du Canada (ARC)': _RETENUES_2025_CAN + _TROPPAYE_2024_CAN,
        'Revenu Québec': _RETENUES_2025_QC + _TROPPAYE_2024_QC,
    }

    # Ajouter retenues + trop-payé 2024 au ytd_2025 (KPI)
    ytd_2025 += sum(_deja_paye_map.values())  # 3,540.71 + 5,039.95 = 8,580.66

    # Distribuer dans les obligations AF2025
    for o in obligations:
        _lbl = o.get('label', '')
        _dd = o.get('due_date')
        if (_lbl in _deja_paye_map
                and _dd is not None
                and hasattr(_dd, 'year') and _dd.year <= 2026):
            o['paid'] = o.get('paid', 0) + _deja_paye_map[_lbl]
            o['pct'] = (o['paid'] / o['total']) if o.get('total', 0) > 0 else 0.0

    # ── Fusionner COTIS_2025 (RRQ+FSS+RQAP) dans Revenu Québec AF2025 ──
    # Ces cotisations sont toutes québécoises → on les absorbe dans
    # l'obligation Revenu Québec pour AF2025 au lieu d'une ligne séparée.
    _cotis_2025 = [o for o in obligations
                   if 'RRQ' in o.get('label', '') or 'COTIS_2025' in str(o.get('source', ''))]
    _rq_2025 = [o for o in obligations
                if o.get('label', '') == 'Revenu Québec'
                and o.get('due_date') is not None
                and hasattr(o['due_date'], 'year') and o['due_date'].year <= 2026]
    if _cotis_2025 and _rq_2025:
        _rq = _rq_2025[0]
        for _co in _cotis_2025:
            _rq['total'] = _rq.get('total', 0) + _co.get('total', 0)
            _rq['paid']  = _rq.get('paid', 0)  + _co.get('paid', 0)
            obligations.remove(_co)
        _rq['pct'] = (_rq['paid'] / _rq['total']) if _rq['total'] > 0 else 0.0

    # Construire le contexte principal (alertes, AR, métriques)
    entries = []
    monthly_revenue: list[dict] = []
    try:
        # load_journal_entries attend le chemin vers VENTES_AAAA.xlsx
        # Si ventes_path est un répertoire VENTES/, remonter au parent et chercher le xlsx
        if os.path.isdir(ventes_path):
            hits = _glob_mod.glob(os.path.join(os.path.dirname(ventes_path), 'VENTES_*.xlsx'))
            xlsx_path = hits[0] if hits else ventes_path
        else:
            xlsx_path = ventes_path

        entries   = load_journal_entries(xlsx_path)
        alert_ctx = build_context(entries, paths, today)
        # Pipeline V3 — injecter les données V3 dans l'AlertContext pour les alertes enrichies
        alert_ctx.expenses_by_month = expenses_by_month
        alerts    = generate_alerts(alert_ctx)
        score_label, score_emoji = compute_score(alerts)
    except Exception as exc:
        from analytics.alerts import AlertContext
        alert_ctx    = AlertContext(reference_date=today)
        alerts       = []
        score_label  = 'OK'
        score_emoji  = '🟢'
        error        = str(exc)

    # Construire les revenus mensuels depuis les entries (STEP 1.2)
    ar_par_mois: dict = {}
    try:
        monthly_revenue = _build_monthly_revenue(entries, annee)
        ar_par_mois = _build_ar_par_mois(entries, annee)
    except Exception:
        pass  # Pas critique

    # ── Pipeline V3 — Phase 4 : Trésorerie ────────────────────────────────────
    current_treasury  = 0.0
    treasury_after_tax = 0.0
    try:
        # Revenus encaissés = somme des freelance + salaire dans monthly_revenue
        rev_enc = sum(m.get('freelance', 0) + m.get('salaire', 0) for m in monthly_revenue)
        # Dépenses totales = somme de expenses_by_month
        dep_total = sum(expenses_by_month.values())

        # Paiements effectués via cash_ledger (tous types, année courante)
        paiements_ledger = 0.0
        try:
            from fiscal.cash_ledger import ytd_payments as _cash_ytd
            paiements_ledger = _cash_ytd(annee)
        except ImportError:
            pass

        current_treasury = round(rev_enc - dep_total - paiements_ledger, 2)

        # Obligations restantes = somme des soldes d'obligations actives
        obl_restantes = sum(o.get('total', 0) - o.get('paid', 0)
                            for o in obligations
                            if (o.get('total', 0) - o.get('paid', 0)) > 0)
        treasury_after_tax = round(current_treasury - obl_restantes, 2)
    except Exception:
        pass

    # ── Pipeline V3 — Phase 5 : Prévision fiscale ──────────────────────────
    fiscal_forecast = None
    try:
        months_elapsed = today.month
        if months_elapsed >= 2:  # Besoin d'au moins 2 mois pour une projection fiable
            rev_ytd = sum(m.get('freelance', 0) + m.get('salaire', 0) for m in monthly_revenue)
            projected_income = round((rev_ytd / months_elapsed) * 12, 2)

            # Estimer l'impôt projeté
            try:
                from fiscal.estimator import (
                    estimate_income_tax_federal,
                    estimate_income_tax_qc,
                    estimate_autonomous_contributions,
                )
                proj_fed   = estimate_income_tax_federal(projected_income)
                proj_qc    = estimate_income_tax_qc(projected_income)
                proj_cotis = estimate_autonomous_contributions(projected_income)
                projected_tax = round(proj_fed + proj_qc + proj_cotis, 2)
            except ImportError:
                projected_tax = 0.0

            # Acomptes déjà versés pour l'année courante
            installments_paid = 0.0
            try:
                from fiscal.cash_ledger import ytd_payments as _cash_ytd2
                installments_paid = _cash_ytd2(annee, payment_type='acompte_federal') + \
                                    _cash_ytd2(annee, payment_type='acompte_provincial')
            except ImportError:
                pass

            remaining = max(0.0, round(projected_tax - installments_paid, 2))

            # Évaluer le risque
            if projected_tax > 0:
                coverage = installments_paid / projected_tax if projected_tax > 0 else 1.0
                if coverage >= 0.6:
                    risk = 'low'
                elif coverage >= 0.3:
                    risk = 'medium'
                else:
                    risk = 'high'
            else:
                risk = 'low'

            note = ''
            if months_elapsed < 4:
                note = 'Projection préliminaire (< 4 mois de données)'

            fiscal_forecast = FiscalForecast(
                projected_income       = projected_income,
                projected_tax          = projected_tax,
                remaining_installments = remaining,
                fiscal_risk            = risk,
                months_elapsed         = months_elapsed,
                note                   = note,
            )
    except Exception:
        pass

    # ── Post-traitement alertes : taxes mensuelles au lieu du cumul annuel ──
    # L'usager veut voir le montant à remettre en fin de mois courant, pas
    # le cumul annuel. On réécrit les alertes TAXES_ELEVEES / TAXES_MODEREES
    # pour afficher taxes_current_month quand disponible.
    if taxes_current_month > 0 and taxes_current_label:
        _mois = taxes_current_label.lower()
        for _a in alerts:
            if _a.type in ('TAXES_ELEVEES', 'TAXES_MODEREES'):
                _a.message = (
                    f'Taxes à remettre ({_mois}) : '
                    f'{taxes_current_month:,.2f}\u202f$ (TPS + TVQ)')
                _a.value = taxes_current_month

    return CockpitContext(
        alert_ctx             = alert_ctx,
        alerts                = alerts,
        score_label           = score_label,
        score_emoji           = score_emoji,
        reference_date        = today,
        obligations           = obligations,
        monthly_revenue       = monthly_revenue,
        ar_par_mois           = ar_par_mois,
        expenses_by_category  = expenses_by_category,
        expenses_by_month     = expenses_by_month,
        taxes_current_month   = taxes_current_month,
        taxes_current_label   = taxes_current_label,
        fiscal_2025_summary   = fiscal_2025_summary,
        ytd_2025              = ytd_2025,
        current_treasury      = current_treasury,
        treasury_after_tax    = treasury_after_tax,
        fiscal_forecast       = fiscal_forecast,
        error                 = error,
    )
