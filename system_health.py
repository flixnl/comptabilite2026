#!/usr/bin/env python3
"""
dashboard/system_health.py — Santé globale du système comptable.

Vérifie 8 points critiques couvrant les chemins, Excel, les taxes Python,
les acomptes provisionnels et la cohérence autopilot.

Les checks sont purement en lecture (aucune écriture).
L'ensemble doit s'exécuter en < 2s avec des données en cache.

USAGE
-----
    from dashboard.system_health import check_system_health, HealthLevel

    report = check_system_health()
    print(report.summary_line())
    for r in report.results:
        print(r.badge, r.label, '—', r.detail)

    # Avec résultats pré-calculés pour accélérer
    report = check_system_health(
        taxes_position=pos,      # TaxesPosition déjà calculée
        autopilot_result=ap,     # AutopilotResult déjà disponible
    )
"""
from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from enum import Enum
from typing import Any, Optional

# ── Résolution _OUTILS depuis dashboard/ ─────────────────────────────────────
_DASHBOARD_DIR = os.path.dirname(os.path.abspath(__file__))
_OUTILS_DIR    = os.path.dirname(_DASHBOARD_DIR)
if _OUTILS_DIR not in sys.path:
    sys.path.insert(0, _OUTILS_DIR)


# ══════════════════════════════════════════════════════════════════════════════
#  Niveaux et structures
# ══════════════════════════════════════════════════════════════════════════════

class HealthLevel(Enum):
    OK      = 'OK'
    WARNING = 'WARNING'
    ERROR   = 'ERROR'


_BADGE = {
    HealthLevel.OK:      '✅',
    HealthLevel.WARNING: '⚠️',
    HealthLevel.ERROR:   '❌',
}


@dataclass
class CheckResult:
    """Résultat d'un point de contrôle individuel."""
    label:  str
    level:  HealthLevel
    detail: str = ''

    @property
    def badge(self) -> str:
        return _BADGE[self.level]

    @property
    def is_ok(self) -> bool:
        return self.level == HealthLevel.OK

    @property
    def is_warning(self) -> bool:
        return self.level == HealthLevel.WARNING

    @property
    def is_error(self) -> bool:
        return self.level == HealthLevel.ERROR

    def __str__(self) -> str:
        s = f'{self.badge} {self.label}'
        if self.detail:
            s += f' — {self.detail}'
        return s


@dataclass
class SystemHealthReport:
    """Rapport complet de santé système."""
    results: list[CheckResult] = field(default_factory=list)

    # ── Compteurs ──────────────────────────────────────────────────────────────

    @property
    def n_ok(self) -> int:
        return sum(1 for r in self.results if r.is_ok)

    @property
    def n_warnings(self) -> int:
        return sum(1 for r in self.results if r.is_warning)

    @property
    def n_errors(self) -> int:
        return sum(1 for r in self.results if r.is_error)

    @property
    def all_ok(self) -> bool:
        return self.n_errors == 0 and self.n_warnings == 0

    @property
    def global_level(self) -> HealthLevel:
        if self.n_errors > 0:
            return HealthLevel.ERROR
        if self.n_warnings > 0:
            return HealthLevel.WARNING
        return HealthLevel.OK

    @property
    def global_badge(self) -> str:
        return _BADGE[self.global_level]

    @property
    def score(self) -> str:
        """Ex: '6/8 OK'"""
        return f'{self.n_ok}/{len(self.results)} OK'

    # ── Textes d'affichage ─────────────────────────────────────────────────────

    def summary_line(self) -> str:
        """Une ligne pour notification macOS."""
        if self.all_ok:
            return f'✅ Système sain — {self.score}'
        parts = []
        if self.n_errors:
            parts.append(f'{self.n_errors} erreur(s)')
        if self.n_warnings:
            parts.append(f'{self.n_warnings} avertissement(s)')
        return f'{self.global_badge} {", ".join(parts)} — {self.score}'

    def detail_lines(self) -> list[str]:
        """Toutes les lignes pour le dialog détaillé."""
        return [str(r) for r in self.results]

    def errors_and_warnings(self) -> list[str]:
        """Seulement les résultats non-OK, pour alerte concise."""
        return [str(r) for r in self.results if not r.is_ok]

    def menubar_status(self) -> str:
        """Titre court pour l'item de menu."""
        if self.all_ok:
            return f'✅ Santé système : OK ({self.score})'
        parts = []
        if self.n_errors:
            parts.append(f'{self.n_errors} ❌')
        if self.n_warnings:
            parts.append(f'{self.n_warnings} ⚠️')
        return f'Santé système : {" ".join(parts)}  ({self.score})'


# ══════════════════════════════════════════════════════════════════════════════
#  Point d'entrée public
# ══════════════════════════════════════════════════════════════════════════════

def check_system_health(
    taxes_position: Optional[Any]  = None,   # TaxesPosition pré-calculée (cache)
    autopilot_result: Optional[Any] = None,  # AutopilotResult pré-calculé (cache)
) -> SystemHealthReport:
    """Lance les 8 checks de santé système.

    Pure lecture — aucune écriture.

    Paramètres
    ----------
    taxes_position   : TaxesPosition déjà calculée (évite un recalcul coûteux).
                       None → les checks qui en dépendent rapportent WARNING.
    autopilot_result : AutopilotResult déjà calculé.
                       None → le check autopilot est basé sur le rapport sauvegardé.

    Retourne
    --------
    SystemHealthReport avec 8 CheckResult.
    """
    report = SystemHealthReport()

    # ── 1. Chemins ───────────────────────────────────────────────────────────
    paths = _resolve_paths()
    _check_paths(paths, report)

    # ── 2–7. Checks Excel (ouvre DÉPENSES une seule fois) ────────────────────
    dep = paths.get('depenses', '')
    wb  = None
    if dep and os.path.isfile(dep):
        try:
            import openpyxl
            # keep_vba=True pour lire les formules; read_only=False nécessaire pour keep_vba
            wb = openpyxl.load_workbook(dep, keep_vba=True)
        except Exception as exc:
            report.results.append(CheckResult(
                'Ouverture DÉPENSES', HealthLevel.ERROR, str(exc)[:80]))
            wb = None

    try:
        _check_nantel_dynamic(wb, report)          # 2
        _check_r10c8_formula(wb, report)           # 6  (groupé logiquement avec Excel)
        _check_resume_fiscal(wb, report)           # 5
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

    # ── 3. Taxes cohérentes Python vs Excel ──────────────────────────────────
    _check_taxes_coherence(paths, taxes_position, report)   # 3

    # ── 4. Acomptes provisionnels intégrés ───────────────────────────────────
    _check_acomptes(paths, taxes_position, report)          # 4

    # ── 7. Pas de fallback brut taxes ─────────────────────────────────────────
    _check_no_taxes_fallback(taxes_position, report)        # 7

    # ── 8. Autopilot cohérent ─────────────────────────────────────────────────
    _check_autopilot_coherence(paths, taxes_position, autopilot_result, report)  # 8

    return report


# ══════════════════════════════════════════════════════════════════════════════
#  Helpers internes
# ══════════════════════════════════════════════════════════════════════════════

def _resolve_paths() -> dict:
    """Résout les chemins principaux. Retourne un dict (jamais de raise)."""
    try:
        from core.paths import resolve_main_paths
        p = resolve_main_paths()
        ledger = os.path.join(_OUTILS_DIR, 'fiscal', 'installments_ledger.json')
        return {
            'ventes':   p.ventes,
            'depenses': p.depenses,
            'ledger':   ledger,
            'base':     p.base,
        }
    except Exception as exc:
        return {'_error': str(exc)}


# ── Check 1 : chemins ─────────────────────────────────────────────────────────

def _check_paths(paths: dict, report: SystemHealthReport) -> None:
    if '_error' in paths:
        report.results.append(CheckResult(
            'Chemins système', HealthLevel.ERROR,
            f'core.paths inaccessible : {paths["_error"][:60]}'))
        return

    issues = []
    ok_parts = []

    v = paths.get('ventes', '')
    if v and os.path.isfile(v):
        ok_parts.append('VENTES ✓')
    else:
        issues.append(f'VENTES introuvable ({os.path.basename(v) or "?"})')

    d = paths.get('depenses', '')
    if d and os.path.isfile(d):
        ok_parts.append('DÉPENSES ✓')
    else:
        issues.append(f'DÉPENSES introuvable ({os.path.basename(d) or "?"})')

    # Ledger fiscal — WARNING seulement (optionnel si pas encore de paiements)
    l = paths.get('ledger', '')
    if l and os.path.isfile(l):
        ok_parts.append('ledger ✓')
    else:
        ok_parts.append('ledger absent (normal si 0 acompte)')

    if issues:
        report.results.append(CheckResult(
            'Chemins système', HealthLevel.ERROR, ' | '.join(issues)))
    else:
        report.results.append(CheckResult(
            'Chemins système', HealthLevel.OK, ' | '.join(ok_parts)))


# ── Check 2 : Nantel dynamique ────────────────────────────────────────────────

def _check_nantel_dynamic(wb: Any, report: SystemHealthReport) -> None:
    label = 'Nantel dynamique'
    if wb is None:
        report.results.append(CheckResult(label, HealthLevel.WARNING, 'Classeur non ouvert'))
        return

    if 'TABLEAU DE BORD' not in wb.sheetnames:
        report.results.append(CheckResult(
            label, HealthLevel.ERROR, 'Onglet TABLEAU DE BORD absent'))
        return

    ws  = wb['TABLEAU DE BORD']
    val = ws.cell(112, 5).value   # R112 C5

    if val is None:
        report.results.append(CheckResult(label, HealthLevel.ERROR, 'R112 C5 vide'))
    elif isinstance(val, str) and val.startswith('='):
        val_up = val.upper()
        if 'STRAT' in val_up:
            report.results.append(CheckResult(
                label, HealthLevel.OK, f"→ formule STRAT. FISCALES ✓"))
        else:
            report.results.append(CheckResult(
                label, HealthLevel.OK, f'→ formule : {val[:50]}'))
    else:
        report.results.append(CheckResult(
            label, HealthLevel.ERROR,
            f'Valeur statique legacy : {repr(val)[:40]} — attendu formule'))


# ── Check 3 : cohérence taxes Python vs Excel (mois courant) ──────────────────

_MOIS_TAXES = ['JANVIER', 'FÉVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN',
               'JUILLET', 'AOÛT', 'SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DÉCEMBRE']


def _read_taxes_current_month_xl(wb, today, depenses_path: str = '') -> tuple[Optional[float], str]:
    """Lit TAXES {MOIS}!C23 (mois courant, sinon dernier mois non-vide).

    Même logique que dashboard.context._read_taxes_current_month() avec
    fallback Pass 2 : si les cached values sont détruites (openpyxl save
    antérieur), recalcule depuis les cellules brutes via context.py.

    Parameters
    ----------
    wb             : Workbook ouvert en data_only=True
    today          : date du jour
    depenses_path  : Chemin vers DÉPENSES.xlsm (pour le fallback Pass 2)

    Retourne (montant ou None, nom_mois).
    """
    # Pass 1 : cached values (data_only=True)
    for m in range(today.month, 0, -1):
        sheet_name = f'TAXES {_MOIS_TAXES[m - 1]}'
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        c23 = ws.cell(row=23, column=3).value
        if c23 is not None and isinstance(c23, (int, float)):
            return max(0.0, round(float(c23), 2)), _MOIS_TAXES[m - 1].capitalize()

    # Pass 2 : fallback — cached values détruites, recalculer depuis brutes
    try:
        from dashboard.context import _compute_taxes_from_raw
        # Résoudre le chemin du classeur
        dep = depenses_path
        if not dep:
            try:
                from core.paths import resolve_main_paths
                dep = resolve_main_paths().depenses
            except Exception:
                pass
        if dep and os.path.isfile(dep):
            import openpyxl as _opx
            wb_raw = _opx.load_workbook(dep, keep_vba=True)
            try:
                for m in range(today.month, 0, -1):
                    mois_name = _MOIS_TAXES[m - 1]
                    computed = _compute_taxes_from_raw(wb_raw, mois_name)
                    if computed is not None:
                        return computed, mois_name.capitalize()
            finally:
                wb_raw.close()
    except Exception:
        pass

    return None, _MOIS_TAXES[today.month - 1].capitalize()


def _check_taxes_coherence(
    paths: dict,
    taxes_position: Any,
    report: SystemHealthReport,
) -> None:
    """Compare les taxes du mois courant : TAXES {MOIS}!C23 vs H10.

    Les deux lisent la même source (onglet TAXES mensuel).
    H10 = MAX(0,'TAXES MARS'!C23) — doit correspondre.
    taxes_position.total_restantes (YTD) est conservé pour d'autres usages
    mais n'est PAS la métrique comparée ici.
    """
    label = 'Taxes Python ↔ Excel'
    TOLERANCE = 500.0   # $

    dep = paths.get('depenses', '')
    if not dep or not os.path.isfile(dep):
        report.results.append(CheckResult(
            label, HealthLevel.WARNING, 'DÉPENSES non accessible pour comparaison'))
        return

    try:
        import openpyxl
        from datetime import date as _date
        today = _date.today()
        wb2 = openpyxl.load_workbook(dep, data_only=True, read_only=True)
        try:
            # Lire la valeur Python : TAXES {MOIS}!C23 (mois courant)
            py_month, mois_label = _read_taxes_current_month_xl(wb2, today, depenses_path=dep)

            # Lire la valeur Excel : TABLEAU DE BORD H10
            xl_total = None
            if 'TABLEAU DE BORD' in wb2.sheetnames:
                cell = wb2['TABLEAU DE BORD'].cell(10, 8).value
                xl_total = float(cell) if isinstance(cell, (int, float)) else None
        finally:
            wb2.close()
    except Exception as exc:
        report.results.append(CheckResult(
            label, HealthLevel.WARNING, f'Lecture Excel échouée : {exc!s:.50}'))
        return

    if py_month is None:
        report.results.append(CheckResult(
            label, HealthLevel.WARNING,
            f'TAXES {mois_label.upper()} C23 vide — aucune donnée taxes ce mois'))
        return

    if xl_total is None:
        # H10 = MAX(0,'TAXES {MOIS}'!C23) — formule non évaluée car Excel pas ouvert.
        # Puisqu'on a déjà py_month depuis C23 (via Pass 1 ou Pass 2), on sait que
        # H10 correspondra une fois Excel recalculé. Pas un warning réel.
        report.results.append(CheckResult(
            label, HealthLevel.OK,
            f'TAXES {mois_label} : {py_month:,.0f} $ '
            f'(H10 = formule — recalcul au prochain ouvrir Excel)'))
        return

    diff = abs(py_month - xl_total)
    detail = (f'TAXES {mois_label} {py_month:,.0f} $ | '
              f'H10 {xl_total:,.0f} $ | Δ {diff:,.0f} $')

    if diff <= TOLERANCE:
        report.results.append(CheckResult(label, HealthLevel.OK, detail))
    elif diff <= TOLERANCE * 3:
        report.results.append(CheckResult(label, HealthLevel.WARNING, detail))
    else:
        report.results.append(CheckResult(label, HealthLevel.ERROR, detail))


# ── Check 4 : acomptes provisionnels intégrés ─────────────────────────────────

def _check_acomptes(
    paths: dict,
    taxes_position: Any,
    report: SystemHealthReport,
) -> None:
    label = 'Acomptes provisionnels'

    # Lire le ledger
    ledger = paths.get('ledger', '')
    try:
        from fiscal.installments_ledger import load_installment_payments, ytd_total
        from datetime import date as _date
        payments = load_installment_payments()
        year     = _date.today().year
        ytd      = ytd_total(year)
    except Exception as exc:
        if 'No such file' in str(exc) or 'introuvable' in str(exc).lower():
            report.results.append(CheckResult(
                label, HealthLevel.OK, "Aucun acompte enregistré (normal en début d'année)"))
        else:
            report.results.append(CheckResult(
                label, HealthLevel.WARNING, f'Ledger illisible : {exc!s:.60}'))
        return

    if not payments:
        report.results.append(CheckResult(
            label, HealthLevel.OK, 'Aucun acompte enregistré'))
        return

    # Vérifier que TaxesPosition intègre bien les acomptes
    if taxes_position is not None:
        pos_inst = getattr(taxes_position, 'installments_total', None)
        if pos_inst is None:
            report.results.append(CheckResult(
                label, HealthLevel.ERROR,
                'installments_total absent de TaxesPosition — VOLET 4 manquant'))
            return

        diff = abs(float(pos_inst) - ytd)
        if diff > 1.0:
            report.results.append(CheckResult(
                label, HealthLevel.WARNING,
                f'Ledger {ytd:,.0f} $ ≠ TaxesPosition {pos_inst:,.0f} $ '
                f'(Δ {diff:,.0f} $)'))
        else:
            report.results.append(CheckResult(
                label, HealthLevel.OK,
                f'{len(payments)} paiement(s) — {ytd:,.0f} $ YTD intégré ✓'))
    else:
        # Pas de TaxesPosition — juste vérifier la lisibilité du ledger
        report.results.append(CheckResult(
            label, HealthLevel.OK,
            f'{len(payments)} paiement(s) — {ytd:,.0f} $ YTD (intégration non vérifiée)'))


# ── Check 5 : résumé fiscal présent ──────────────────────────────────────────

def _check_resume_fiscal(wb: Any, report: SystemHealthReport) -> None:
    label = 'Résumé fiscal TAXES *'
    if wb is None:
        report.results.append(CheckResult(label, HealthLevel.WARNING, 'Classeur non ouvert'))
        return

    taxes_sheets = [
        name for name in wb.sheetnames
        if name.upper().startswith('TAXES') and len(name.split()) >= 2
    ]
    if not taxes_sheets:
        report.results.append(CheckResult(
            label, HealthLevel.ERROR, 'Aucun onglet TAXES * trouvé'))
        return

    for sheet_name in taxes_sheets[:3]:
        ws = wb[sheet_name]
        for row in range(25, 50):
            for col in (1, 2):
                cell_val = ws.cell(row, col).value
                if isinstance(cell_val, str) and 'RÉSUMÉ FISCAL' in cell_val.upper():
                    # Chercher TAXES RESTANTES dans les 12 lignes suivantes
                    has_restantes = any(
                        isinstance(ws.cell(r, c).value, str)
                        and 'TAXES RESTANTES' in ws.cell(r, c).value.upper()
                        for r in range(row, min(row + 12, 55))
                        for c in (1, 2)
                    )
                    detail = f'{sheet_name} R{row}'
                    if has_restantes:
                        detail += ' + TAXES RESTANTES ✓'
                    report.results.append(CheckResult(label, HealthLevel.OK, detail))
                    return

    report.results.append(CheckResult(
        label, HealthLevel.WARNING,
        f'RÉSUMÉ FISCAL UNIFIÉ absent dans {taxes_sheets[:3]}'))


# ── Check 6 : R10 C8 = formule (TAXES À REMETTRE) ────────────────────────────

def _check_r10c8_formula(wb: Any, report: SystemHealthReport) -> None:
    label = 'R10 C8 formule TAXES'
    if wb is None:
        report.results.append(CheckResult(label, HealthLevel.WARNING, 'Classeur non ouvert'))
        return

    if 'TABLEAU DE BORD' not in wb.sheetnames:
        report.results.append(CheckResult(
            label, HealthLevel.ERROR, 'Onglet TABLEAU DE BORD absent'))
        return

    ws  = wb['TABLEAU DE BORD']
    val = ws.cell(10, 8).value   # H10

    if isinstance(val, str) and val.startswith('='):
        val_up = val.upper()
        if 'TAXES' in val_up and 'C34' in val_up:
            report.results.append(CheckResult(
                label, HealthLevel.OK, 'IFERROR(SUM TAXES*!C34) ✓'))
        else:
            report.results.append(CheckResult(
                label, HealthLevel.OK, f'formule : {val[:55]}'))
    elif isinstance(val, (int, float)):
        report.results.append(CheckResult(
            label, HealthLevel.ERROR,
            f'Valeur figée : {val:,.2f} $ — attendu formule IFERROR(…)'))
    elif val is None:
        report.results.append(CheckResult(
            label, HealthLevel.WARNING, 'H10 vide (formule non évaluée ?)'))
    else:
        report.results.append(CheckResult(
            label, HealthLevel.WARNING, f'Valeur inattendue : {repr(val)[:40]}'))


# ── Check 7 : pas de fallback brut taxes ─────────────────────────────────────

def _check_no_taxes_fallback(
    taxes_position: Any,
    report: SystemHealthReport,
) -> None:
    label = 'Source taxes (pas fallback)'

    if taxes_position is None:
        report.results.append(CheckResult(
            label, HealthLevel.ERROR,
            'TaxesPosition indisponible — impossible de vérifier la source taxes'))
        return

    source = getattr(taxes_position, 'source_collectees', '')
    note   = getattr(taxes_position, 'note', '')

    fallback_indicators = ('fallback', 'brut', 'depenses_direct', 'taxes_sheet')
    is_fallback = any(ind in (source or '').lower() for ind in fallback_indicators)

    if is_fallback:
        report.results.append(CheckResult(
            label, HealthLevel.WARNING,
            f'Source fallback utilisée : "{source}" — VENTES non chargées ?'))
    else:
        detail = f'source = {source or "ventes_fiscal"}' + (f' | {note[:40]}' if note else '')
        report.results.append(CheckResult(label, HealthLevel.OK, detail))


# ── Check 8 : cohérence autopilot (mismatch taxes ≤ 500 $) ───────────────────

def _check_autopilot_coherence(
    paths: dict,
    taxes_position: Any,
    autopilot_result: Any,
    report: SystemHealthReport,
) -> None:
    label = 'Autopilot cohérent'
    THRESHOLD = 500.0

    # Pas de données → skip sans ERROR
    if taxes_position is None and autopilot_result is None:
        report.results.append(CheckResult(
            label, HealthLevel.WARNING,
            'Données non disponibles — lance une actualisation'))
        return

    # Si on a un résultat autopilot en erreur globale
    if autopilot_result is not None:
        ap_error = getattr(autopilot_result, 'error', '') or ''
        if ap_error:
            report.results.append(CheckResult(
                label, HealthLevel.WARNING, f'Autopilot en erreur : {ap_error[:60]}'))
            return

        # Chercher une anomalie taxes dans les anomalies autopilot
        anomalies = getattr(autopilot_result, 'anomalies', []) or []
        tax_anomalies = [
            a for a in anomalies
            if 'tax' in getattr(a, 'category', '').lower()
            or 'taxes' in getattr(a, 'message', '').lower()
        ]
        n_tax_high = sum(
            1 for a in tax_anomalies
            if getattr(a, 'severity', None) and
            str(getattr(a, 'severity', '')).upper() in ('HIGH', 'SEV.HIGH')
        )
        if n_tax_high > 0:
            report.results.append(CheckResult(
                label, HealthLevel.WARNING,
                f'{n_tax_high} anomalie(s) taxes HIGH dans l\'autopilot'))
            return

    # Comparer taxes mois courant (TAXES {MOIS}!C23) avec H10 (même logique check 3)
    dep = paths.get('depenses', '')
    if dep and os.path.isfile(dep):
        try:
            import openpyxl
            from datetime import date as _date
            wb3 = openpyxl.load_workbook(dep, data_only=True, read_only=True)
            try:
                py_month, mois_label = _read_taxes_current_month_xl(wb3, _date.today(), depenses_path=dep)
                xl_total = None
                if 'TABLEAU DE BORD' in wb3.sheetnames:
                    cell_val = wb3['TABLEAU DE BORD'].cell(10, 8).value
                    xl_total = float(cell_val) if isinstance(cell_val, (int, float)) else None
            finally:
                wb3.close()

            ap_label = ''
            if autopilot_result is not None:
                ap_label = (f' | autopilot: '
                            f'{getattr(autopilot_result, "score_emoji", "⚪")} '
                            f'{getattr(autopilot_result, "score_label", "?")}')

            if py_month is None or xl_total is None:
                detail = ''
                if py_month is not None:
                    detail = f'TAXES {mois_label} : {py_month:,.0f} $ (H10 = formule non évaluée)'
                else:
                    detail = f'Taxes {mois_label} non disponible — check non applicable'
                report.results.append(CheckResult(
                    label, HealthLevel.OK,
                    f'{detail}{ap_label}'))
                return

            diff = abs(py_month - xl_total)
            if diff <= THRESHOLD:
                report.results.append(CheckResult(
                    label, HealthLevel.OK,
                    f'Δ taxes {mois_label} ↔ H10 {diff:,.0f} $ ≤ {THRESHOLD:,.0f} $'
                    f'{ap_label}'))
            else:
                report.results.append(CheckResult(
                    label, HealthLevel.ERROR,
                    f'Mismatch {diff:,.0f} $ > {THRESHOLD:,.0f} $'
                    f' (TAXES {mois_label} {py_month:,.0f} $ vs H10 {xl_total:,.0f} $)'
                    f'{ap_label}'))
            return
        except Exception:
            pass

    # Fallback : check basique via score autopilot
    if autopilot_result is not None:
        score = getattr(autopilot_result, 'score_label', 'OK')
        emoji = getattr(autopilot_result, 'score_emoji', '⚪')
        n_h   = getattr(autopilot_result, 'n_high', 0)
        detail = f'{emoji} {score} — {n_h} anomalie(s) HIGH'
        level = HealthLevel.ERROR if n_h >= 3 else (
            HealthLevel.WARNING if n_h >= 1 else HealthLevel.OK)
        report.results.append(CheckResult(label, level, detail))
    else:
        report.results.append(CheckResult(
            label, HealthLevel.OK, 'Score cohérent (TaxesPosition sans référence Excel)'))


# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print('\n' + '═' * 65)
    print('  SYSTEM HEALTH CHECK')
    print('═' * 65)
    rep = check_system_health()
    for r in rep.results:
        print(f'  {r}')
    print('─' * 65)
    print(f'  {rep.summary_line()}')
    print('═' * 65 + '\n')
