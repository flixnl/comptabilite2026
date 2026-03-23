"""
dashboard/health_check.py — Routine de santé du classeur DÉPENSES 2026.xlsm.

Vérifie les points critiques post-migration :
  1. Chemin DÉPENSES résolu (via core.paths)
  2. Nantel dynamique  — OBLIGATIONS R112 C5 est une formule, pas une valeur figée
  3. R10 C8 formule active — TAXES À REMETTRE annuel = formule IFERROR(…TAXES*…C34…)
  4. Résumé fiscal présent — au moins 1 onglet TAXES * contient RÉSUMÉ FISCAL UNIFIÉ
  5. Cohérence Nantel — valeur lue depuis STRATÉGIES FISCALES C102 (nb paiements)

USAGE
-----
    from dashboard.health_check import run_excel_health_check

    report = run_excel_health_check()
    for line in report.summary_lines():
        print(line)
    print(f'Score : {report.score}  —  {"✅ OK" if report.all_ok else "⚠️ Problèmes détectés"}')
"""
from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from typing import List, Optional

# ── Résolution du dossier _OUTILS depuis core/ ────────────────────────────────
_DASHBOARD_DIR = os.path.dirname(os.path.abspath(__file__))
_OUTILS_DIR    = os.path.dirname(_DASHBOARD_DIR)
if _OUTILS_DIR not in sys.path:
    sys.path.insert(0, _OUTILS_DIR)


# ══════════════════════════════════════════════════════════════════════════════
#  Structures de résultat
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class CheckResult:
    """Résultat d'un seul point de contrôle."""
    label:  str         # ex: "Nantel dynamique"
    ok:     bool        # True → ✅,  False → ❌
    detail: str = ''    # info complémentaire ou message d'erreur


@dataclass
class HealthReport:
    """Rapport complet de santé du classeur."""
    checks:        List[CheckResult] = field(default_factory=list)
    depenses_path: str = ''   # chemin réel utilisé (vide si non résolu)

    # ── Propriétés synthèse ────────────────────────────────────────────────────

    @property
    def all_ok(self) -> bool:
        """True si tous les checks sont ✅."""
        return bool(self.checks) and all(c.ok for c in self.checks)

    @property
    def n_ok(self) -> int:
        return sum(1 for c in self.checks if c.ok)

    @property
    def n_fail(self) -> int:
        return sum(1 for c in self.checks if not c.ok)

    @property
    def score(self) -> str:
        """ex: '4/4' ou '3/4'"""
        return f'{self.n_ok}/{len(self.checks)}'

    @property
    def emoji(self) -> str:
        if self.n_fail == 0:
            return '✅'
        elif self.n_fail <= 1:
            return '⚠️'
        return '❌'

    def summary_lines(self) -> List[str]:
        """Liste de lignes pour affichage (notification / alert / log)."""
        lines = []
        for c in self.checks:
            prefix = '✅' if c.ok else '❌'
            line   = f'{prefix} {c.label}'
            if c.detail:
                line += f' — {c.detail}'
            lines.append(line)
        return lines

    def short_summary(self) -> str:
        """Une seule ligne concise pour notification macOS."""
        if self.all_ok:
            return f'✅ Tout OK ({self.score})'
        fails = [c.label for c in self.checks if not c.ok]
        return f'{self.emoji} {self.n_fail} problème(s) : {", ".join(fails[:2])}'


# ══════════════════════════════════════════════════════════════════════════════
#  Point d'entrée public
# ══════════════════════════════════════════════════════════════════════════════

def run_excel_health_check(depenses_path: Optional[str] = None) -> HealthReport:
    """Lance la routine de santé du classeur DÉPENSES 2026.xlsm.

    Paramètres
    ----------
    depenses_path : str, optional
        Chemin explicite vers DÉPENSES 2026.xlsm. Si absent, résolu via core.paths.

    Retourne
    --------
    HealthReport avec les résultats de 4 à 5 checks.
    """
    report = HealthReport()

    # ── CHECK 1 : résolution du chemin ────────────────────────────────────────
    depenses_path = _check_path(depenses_path, report)
    if depenses_path is None:
        # Impossible d'aller plus loin sans le fichier
        return report

    # ── Ouvrir le classeur (keep_vba=True — formules lisibles) ────────────────
    try:
        import openpyxl
        wb = openpyxl.load_workbook(depenses_path, keep_vba=True)
    except Exception as exc:
        report.checks.append(CheckResult(
            'Ouverture classeur', False, str(exc)[:120]))
        return report

    try:
        _check_nantel(wb, report)
        _check_r10c8(wb, report)
        _check_resume_fiscal(wb, report)
        _check_nantel_count(wb, report)
        _check_cash_ledger(wb, report)
    finally:
        wb.close()

    return report


# ══════════════════════════════════════════════════════════════════════════════
#  Checks individuels
# ══════════════════════════════════════════════════════════════════════════════

def _check_path(depenses_path: Optional[str], report: HealthReport) -> Optional[str]:
    """Check 1 — résout et valide le chemin du classeur. Retourne le chemin ou None."""
    try:
        if depenses_path:
            if os.path.isfile(depenses_path):
                report.depenses_path = depenses_path
                report.checks.append(CheckResult(
                    'Chemin DÉPENSES résolu', True,
                    os.path.basename(depenses_path)))
                return depenses_path
            else:
                report.checks.append(CheckResult(
                    'Chemin DÉPENSES résolu', False,
                    f'Fichier introuvable : {depenses_path}'))
                return None
        else:
            from core.paths import resolve_main_paths
            _paths   = resolve_main_paths()
            resolved = _paths.depenses
            if not resolved or not os.path.isfile(resolved):
                raise FileNotFoundError(
                    f'DÉPENSES introuvable via core.paths '
                    f'(base={_paths.base})')
            report.depenses_path = resolved
            report.checks.append(CheckResult(
                'Chemin DÉPENSES résolu', True,
                os.path.basename(resolved)))
            return resolved
    except FileNotFoundError as exc:
        report.checks.append(CheckResult(
            'Chemin DÉPENSES résolu', False, str(exc).split('\n')[0][:80]))
        return None
    except Exception as exc:
        report.checks.append(CheckResult(
            'Chemin DÉPENSES résolu', False, str(exc)[:80]))
        return None


def _check_nantel(wb, report: HealthReport) -> None:
    """Check 2 — TABLEAU DE BORD R112 C5 est une formule dynamique (pas statique)."""
    label = 'Nantel dynamique (OBLIGATIONS R112 C5)'
    if 'TABLEAU DE BORD' not in wb.sheetnames:
        report.checks.append(CheckResult(label, False, 'Onglet TABLEAU DE BORD absent'))
        return

    ws  = wb['TABLEAU DE BORD']
    val = ws.cell(112, 5).value

    if val is None:
        report.checks.append(CheckResult(label, False, 'R112 C5 est vide'))
    elif isinstance(val, str) and val.startswith('='):
        val_upper = val.upper()
        if 'STRAT' in val_upper and 'C101' in val_upper:
            report.checks.append(CheckResult(label, True,
                                             "→ ='STRATÉGIES FISCALES'!C101 ✓"))
        else:
            report.checks.append(CheckResult(label, True,
                                             f'formule : {val[:50]}'))
    else:
        report.checks.append(CheckResult(
            label, False,
            f'Valeur statique legacy : {val} — attendu formule =STRATÉGIES FISCALES!C101'))


def _check_r10c8(wb, report: HealthReport) -> None:
    """Check 3 — TABLEAU DE BORD R10 C8 est une formule TAXES À REMETTRE annuel."""
    label = 'R10 C8 formule TAXES active'
    if 'TABLEAU DE BORD' not in wb.sheetnames:
        report.checks.append(CheckResult(label, False, 'Onglet TABLEAU DE BORD absent'))
        return

    ws  = wb['TABLEAU DE BORD']
    val = ws.cell(10, 8).value   # ancre de la fusion H10:J10

    if isinstance(val, str) and val.startswith('='):
        val_upper = val.upper()
        if 'TAXES' in val_upper and 'C34' in val_upper:
            report.checks.append(CheckResult(label, True,
                                             'IFERROR(SUM TAXES*!C34,0) ✓'))
        else:
            report.checks.append(CheckResult(label, True,
                                             f'formule : {val[:60]}'))
    else:
        report.checks.append(CheckResult(
            label, False,
            f'Valeur figée : {repr(val)[:40]} — attendu =IFERROR(…TAXES*…C34…,0)'))


def _check_resume_fiscal(wb, report: HealthReport) -> None:
    """Check 4 — Au moins un onglet TAXES * contient RÉSUMÉ FISCAL UNIFIÉ."""
    label = 'Résumé fiscal TAXES * présent'

    taxes_sheets = [
        name for name in wb.sheetnames
        if name.upper().startswith('TAXES') and len(name.split()) >= 2
    ]
    if not taxes_sheets:
        report.checks.append(CheckResult(label, False, 'Aucun onglet TAXES * trouvé'))
        return

    for sheet_name in taxes_sheets[:3]:   # 3 premiers suffisent
        ws = wb[sheet_name]
        # Le RÉSUMÉ FISCAL UNIFIÉ est écrit en colonne A (col 1) par migrate_taxes_mars.py
        # Scan aussi la colonne B (col 2) pour compatibilité avec d'éventuels futurs formats
        for row in range(27, 45):
            for col in (1, 2):
                cell_val = ws.cell(row, col).value
                if isinstance(cell_val, str) and 'RÉSUMÉ FISCAL' in cell_val.upper():
                    # Vérifier TAXES RESTANTES dans les 8 lignes suivantes (col 1 ou 2)
                    has_restantes = any(
                        isinstance(ws.cell(r, c).value, str)
                        and 'TAXES RESTANTES' in ws.cell(r, c).value.upper()
                        for r in range(row, min(row + 10, 50))
                        for c in (1, 2)
                    )
                    detail = f'{sheet_name} R{row}'
                    if has_restantes:
                        detail += ' + TAXES RESTANTES ✓'
                    report.checks.append(CheckResult(label, True, detail))
                    return

    report.checks.append(CheckResult(
        label, False,
        f'RÉSUMÉ FISCAL UNIFIÉ absent dans les onglets scannés : {taxes_sheets[:3]}'))


def _check_nantel_count(wb, report: HealthReport) -> None:
    """Check 5 (info) — Lit le nb de paiements Nantel depuis STRATÉGIES FISCALES."""
    label = 'Nantel — nb paiements (info)'
    if 'STRATÉGIES FISCALES' not in wb.sheetnames:
        report.checks.append(CheckResult(label, False, 'Onglet STRATÉGIES FISCALES absent'))
        return

    ws  = wb['STRATÉGIES FISCALES']
    val = ws.cell(102, 3).value   # R102 C3 = Nombre de paiements effectués

    if isinstance(val, (int, float)) and val > 0:
        total_paye = int(val) * 600
        restants   = max(0, 60 - int(val))
        report.checks.append(CheckResult(
            label, True,
            f'{int(val)}/60 paiements effectués — {total_paye:,} $ payés — {restants} restants'))
    elif val is None:
        report.checks.append(CheckResult(label, False, 'R102 C3 est vide'))
    else:
        report.checks.append(CheckResult(label, False,
                                          f'Valeur inattendue : {repr(val)[:40]}'))


def _check_cash_ledger(wb, report: HealthReport) -> None:
    """Check 6 (Pipeline V3) — Cash ledger cohérent per-obligation avec OBLIGATIONS.

    Compare obligation par obligation (pas un total agrégé) car le cash_ledger
    peut ne couvrir qu'un sous-ensemble des obligations (ex: acomptes fédéraux
    uniquement). Un delta global serait trompeur.
    """
    label = 'Cash ledger V3'
    try:
        from fiscal.cash_ledger import load_ledger
        entries = load_ledger()
        if not entries:
            report.checks.append(CheckResult(label, True,
                                              'Cash ledger vide (aucune entrée)'))
            return

        active_entries = [e for e in entries if e.is_active]
        total_ledger = sum(e.amount for e in active_entries)

        # Regrouper le ledger par obligation_id
        ledger_by_obl: dict[str, float] = {}
        for e in active_entries:
            oid = e.obligation_id or '_sans_obligation'
            ledger_by_obl[oid] = ledger_by_obl.get(oid, 0.0) + e.amount

        # Lire OBLIGATIONS per-row pour comparaison ciblée
        if 'OBLIGATIONS' not in wb.sheetnames:
            report.checks.append(CheckResult(
                label, True,
                f'{len(active_entries)} entrée(s), {total_ledger:,.0f}$ (OBLIGATIONS absent)'))
            return

        ws_obl = wb['OBLIGATIONS']
        obl_rows: dict[str, float] = {}
        for r in range(2, 20):
            obl_id = ws_obl.cell(r, 1).value
            if not obl_id:
                break
            paid = ws_obl.cell(r, 11).value  # col K = TOTAL_PAYÉ
            if paid is not None and isinstance(paid, (int, float)):
                obl_rows[str(obl_id)] = float(paid)

        # Comparer UNIQUEMENT les obligations couvertes par le ledger
        covered_ids = set(ledger_by_obl.keys()) & set(obl_rows.keys())
        uncovered_ids = set(obl_rows.keys()) - set(ledger_by_obl.keys())

        if not covered_ids:
            # Ledger existe mais ne couvre aucune obligation du tableau
            uncov_total = sum(obl_rows.get(oid, 0) for oid in uncovered_ids)
            report.checks.append(CheckResult(
                label, True,
                f'{len(active_entries)} entrée(s) ledger ({total_ledger:,.0f}$) — '
                f'pas encore lié aux {len(obl_rows)} obligations ({uncov_total:,.0f}$)'))
            return

        # Per-obligation comparison
        mismatches = []
        for oid in sorted(covered_ids):
            ledger_val = ledger_by_obl[oid]
            obl_val = obl_rows[oid]
            delta = abs(ledger_val - obl_val)
            if delta > 100:
                mismatches.append(f'{oid}: ledger={ledger_val:,.0f}$ vs obl={obl_val:,.0f}$')

        if uncovered_ids:
            uncov_note = f' | {len(uncovered_ids)} obligation(s) hors ledger (normal)'
        else:
            uncov_note = ''

        if not mismatches:
            report.checks.append(CheckResult(
                label, True,
                f'{len(covered_ids)} obligation(s) vérifiée(s) OK — '
                f'ledger={total_ledger:,.0f}${uncov_note}'))
        else:
            report.checks.append(CheckResult(
                label, False,
                f'{len(mismatches)} divergence(s) : {"; ".join(mismatches[:3])}{uncov_note}'))

    except ImportError:
        report.checks.append(CheckResult(label, True,
                                          'Module cash_ledger non installé (V3 en attente)'))
    except Exception as exc:
        report.checks.append(CheckResult(label, False, str(exc)[:80]))


# ══════════════════════════════════════════════════════════════════════════════
#  Helpers pour l'action "Incrémenter Nantel"
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class NantelInfo:
    """Données lues depuis STRATÉGIES FISCALES pour l'affichage Nantel."""
    nb_paiements:  int   = 0
    mensualite:    float = 600.0
    nb_total:      int   = 60
    total_paye:    float = 0.0
    restants:      int   = 0
    solde_restant: float = 0.0


def read_nantel_info(depenses_path: str) -> NantelInfo:
    """Lit les informations Nantel depuis STRATÉGIES FISCALES.

    Retourne un NantelInfo avec nb_paiements = 0 en cas d'erreur.
    """
    info = NantelInfo()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(depenses_path, keep_vba=True)
        try:
            if 'STRATÉGIES FISCALES' in wb.sheetnames:
                ws = wb['STRATÉGIES FISCALES']
                nb   = ws.cell(102, 3).value   # R102 C3 = nb paiements effectués
                mens = ws.cell(103, 3).value   # R103 C3 = mensualité (600)
                tot  = ws.cell(104, 3).value   # R104 C3 = durée totale (60)

                if isinstance(nb, (int, float)) and nb > 0:
                    info.nb_paiements = int(nb)
                if isinstance(mens, (int, float)) and mens > 0:
                    info.mensualite = float(mens)
                if isinstance(tot, (int, float)) and tot > 0:
                    info.nb_total = int(tot)

                info.total_paye    = info.nb_paiements * info.mensualite
                info.restants      = max(0, info.nb_total - info.nb_paiements)
                info.solde_restant = info.restants * info.mensualite
        finally:
            wb.close()
    except Exception:
        pass
    return info


def increment_nantel_payment(depenses_path: str) -> tuple[bool, str, NantelInfo]:
    """Incrémente de +1 le nombre de paiements Nantel dans STRATÉGIES FISCALES.

    Paramètres
    ----------
    depenses_path : str
        Chemin résolu vers DÉPENSES 2026.xlsm.

    Retourne
    --------
    (success: bool, message: str, new_info: NantelInfo)
    """
    import shutil
    from datetime import datetime

    # ── Lecture de la valeur actuelle ────────────────────────────────────────
    try:
        import openpyxl
        wb = openpyxl.load_workbook(depenses_path, keep_vba=True)
    except Exception as exc:
        return False, f'Impossible d\'ouvrir le classeur : {exc}', NantelInfo()

    try:
        if 'STRATÉGIES FISCALES' not in wb.sheetnames:
            return False, 'Onglet STRATÉGIES FISCALES introuvable', NantelInfo()

        ws  = wb['STRATÉGIES FISCALES']
        val = ws.cell(102, 3).value

        if not isinstance(val, (int, float)):
            return (False,
                    f'R102 C3 invalide : {repr(val)[:40]} — attendu un entier',
                    NantelInfo())

        current = int(val)
        if current < 0 or current > 60:
            return (False,
                    f'Valeur hors plage : {current} (attendu 0‒60)',
                    NantelInfo())

        new_val = current + 1
    finally:
        wb.close()

    # ── Backup avant écriture ─────────────────────────────────────────────────
    bckp_dir = os.path.join(os.path.dirname(depenses_path), '_bckp')
    os.makedirs(bckp_dir, exist_ok=True)
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    bckp = os.path.join(bckp_dir,
                        f'DÉPENSES 2026.xlsm.pre_nantel_incr_{ts}')
    try:
        shutil.copy2(depenses_path, bckp)
    except Exception as exc:
        return False, f'Backup échoué : {exc}', NantelInfo()

    # ── Écriture ──────────────────────────────────────────────────────────────
    try:
        wb2 = openpyxl.load_workbook(depenses_path, keep_vba=True)
        wb2['STRATÉGIES FISCALES'].cell(102, 3).value = new_val
        # Forcer recalcul des formules (H10 TABLEAU DE BORD, etc.) à la prochaine ouverture
        try:
            wb2.calculation.fullCalcOnLoad = True
        except Exception:
            pass
        wb2.save(depenses_path)
        wb2.close()
    except Exception as exc:
        return False, f'Écriture échouée : {exc}', NantelInfo()

    # ── Relecture post-save ────────────────────────────────────────────────────
    new_info = read_nantel_info(depenses_path)
    if new_info.nb_paiements != new_val:
        return (False,
                f'Post-save mismatch : lu {new_info.nb_paiements}, attendu {new_val}',
                new_info)

    msg = (f'Nantel mis à jour : {new_val} paiements — '
           f'{int(new_info.total_paye):,} $ payés — '
           f'{new_info.restants} restants')
    return True, msg, new_info


# ══════════════════════════════════════════════════════════════════════════════
#  CLI de diagnostic (python3 -m dashboard.health_check)
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Santé du classeur DÉPENSES 2026.xlsm')
    parser.add_argument('--excel', metavar='CHEMIN',
                        help='Chemin explicite vers .xlsm')
    args = parser.parse_args()

    report = run_excel_health_check(depenses_path=args.excel)

    print()
    print('═' * 60)
    print('  EXCEL HEALTH CHECK — DÉPENSES 2026.xlsm')
    print('═' * 60)
    if report.depenses_path:
        print(f'  Fichier : {report.depenses_path}')
    for line in report.summary_lines():
        print(f'  {line}')
    print('─' * 60)
    print(f'  Score : {report.score}  — {report.emoji} '
          f'{"Tout OK" if report.all_ok else f"{report.n_fail} problème(s)"}')
    print('═' * 60)
    print()
