#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""dashboard/write_dashboard_v2.py
=========================================================
Écrit l'onglet DASHBOARD V2 dans DÉPENSES 2026.xlsm.

100 % Python-driven via CockpitContext — zéro formule Excel.
Design premium inspiré Notion / Stripe (light theme).

Usage
─────
    from dashboard.write_dashboard_v2 import write_dashboard_v2
    write_dashboard_v2(ctx, depenses_path)
"""
from __future__ import annotations

import os
import shutil
import tempfile
import zipfile
from datetime import datetime
from typing import TYPE_CHECKING, Optional
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from dashboard.context import CockpitContext

# ── Palette Notion / Stripe (light) ─────────────────────────────────────────
_WHITE    = 'FFFFFF'
_NEAR_BLK = '111827'
_G700     = '374151'
_G500     = '6B7280'
_G400     = '9CA3AF'
_G300     = 'D1D5DB'
_G200     = 'E5E7EB'
_G150     = 'F3F4F6'   # card background
_G100     = 'F9FAFB'
_GFAFA    = 'FAFAFA'   # section alt background

_RED50    = 'FEF2F2'
_RED800   = '991B1B'
_YEL50    = 'FFFBEB'
_YEL800   = '92400E'
_GRN50    = 'F0FDF4'
_GRN700   = '15803D'
_BLU50    = 'EFF6FF'
_BLU700   = '1D4ED8'

# ── Section background mapping (contraste visible) ──────────────────────────
_SEC_BG_KPI    = 'E8EDF3'  # bleu-gris léger — KPI + obligations
_SEC_BG_REV    = _WHITE    # blanc pur — revenus
_SEC_BG_DEP    = 'F5F5F5'  # gris très clair — dépenses
_SEC_BG_OBL    = 'E8EDF3'  # bleu-gris léger — obligations
_SEC_BG_ALERT  = _WHITE    # blanc pur — alertes

_SHEET_NAME = 'DASHBOARD V2'

# ── Noms de mois FR ──────────────────────────────────────────────────────────
_MOIS_NOMS = ['JAN', 'FÉV', 'MAR', 'AVR', 'MAI', 'JUN',
              'JUL', 'AOÛ', 'SEP', 'OCT', 'NOV', 'DÉC']

# ── Helpers de style ─────────────────────────────────────────────────────────

def _font(color: str = _NEAR_BLK, bold: bool = False, size: int = 11,
          name: str = 'Calibri', italic: bool = False) -> Font:
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def _fill(color: str) -> PatternFill:
    return PatternFill('solid', fgColor=color)

def _align(h: str = 'left', v: str = 'center', wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

_BORDER_NONE = Border()
_BORDER_SEP  = Border(bottom=Side(style='thin', color=_G200))

def _fmt_money(val: Optional[float]) -> str:
    if val is None:
        return '—'
    if val == 0:
        return '0 $'
    return f'{val:,.0f} $'

def _fmt_pct(val: float) -> str:
    return f'{val * 100:.0f} %'


# ── Setup colonnes ───────────────────────────────────────────────────────────

def _setup_columns(ws):
    widths = {'A': 2.5, 'B': 24, 'C': 14, 'D': 14, 'E': 16,
              'F': 12, 'G': 16, 'H': 16, 'I': 10, 'J': 10, 'K': 18, 'L': 3}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.sheet_properties.tabColor = '1D4ED8'


# ── Section writers ──────────────────────────────────────────────────────────

def _clear_row(ws, row: int, bg: str = _WHITE, cols: int = 12):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(bg)
        cell.border = _BORDER_NONE
        cell.font = _font(bg, size=2)  # invisible placeholder
        cell.alignment = _align()

def _spacer(ws, row: int, height: float = 8, bg: str = _WHITE):
    ws.row_dimensions[row].height = height
    _clear_row(ws, row, bg)


def _section_separator(ws, row: int, bg_above: str = _WHITE,
                       bg_below: str = _WHITE) -> int:
    """Ligne de séparation fine entre sections (transition de fonds)."""
    _spacer(ws, row, 8, bg_above)
    row += 1
    ws.row_dimensions[row].height = 1.5
    for c in range(1, 13):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(_WHITE)
        cell.border = Border(bottom=Side(style='thin', color=_G200))
    row += 1
    _spacer(ws, row, 12, bg_below)
    return row + 1


def _section_title(ws, row: int, title: str, subtitle: str = '',
                   bg: str = _WHITE) -> int:
    """Titre de section avec style uniforme — 13pt bold, plus d'espace."""
    # Extra top breathing room
    _spacer(ws, row, 6, bg)
    row += 1

    ws.row_dimensions[row].height = 30
    _clear_row(ws, row, bg)
    cell = ws.cell(row=row, column=2)
    cell.value = title
    cell.font = _font(_NEAR_BLK, bold=True, size=13)
    cell.alignment = _align('left', 'bottom')
    cell.fill = _fill(bg)
    if subtitle:
        sc = ws.cell(row=row, column=7)
        sc.value = subtitle
        sc.font = _font(_G400, size=9)
        sc.alignment = _align('right', 'bottom')
        sc.fill = _fill(bg)
    row += 1
    # Thin underline
    ws.row_dimensions[row].height = 2
    for c in range(2, 11):
        ws.cell(row=row, column=c).border = Border(
            bottom=Side(style='thin', color=_G300))
        ws.cell(row=row, column=c).fill = _fill(bg)
    row += 1
    _spacer(ws, row, 6, bg)
    return row + 1


def _write_header(ws, ctx: 'CockpitContext', row: int) -> int:
    """Header : titre + score badge."""
    # Spacer top
    _spacer(ws, row, 10)
    row += 1

    ws.row_dimensions[row].height = 36
    _clear_row(ws, row)

    # Titre
    cell = ws.cell(row=row, column=2)
    cell.value = 'COCKPIT · 2026'
    cell.font = _font(_NEAR_BLK, bold=True, size=18)
    cell.alignment = _align('left', 'center')

    # Score badge
    score_text = f'{ctx.score_emoji}  {ctx.score_label}'
    badge = ws.cell(row=row, column=9)
    badge.value = score_text
    badge.font = _font(_NEAR_BLK, bold=True, size=14)
    badge.alignment = _align('right', 'center')

    row += 1
    _spacer(ws, row, 8)
    return row + 1


def _write_priority(ws, ctx: 'CockpitContext', row: int) -> int:
    """Barre de priorité colorée — accent gauche épais."""
    ac = ctx.alert_ctx if not isinstance(ctx.alert_ctx, str) else None
    text, pale_bg, dark_fg = '', _GRN50, _GRN700

    # Acompte en retard
    if (ac and ac.next_installment and ac.installments
            and ac.installments.required):
        nxt = ac.next_installment
        if nxt.status == 'EN RETARD':
            text = (f'⚠  Acompte en retard de {-nxt.days_left} j '
                    f'— {nxt.period.amount:,.0f} $ dû le {nxt.period.due_date}')
            pale_bg, dark_fg = _RED50, _RED800
        elif nxt.status == 'IMMINENT':
            text = (f'⏳  Acompte dans {nxt.days_left} j '
                    f'— {nxt.period.amount:,.0f} $ le {nxt.period.due_date}')
            pale_bg, dark_fg = _YEL50, _YEL800

    # Fallback : alerte la plus sévère
    if not text and ctx.alerts:
        top = ctx.alerts[0]
        text = top.message
        if top.severity == 'HIGH':
            pale_bg, dark_fg = _RED50, _RED800
        elif top.severity == 'MEDIUM':
            pale_bg, dark_fg = _YEL50, _YEL800

    if not text:
        text = '✓  Aucune alerte critique — tout est en ordre'

    # Accent bar (col A = accent, B-K = message background)
    ws.row_dimensions[row].height = 32
    ws.cell(row=row, column=1).fill = _fill(dark_fg)
    for c in range(2, 12):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(pale_bg)
        cell.border = _BORDER_NONE
    prio_cell = ws.cell(row=row, column=2)
    prio_cell.value = f'  {text}'
    prio_cell.font = _font(dark_fg, bold=True, size=11)
    prio_cell.alignment = _align('left', 'center')

    row += 1
    _spacer(ws, row, 12)
    return row + 1


def _write_kpi_strip(ws, ctx: 'CockpitContext', row: int) -> int:
    """8 KPI cards sur 2 rangées de 4, fond gris clair, plus aéré."""
    ac = ctx.alert_ctx if not isinstance(ctx.alert_ctx, str) else None

    # Calculer les KPI (logique métier inchangée)
    rev_freelance = sum(m.get('freelance', 0) for m in ctx.monthly_revenue)
    rev_salaire   = sum(m.get('salaire', 0) for m in ctx.monthly_revenue)
    rev_total     = rev_freelance + rev_salaire
    _ar_raw       = ac.ar.total_st if ac and ac.ar else None
    ar_total      = float(_ar_raw) if _ar_raw is not None else 0.0
    dep_total     = sum(ctx.expenses_by_category.values())
    profit        = rev_total - dep_total

    taxes_rest = ctx.taxes_current_month
    taxes_label_mois = ctx.taxes_current_label.lower() if ctx.taxes_current_label else ''

    impot_est = 0.0
    if ac and ac.tax_estimate:
        te = ac.tax_estimate
        impot_est = te.impot_federal + te.impot_qc + te.cotisations_autonomes

    acomptes_ytd = ctx.ytd_2025

    prochain_acompte = '—'
    if ac and ac.next_installment and ac.installments and ac.installments.required:
        nxt = ac.next_installment
        prochain_acompte = f'{nxt.period.due_date.strftime("%-d %b %Y")}'

    kpis = [
        [('Revenus encaissés', _fmt_money(rev_total)),
         ('À recevoir', _fmt_money(ar_total)),
         ('Dépenses déductibles', _fmt_money(dep_total)),
         ('Profit brut', _fmt_money(profit))],
        [(f'TPS/TVQ à remettre — {taxes_label_mois}' if taxes_label_mois else 'TPS/TVQ restantes',
          _fmt_money(taxes_rest)),
         ('Impôt estimé 2026', _fmt_money(impot_est)),
         ('Acomptes versés AF25', _fmt_money(acomptes_ytd)),
         ('Prochain acompte', prochain_acompte)],
    ]

    _KPI_BG = _SEC_BG_KPI    # section background
    _CARD   = _G200           # card inner bg (slightly darker than section)

    for kpi_row_idx, kpi_row in enumerate(kpis):
        # Padding top (section bg)
        _spacer(ws, row, 8, _KPI_BG)
        row += 1

        # Card top padding
        _spacer(ws, row, 6, _KPI_BG)
        row += 1

        # Labels row — small, grey, uppercase
        ws.row_dimensions[row].height = 20
        _clear_row(ws, row, _KPI_BG)
        for i, (label, _) in enumerate(kpi_row):
            col = 2 + i * 3
            cell = ws.cell(row=row, column=col)
            cell.value = label.upper()
            cell.font = _font(_G700, bold=False, size=9)
            cell.alignment = _align('left', 'bottom')
            cell.fill = _fill(_CARD)
            # Extend card bg to the 2 adjacent columns
            for dc in (1, 2):
                ws.cell(row=row, column=col + dc).fill = _fill(_CARD)
        row += 1

        # Values row — large, bold, dark, taller for visual impact
        ws.row_dimensions[row].height = 44
        _clear_row(ws, row, _KPI_BG)
        for i, (_, value) in enumerate(kpi_row):
            col = 2 + i * 3
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = _font(_NEAR_BLK, bold=True, size=20)
            cell.alignment = _align('left', 'center')
            cell.fill = _fill(_CARD)
            # Extend card bg to the 2 adjacent columns
            for dc in (1, 2):
                ws.cell(row=row, column=col + dc).fill = _fill(_CARD)
        row += 1

        # Card bottom padding
        _spacer(ws, row, 8, _KPI_BG)
        row += 1

        # Thin divider between KPI rows (still in section bg)
        if kpi_row_idx == 0:
            ws.row_dimensions[row].height = 1.5
            _clear_row(ws, row, _KPI_BG)
            for c in range(2, 11):
                ws.cell(row=row, column=c).border = Border(
                    bottom=Side(style='thin', color=_G200))
            row += 1

    # Section end separator — transition KPI bg → REV bg
    return _section_separator(ws, row, _KPI_BG, _SEC_BG_REV)


def _write_revenus_mois(ws, ctx: 'CockpitContext', row: int) -> int:
    """Table revenus par mois — lignes alternées, fond blanc."""
    _BG = _SEC_BG_REV
    _ALT = 'EDF0F5'   # bleu-gris très léger vs blanc pur
    row = _section_title(ws, row, 'REVENUS PAR MOIS', bg=_BG)

    # Column headers
    headers = [('B', 'Mois'), ('C', 'Freelance'), ('D', 'Salaire'),
               ('E', 'Dépenses'), ('F', 'Profit')]
    ws.row_dimensions[row].height = 22
    _clear_row(ws, row, _BG)
    for col_letter, hdr in headers:
        col_idx = ord(col_letter) - ord('A') + 1
        cell = ws.cell(row=row, column=col_idx)
        cell.value = hdr
        cell.font = _font(_G500, bold=True, size=9)
        cell.alignment = _align('right' if col_letter != 'B' else 'left', 'center')
        cell.border = _BORDER_SEP
        cell.fill = _fill(_BG)
    row += 1

    # Build monthly lookup (logique inchangée)
    rev_by_m = {m['mois'] % 100: m for m in ctx.monthly_revenue}
    dep_by_m = ctx.expenses_by_month
    totals = {'freelance': 0, 'salaire': 0, 'depenses': 0, 'profit': 0}
    row_idx = 0

    for m in range(1, 13):
        rev = rev_by_m.get(m, {})
        freelance = rev.get('freelance', 0)
        salaire   = rev.get('salaire', 0)
        dep       = dep_by_m.get(m, 0)
        profit    = freelance + salaire - dep

        if freelance == 0 and salaire == 0 and dep == 0:
            continue

        totals['freelance'] += freelance
        totals['salaire']   += salaire
        totals['depenses']  += dep
        totals['profit']    += profit

        row_bg = _ALT if row_idx % 2 == 1 else _BG
        ws.row_dimensions[row].height = 28
        _clear_row(ws, row, row_bg)

        ws.cell(row=row, column=2).value = _MOIS_NOMS[m - 1]
        ws.cell(row=row, column=2).font = _font(_G700, bold=False, size=10)
        ws.cell(row=row, column=2).alignment = _align('left', 'center')
        ws.cell(row=row, column=2).fill = _fill(row_bg)

        for ci, val in [(3, freelance), (4, salaire), (5, dep), (6, profit)]:
            c = ws.cell(row=row, column=ci)
            c.value = _fmt_money(val) if val is not None else '—'
            c.font = _font(_NEAR_BLK, size=10)
            c.alignment = _align('right', 'center')
            c.fill = _fill(row_bg)

        row += 1
        row_idx += 1

    # Total row
    _spacer(ws, row, 2, _BG)
    row += 1
    ws.row_dimensions[row].height = 30
    _clear_row(ws, row, _BG)
    _top_border = Border(top=Side(style='medium', color=_G300))
    ws.cell(row=row, column=2).value = 'TOTAL'
    ws.cell(row=row, column=2).font = _font(_NEAR_BLK, bold=True, size=11)
    ws.cell(row=row, column=2).border = _top_border
    ws.cell(row=row, column=2).fill = _fill(_BG)
    for ci, key in [(3, 'freelance'), (4, 'salaire'), (5, 'depenses'), (6, 'profit')]:
        c = ws.cell(row=row, column=ci)
        c.value = _fmt_money(totals[key])
        c.font = _font(_NEAR_BLK, bold=True, size=11)
        c.alignment = _align('right', 'center')
        c.border = _top_border
        c.fill = _fill(_BG)
    row += 1

    return _section_separator(ws, row, _BG, _SEC_BG_REV)  # → chart section


# ── Graphique colonnes empilées ─────────────────────────────────────────────

_DATA_ZONE_ROW = 200   # zone cachée pour les données du graphique

def _write_chart_revenus_depenses(ws, ctx: 'CockpitContext', row: int) -> int:
    """Graphique colonnes empilées — Freelance + Salaire + À recevoir par mois.

    Reproduit le graphique de l'ancien onglet TABLEAU DE BORD dans DASHBOARD V2.
    Les données source sont écrites dans une zone cachée (lignes 200+) et le
    graphique est ancré dans la zone visible.
    """
    _BG = _SEC_BG_REV

    # ── Construire les données mensuelles ────────────────────────────────────
    rev_by_m = {m['mois'] % 100: m for m in ctx.monthly_revenue}
    ar_by_m  = ctx.ar_par_mois or {}

    months_with_data = sorted({
        m for m in range(1, 13)
        if rev_by_m.get(m, {}).get('freelance', 0) > 0
        or rev_by_m.get(m, {}).get('salaire', 0) > 0
        or ar_by_m.get(m, 0) > 0
    })

    if not months_with_data:
        return _section_separator(ws, row, _BG, _SEC_BG_DEP)

    ds = _DATA_ZONE_ROW   # data start row
    n  = len(months_with_data)

    # Headers
    ws.cell(row=ds, column=1).value = 'Mois'
    ws.cell(row=ds, column=2).value = 'Freelance'
    ws.cell(row=ds, column=3).value = 'Salaire'
    ws.cell(row=ds, column=4).value = 'À recevoir'

    for i, m in enumerate(months_with_data):
        r   = ds + 1 + i
        rev = rev_by_m.get(m, {})
        ws.cell(row=r, column=1).value = _MOIS_NOMS[m - 1]
        ws.cell(row=r, column=2).value = rev.get('freelance', 0)
        ws.cell(row=r, column=3).value = rev.get('salaire', 0)
        ws.cell(row=r, column=4).value = ar_by_m.get(m, 0)

    # Cacher les lignes de données
    for r in range(ds, ds + 1 + n):
        ws.row_dimensions[r].hidden = True

    # ── NE PAS utiliser ws.add_chart() — injection raw XML dans le ZIP ──────
    # Stocker les données pour la génération raw du chart XML
    ws._chart_data_cache = {
        'months': [_MOIS_NOMS[m - 1] for m in months_with_data],
        'series': [
            ('Freelance', _BLU700, [rev_by_m.get(m, {}).get('freelance', 0) for m in months_with_data]),
            ('Salaire',   _GRN700, [rev_by_m.get(m, {}).get('salaire', 0) for m in months_with_data]),
            ('À recevoir', _G400,  [ar_by_m.get(m, 0) for m in months_with_data]),
        ],
        'ds': ds,
        'n': n,
        'anchor_row': row,   # row 0-indexed for drawing = row-1
    }

    # Réserver l'espace visuel (~18 lignes)
    chart_rows = 18
    for r in range(row, row + chart_rows):
        _clear_row(ws, r, _BG)
        ws.row_dimensions[r].height = 15
    row += chart_rows

    return _section_separator(ws, row, _BG, _SEC_BG_DEP)


def _write_depenses_cat(ws, ctx: 'CockpitContext', row: int) -> int:
    """Table dépenses par catégorie — fond gris clair, montants bold."""
    _BG = _SEC_BG_DEP
    _ALT = 'EBEBEB'   # gris légèrement plus foncé vs #F5F5F5
    row = _section_title(ws, row, 'DÉPENSES PAR CATÉGORIE', bg=_BG)

    # Headers
    ws.row_dimensions[row].height = 22
    _clear_row(ws, row, _BG)
    for col_letter, hdr in [('B', 'Catégorie'), ('E', 'Montant'), ('F', 'Part')]:
        col_idx = ord(col_letter) - ord('A') + 1
        cell = ws.cell(row=row, column=col_idx)
        cell.value = hdr
        cell.font = _font(_G500, bold=True, size=9)
        cell.alignment = _align('right' if col_letter != 'B' else 'left', 'center')
        cell.border = _BORDER_SEP
        cell.fill = _fill(_BG)
    row += 1

    total_dep = sum(ctx.expenses_by_category.values()) or 1
    sorted_cats = sorted(ctx.expenses_by_category.items(), key=lambda x: -x[1])

    for i, (cat, montant) in enumerate(sorted_cats):
        pct = montant / total_dep
        row_bg = _ALT if i % 2 == 1 else _BG
        ws.row_dimensions[row].height = 28
        _clear_row(ws, row, row_bg)

        c_cat = ws.cell(row=row, column=2)
        c_cat.value = str(cat)[:30]
        c_cat.font = _font(_G700, size=10)
        c_cat.alignment = _align('left', 'center')
        c_cat.fill = _fill(row_bg)

        c_montant = ws.cell(row=row, column=5)
        c_montant.value = _fmt_money(montant)
        c_montant.font = _font(_NEAR_BLK, bold=True, size=11)
        c_montant.alignment = _align('right', 'center')
        c_montant.fill = _fill(row_bg)

        c_pct = ws.cell(row=row, column=6)
        c_pct.value = _fmt_pct(pct)
        pct_color = _NEAR_BLK if pct >= 0.2 else (_G500 if pct >= 0.05 else _G400)
        c_pct.font = _font(pct_color, bold=(pct >= 0.15), size=10)
        c_pct.alignment = _align('right', 'center')
        c_pct.fill = _fill(row_bg)

        # Barre de progression inline (caractères pleins, plus longue)
        bar_len = 20
        filled = max(1, round(pct * bar_len)) if pct > 0.01 else 0
        bar_char_on  = '\u2588'   # █ plein
        bar_char_off = '\u2591'   # ░ léger
        bar_color = _NEAR_BLK if pct >= 0.20 else (_G500 if pct >= 0.08 else _G300)
        c_bar = ws.cell(row=row, column=7)
        c_bar.value = bar_char_on * filled + bar_char_off * (bar_len - filled)
        c_bar.font = Font(name='Calibri', size=7, color=bar_color)
        c_bar.alignment = _align('left', 'center')
        c_bar.fill = _fill(row_bg)

        row += 1

    # Total
    _spacer(ws, row, 2, _BG)
    row += 1
    ws.row_dimensions[row].height = 30
    _clear_row(ws, row, _BG)
    _top_border = Border(top=Side(style='medium', color=_G300))
    ws.cell(row=row, column=2).value = 'TOTAL'
    ws.cell(row=row, column=2).font = _font(_NEAR_BLK, bold=True, size=11)
    ws.cell(row=row, column=2).border = _top_border
    ws.cell(row=row, column=2).fill = _fill(_BG)
    c_tot = ws.cell(row=row, column=5)
    c_tot.value = _fmt_money(total_dep if total_dep != 1 else 0)
    c_tot.font = _font(_NEAR_BLK, bold=True, size=11)
    c_tot.alignment = _align('right', 'center')
    c_tot.border = _top_border
    c_tot.fill = _fill(_BG)
    row += 1

    return _section_separator(ws, row, _BG, _SEC_BG_OBL)


def _write_obligations(ws, ctx: 'CockpitContext', row: int) -> int:
    """Obligations financières — séparées AF2025 / AF2026, fond bleu-gris."""
    _BG  = _SEC_BG_OBL
    _ALT = 'D8DFE9'   # bleu-gris plus foncé vs fond bleu-gris léger
    row = _section_title(ws, row, 'OBLIGATIONS FINANCIÈRES', bg=_BG)

    if not ctx.obligations:
        _clear_row(ws, row, _BG)
        ws.cell(row=row, column=2).value = 'Aucune obligation active'
        ws.cell(row=row, column=2).font = _font(_G400, size=10)
        ws.cell(row=row, column=2).fill = _fill(_BG)
        row += 1
        return _section_separator(ws, row, _BG, _SEC_BG_ALERT)

    # Regrouper par année fiscale (due_date.year - 1 = AF)
    groups: dict[str, list] = {}
    for obl in ctx.obligations:
        dd = obl.get('due_date')
        # Resolve year from date or serial number
        dd_year = None
        if dd is not None:
            if hasattr(dd, 'year'):
                dd_year = dd.year
            elif isinstance(dd, (int, float)) and 40000 < dd < 60000:
                from datetime import datetime as _dt, timedelta as _td
                dd_year = (_dt(1899, 12, 30) + _td(days=int(dd))).year
        if dd_year is not None:
            if dd_year <= 2026:
                key = 'OBLIGATIONS AF2025'
            else:
                key = 'OBLIGATIONS AF2026 (estimations)'
        else:
            key = 'AUTRES'
        # Nantel/BIA → catégorie séparée
        label = obl.get('label', '')
        if 'Nantel' in label or 'BIA' in label:
            key = 'PRÊT / FINANCEMENT'
        groups.setdefault(key, []).append(obl)

    first_group = True
    for group_name, obls in groups.items():
        if not first_group:
            _spacer(ws, row, 16, _BG)
            row += 1
        else:
            _spacer(ws, row, 6, _BG)
            row += 1

        # Group subtitle
        ws.row_dimensions[row].height = 24
        _clear_row(ws, row, _BG)
        ws.cell(row=row, column=2).value = group_name
        ws.cell(row=row, column=2).font = _font(_G500, bold=True, size=9)
        ws.cell(row=row, column=2).alignment = _align('left', 'bottom')
        ws.cell(row=row, column=2).fill = _fill(_BG)
        row += 1

        # Headers
        ws.row_dimensions[row].height = 20
        _clear_row(ws, row, _BG)
        for col_letter, hdr in [('B', 'Obligation'), ('E', 'Payé / Total'),
                                 ('G', 'Progression'), ('I', '%'),
                                 ('K', 'Reste')]:
            col_idx = ord(col_letter) - ord('A') + 1
            cell = ws.cell(row=row, column=col_idx)
            cell.value = hdr
            cell.font = _font(_G500, bold=True, size=9)
            cell.alignment = _align('right' if col_letter not in ('B', 'G') else 'left',
                                    'center')
            cell.border = _BORDER_SEP
            cell.fill = _fill(_BG)
        row += 1

        for i, obl in enumerate(obls):
            pct = obl.get('pct', 0)
            row_bg = _ALT if i % 2 == 1 else _BG
            ws.row_dimensions[row].height = 32
            _clear_row(ws, row, row_bg)

            # Label
            c_label = ws.cell(row=row, column=2)
            c_label.value = obl.get('label', '?')[:32]
            c_label.font = _font(_G700, size=10)
            c_label.alignment = _align('left', 'center')
            c_label.fill = _fill(row_bg)

            # Payé / Total
            paid  = obl.get('paid', 0)
            total = obl.get('total', 0)
            c_pt = ws.cell(row=row, column=5)
            c_pt.value = f'{paid:,.0f} / {total:,.0f} $'
            c_pt.font = _font(_NEAR_BLK, bold=True, size=11)
            c_pt.alignment = _align('right', 'center')
            c_pt.fill = _fill(row_bg)

            # Progress bar — longer (18 chars), stronger contrast
            bar_len = 18
            filled = round(pct * bar_len)
            if pct > 0.5:
                bar_color = _GRN700
            elif pct >= 0.1:
                bar_color = _YEL800
            elif pct > 0:
                bar_color = _G500
            else:
                bar_color = _G300
            bar = '█' * filled + '░' * (bar_len - filled)
            c_bar = ws.cell(row=row, column=7)
            c_bar.value = bar
            c_bar.font = Font(name='Courier New', size=9, color=bar_color)
            c_bar.alignment = _align('left', 'center')
            c_bar.fill = _fill(row_bg)

            # Pct — bold and colored
            c_pct = ws.cell(row=row, column=9)
            c_pct.value = _fmt_pct(pct)
            pct_color = _GRN700 if pct > 0.5 else (_YEL800 if pct >= 0.1 else _G400)
            c_pct.font = _font(pct_color, bold=True, size=11)
            c_pct.alignment = _align('right', 'center')
            c_pct.fill = _fill(row_bg)

            # Reste (solde)
            reste = max(total - paid, 0)
            c_reste = ws.cell(row=row, column=11)
            c_reste.value = f'{reste:,.0f} $'
            reste_color = _GRN700 if reste == 0 else _RED800
            c_reste.font = _font(reste_color, bold=True, size=11)
            c_reste.alignment = _align('right', 'center')
            c_reste.fill = _fill(row_bg)

            row += 1

        # Total line for this group
        if len(obls) > 1:
            _spacer(ws, row, 2, _BG)
            row += 1
            ws.row_dimensions[row].height = 28
            _clear_row(ws, row, _BG)
            _top_border = Border(top=Side(style='medium', color=_G300))
            ws.cell(row=row, column=2).value = 'TOTAL'
            ws.cell(row=row, column=2).font = _font(_NEAR_BLK, bold=True, size=10)
            ws.cell(row=row, column=2).border = _top_border
            ws.cell(row=row, column=2).fill = _fill(_BG)
            grp_paid  = sum(o.get('paid', 0) for o in obls)
            grp_total = sum(o.get('total', 0) for o in obls)
            grp_pct   = (grp_paid / grp_total) if grp_total > 0 else 0
            c_pt = ws.cell(row=row, column=5)
            c_pt.value = f'{grp_paid:,.0f} / {grp_total:,.0f} $'
            c_pt.font = _font(_NEAR_BLK, bold=True, size=10)
            c_pt.alignment = _align('right', 'center')
            c_pt.border = _top_border
            c_pt.fill = _fill(_BG)
            c_pct = ws.cell(row=row, column=9)
            c_pct.value = _fmt_pct(grp_pct)
            pct_color = _GRN700 if grp_pct > 0.5 else (_YEL800 if grp_pct >= 0.1 else _G400)
            c_pct.font = _font(pct_color, bold=True, size=11)
            c_pct.alignment = _align('right', 'center')
            c_pct.border = _top_border
            c_pct.fill = _fill(_BG)
            grp_reste = max(grp_total - grp_paid, 0)
            c_reste = ws.cell(row=row, column=11)
            c_reste.value = f'{grp_reste:,.0f} $'
            c_reste.font = _font(_RED800 if grp_reste > 0 else _GRN700, bold=True, size=10)
            c_reste.alignment = _align('right', 'center')
            c_reste.border = _top_border
            c_reste.fill = _fill(_BG)
            row += 1

        first_group = False

    # ── GRAND TOTAL DETTES ──────────────────────────────────────────────────
    # Somme de toutes les obligations + taxes mois courant − déjà payé
    all_obls = ctx.obligations or []
    grand_total = sum(o.get('total', 0) for o in all_obls)
    grand_paid  = sum(o.get('paid', 0)  for o in all_obls)

    # Ajouter les taxes du mois courant (TPS/TVQ à remettre)
    taxes_mois  = ctx.taxes_current_month or 0.0
    taxes_label = ctx.taxes_current_label or ''

    grand_total += taxes_mois
    grand_solde  = round(grand_total - grand_paid, 2)

    _spacer(ws, row, 10, _BG)
    row += 1
    ws.row_dimensions[row].height = 36
    _clear_row(ws, row, _BG)
    _heavy_border = Border(top=Side(style='double', color=_G500))

    # Label
    c_gt_label = ws.cell(row=row, column=2)
    c_gt_label.value = 'TOTAL DETTES'
    if taxes_mois > 0 and taxes_label:
        c_gt_extra = ws.cell(row=row, column=3)
        c_gt_extra.value = f'(incl. TPS/TVQ {taxes_label} : {taxes_mois:,.0f} $)'
        c_gt_extra.font = _font(_G400, size=9, italic=True)
        c_gt_extra.alignment = _align('left', 'center')
        c_gt_extra.fill = _fill(_BG)
    c_gt_label.font = _font(_NEAR_BLK, bold=True, size=12)
    c_gt_label.alignment = _align('left', 'center')
    c_gt_label.border = _heavy_border
    c_gt_label.fill = _fill(_BG)

    # Payé / Total
    c_gt_pt = ws.cell(row=row, column=5)
    c_gt_pt.value = f'{grand_paid:,.0f} / {grand_total:,.0f} $'
    c_gt_pt.font = _font(_NEAR_BLK, bold=True, size=12)
    c_gt_pt.alignment = _align('right', 'center')
    c_gt_pt.border = _heavy_border
    c_gt_pt.fill = _fill(_BG)

    # Solde restant
    solde_color = _RED800 if grand_solde > 10000 else (_YEL800 if grand_solde > 5000 else _G700)
    c_gt_solde = ws.cell(row=row, column=7)
    c_gt_solde.value = f'Solde : {grand_solde:,.0f} $'
    c_gt_solde.font = _font(solde_color, bold=True, size=12)
    c_gt_solde.alignment = _align('left', 'center')
    c_gt_solde.border = _heavy_border
    c_gt_solde.fill = _fill(_BG)

    # Pct
    grand_pct = (grand_paid / grand_total) if grand_total > 0 else 0
    c_gt_pct = ws.cell(row=row, column=9)
    c_gt_pct.value = _fmt_pct(grand_pct)
    pct_color = _GRN700 if grand_pct > 0.5 else (_YEL800 if grand_pct >= 0.1 else _G400)
    c_gt_pct.font = _font(pct_color, bold=True, size=12)
    c_gt_pct.alignment = _align('right', 'center')
    c_gt_pct.border = _heavy_border
    c_gt_pct.fill = _fill(_BG)
    # Reste grand total
    c_gt_reste = ws.cell(row=row, column=11)
    c_gt_reste.value = f'{grand_solde:,.0f} $'
    c_gt_reste.font = _font(solde_color, bold=True, size=12)
    c_gt_reste.alignment = _align('right', 'center')
    c_gt_reste.border = _heavy_border
    c_gt_reste.fill = _fill(_BG)
    row += 1

    return _section_separator(ws, row, _BG, _SEC_BG_ALERT)


def _write_alertes(ws, ctx: 'CockpitContext', row: int) -> int:
    """Section alertes — cards visuelles avec fond teinté, fond blanc."""
    _BG = _SEC_BG_ALERT
    active = [a for a in ctx.alerts if a.severity in ('HIGH', 'MEDIUM', 'LOW')]
    count = len(active)

    count_text = f'{count} active{"s" if count != 1 else ""}'
    row = _section_title(ws, row, 'ALERTES', count_text, bg=_BG)

    _sev_bg = {'HIGH': _RED50, 'MEDIUM': _YEL50, 'LOW': _BLU50}
    _sev_fg = {'HIGH': _RED800, 'MEDIUM': _YEL800, 'LOW': _BLU700}

    display = active[:6] if active else []
    if not display:
        # Card-like "all clear" block
        _spacer(ws, row, 4, _BG)
        row += 1
        _clear_row(ws, row, _GRN50)
        ws.row_dimensions[row].height = 36
        ws.cell(row=row, column=1).fill = _fill(_GRN700)  # accent left
        ws.cell(row=row, column=2).value = '  ✓  Aucune alerte — tout va bien'
        ws.cell(row=row, column=2).font = _font(_GRN700, size=11)
        ws.cell(row=row, column=2).fill = _fill(_GRN50)
        ws.cell(row=row, column=2).alignment = _align('left', 'center')
        row += 1
        _spacer(ws, row, 4, _BG)
        row += 1
    else:
        for a in display:
            bg = _sev_bg.get(a.severity, _WHITE)
            fg = _sev_fg.get(a.severity, _G400)

            # Top padding for card effect
            _spacer(ws, row, 6, _BG)
            row += 1

            # Card content row — taller, with accent bar
            ws.row_dimensions[row].height = 38
            _clear_row(ws, row, bg)
            # Accent bar left
            ws.cell(row=row, column=1).fill = _fill(fg)
            c = ws.cell(row=row, column=2)
            # Enrichir l'alerte taxes avec le mois courant (garder le cumul annuel)
            msg = a.message
            if a.type in ('TAXES_ELEVEES', 'TAXES_MODEREES'):
                _t_mois = ctx.taxes_current_label or ''
                _t_val = ctx.taxes_current_month
                if _t_val is not None and _t_val > 0 and _t_mois:
                    # Montrer le mois courant en complément, sans écraser le cumul annuel
                    msg = (f'Prochaine remise ({_t_mois.lower()}) : '
                           f'{_t_val:,.2f}\u202f$ — {a.message}')
            c.value = f'  {a.icon}  {msg}'
            c.font = _font(fg, bold=(a.severity == 'HIGH'), size=10)
            c.alignment = _align('left', 'center', wrap=True)
            c.fill = _fill(bg)
            row += 1

            # Bottom padding for card effect
            _spacer(ws, row, 6, _BG)
            row += 1

    _spacer(ws, row, 8, _BG)
    return row + 1


def _write_footer(ws, row: int) -> int:
    """Footer discret avec timestamp."""
    # Separator
    ws.row_dimensions[row].height = 1.5
    for c in range(2, 11):
        ws.cell(row=row, column=c).border = Border(
            bottom=Side(style='thin', color=_G200))
        ws.cell(row=row, column=c).fill = _fill(_WHITE)
    row += 1

    _spacer(ws, row, 8, _SEC_BG_ALERT)
    row += 1

    ws.row_dimensions[row].height = 16
    _clear_row(ws, row)
    _MOIS_FR = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin',
                 'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']
    now = datetime.now()
    ts = f'{now.day} {_MOIS_FR[now.month - 1]} {now.year} à {now:%H:%M}'
    cell = ws.cell(row=row, column=2)
    cell.value = f'Mis à jour : {ts}'
    cell.font = _font(_G400, size=8)
    cell.alignment = _align('left', 'center')

    row += 1
    ws.row_dimensions[row].height = 14
    _clear_row(ws, row)
    cell2 = ws.cell(row=row, column=2)
    cell2.value = 'Source : CockpitContext · Données privées'
    cell2.font = _font(_G400, size=7)
    cell2.alignment = _align('left', 'center')

    row += 1
    _spacer(ws, row, 10)
    return row + 1


# ── Injection ZIP (préserve les cached values des autres feuilles) ───────────

_XLSX_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_REL_NS  = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_PKG_NS  = 'http://schemas.openxmlformats.org/package/2006/relationships'
_CT_NS   = 'http://schemas.openxmlformats.org/package/2006/content-types'
_WS_TYPE = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet'
_WS_CT   = 'application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml'


# ── Helpers pour tracer les assets d'une feuille (drawings, charts, rels) ────

def _resolve_rel_path(parent_path: str, target: str) -> Optional[str]:
    """Résout un chemin relatif (.rels Target) par rapport au parent."""
    if not target or target.startswith('http'):
        return None
    if target.startswith('/'):
        return target.lstrip('/')
    parent_dir = parent_path.rsplit('/', 1)[0] if '/' in parent_path else ''
    if target.startswith('../'):
        resolved = os.path.normpath(os.path.join(parent_dir, target))
    else:
        resolved = f'{parent_dir}/{target}' if parent_dir else target
    return resolved.replace('\\', '/')


def _collect_zip_refs(z: zipfile.ZipFile, rels_path: str,
                      parent_path: str, assets: dict) -> None:
    """Collecte récursivement les fichiers référencés dans un .rels (drawings, charts)."""
    rels_xml = ET.fromstring(z.read(rels_path))
    for rel in rels_xml:
        if not (rel.tag.endswith('}Relationship') or rel.tag == 'Relationship'):
            continue
        resolved = _resolve_rel_path(parent_path, rel.get('Target', ''))
        if not resolved or resolved not in z.namelist() or resolved in assets:
            continue
        assets[resolved] = z.read(resolved)
        # Sous-rels (ex: drawing → chart)
        r_dir, r_file = (resolved.rsplit('/', 1) if '/' in resolved
                         else ('', resolved))
        sub_rels = f'{r_dir}/_rels/{r_file}.rels'
        if sub_rels in z.namelist() and sub_rels not in assets:
            assets[sub_rels] = z.read(sub_rels)
            _collect_zip_refs(z, sub_rels, resolved, assets)


def _collect_zip_ref_paths(z: zipfile.ZipFile, rels_path: str,
                           parent_path: str, paths: set) -> None:
    """Comme _collect_zip_refs mais ne collecte que les chemins (set)."""
    rels_xml = ET.fromstring(z.read(rels_path))
    for rel in rels_xml:
        if not (rel.tag.endswith('}Relationship') or rel.tag == 'Relationship'):
            continue
        resolved = _resolve_rel_path(parent_path, rel.get('Target', ''))
        if not resolved or resolved not in z.namelist() or resolved in paths:
            continue
        paths.add(resolved)
        r_dir, r_file = (resolved.rsplit('/', 1) if '/' in resolved
                         else ('', resolved))
        sub_rels = f'{r_dir}/_rels/{r_file}.rels'
        if sub_rels in z.namelist() and sub_rels not in paths:
            paths.add(sub_rels)
            _collect_zip_ref_paths(z, sub_rels, resolved, paths)


def _ensure_content_type_for_shared_strings(xlsm_path: str) -> None:
    """Ajoute l'Override pour sharedStrings.xml dans [Content_Types].xml si manquant."""
    import re as _re
    with zipfile.ZipFile(xlsm_path, 'r') as z:
        ct_bytes = z.read('[Content_Types].xml')
    ct = ct_bytes.decode('utf-8')
    if 'sharedStrings' in ct:
        return  # Déjà présent
    # Ajouter avant </Types>
    ss_override = ('<Override PartName="/xl/sharedStrings.xml" '
                   'ContentType="application/vnd.openxmlformats-officedocument.'
                   'spreadsheetml.sharedStrings+xml"/>')
    ct = ct.replace('</Types>', f'{ss_override}</Types>')
    # Réécrire le ZIP
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsm')
    os.close(tmp_fd)
    with zipfile.ZipFile(xlsm_path, 'r') as src, \
         zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            if item.filename == '[Content_Types].xml':
                dst.writestr(item, ct.encode('utf-8'))
            else:
                dst.writestr(item, src.read(item.filename))
    shutil.move(tmp_path, xlsm_path)


def _merge_content_types(orig_ct: bytes, temp_ct: bytes) -> bytes:
    """Fusionne [Content_Types].xml : original comme base + ajouts du temp.

    L'original contient les entrées pour comments, vmlDrawings, etc.
    Le temp contient les entrées pour chart, drawing (DV2).
    Le résultat a les deux.
    """
    import re as _re
    orig = orig_ct.decode('utf-8')
    temp = temp_ct.decode('utf-8')

    # PartNames existants dans l'original
    orig_parts = set(_re.findall(r'PartName="([^"]+)"', orig))
    # Extensions existantes dans l'original
    orig_exts = set(_re.findall(r'<Default[^>]+Extension="([^"]+)"', orig))

    additions = []
    # Overrides du temp absents de l'original
    for m in _re.finditer(r'<Override\s+[^>]+/>', temp):
        part_m = _re.search(r'PartName="([^"]+)"', m.group())
        if part_m and part_m.group(1) not in orig_parts:
            additions.append(m.group())
    # Defaults du temp absents de l'original
    for m in _re.finditer(r'<Default\s+[^>]+/>', temp):
        ext_m = _re.search(r'Extension="([^"]+)"', m.group())
        if ext_m and ext_m.group(1) not in orig_exts:
            additions.append(m.group())

    if additions:
        orig = orig.replace('</Types>', ''.join(additions) + '</Types>')

    return orig.encode('utf-8')


def _convert_shared_to_inline(sheet_xml: bytes, shared_strings: list[str]) -> bytes:
    """Convertit toutes les cellules t='s' d'un sheet XML en inline strings.

    openpyxl réindexe les shared strings lors du save. Quand on restaure le
    sharedStrings.xml original, les indices du temp ne correspondent plus.
    Solution : convertir toutes les refs shared strings en inline strings
    dans le sheet DV2 (le seul sheet provenant du temp).

    Parameters
    ----------
    sheet_xml       : bytes du sheet XML (provenant du temp)
    shared_strings  : table des shared strings du TEMP (par index)
    """
    import re
    text = sheet_xml.decode('utf-8')

    def _replace_shared(m):
        full_match = m.group(0)
        # Extraire l'index de la shared string
        v_match = re.search(r'<v>(\d+)</v>', full_match)
        if not v_match:
            return full_match
        idx = int(v_match.group(1))
        if idx >= len(shared_strings):
            return full_match
        val = shared_strings[idx]
        # Échapper les caractères XML
        val = val.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        # Remplacer t="s" par t="inlineStr" et <v>N</v> par <is><t>VAL</t></is>
        new = re.sub(r' t="s"', ' t="inlineStr"', full_match, count=1)
        new = re.sub(r'<v>\d+</v>', f'<is><t>{val}</t></is>', new, count=1)
        return new

    # Matcher chaque cellule <c ... t="s" ...>...<v>N</v>...</c>
    text = re.sub(r'<c [^>]*t="s"[^>]*>.*?</c>', _replace_shared, text, flags=re.DOTALL)

    return text.encode('utf-8')


def _remap_drawing_in_rels(orig_rels: bytes, temp_rels: bytes) -> bytes:
    """Utilise les .rels originaux mais remappe les Targets drawing vers le temp.

    openpyxl peut réassigner les numéros de drawing (drawing1 → drawing2) quand
    il ajoute DV2 en position 0. Les .rels originaux pointent vers les anciens
    chemins (ex: drawing1.xml), mais dans le ZIP final, le contenu de TDB est
    dans le chemin du temp (ex: drawing2.xml).

    Cette fonction préserve toutes les références ancillaires (vmlDrawing,
    comments, printerSettings) de l'original mais remplace le Target du
    drawing régulier par celui du temp.
    """
    import re as _re

    # Trouver le Target du drawing régulier dans le temp
    temp_str = temp_rels.decode('utf-8')
    temp_draw = None
    for m in _re.finditer(r'<Relationship\s+[^>]*/>', temp_str):
        elem = m.group()
        if 'relationships/drawing"' in elem and 'vmlDrawing' not in elem:
            t = _re.search(r'Target="([^"]+)"', elem)
            if t:
                temp_draw = t.group(1)
                break

    if not temp_draw:
        return orig_rels  # Pas de drawing dans le temp → rien à remapper

    # Remplacer le Target du drawing dans l'original
    orig_str = orig_rels.decode('utf-8')

    def _replace(m):
        elem = m.group()
        if 'relationships/drawing"' in elem and 'vmlDrawing' not in elem:
            return _re.sub(r'Target="[^"]+"', f'Target="{temp_draw}"', elem)
        return elem

    result = _re.sub(r'<Relationship\s+[^>]*/>', _replace, orig_str)
    return result.encode('utf-8')


def _find_sheet_target(xlsm_path: str, sheet_name: str) -> Optional[str]:
    """Trouve le fichier XML d'une feuille existante dans le .xlsm.

    Retourne ex: 'worksheets/sheet1.xml' ou None si la feuille n'existe pas.
    """
    with zipfile.ZipFile(xlsm_path, 'r') as z:
        wb_xml = ET.fromstring(z.read('xl/workbook.xml'))
        rels_xml = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))

        # Trouver le rId de la feuille
        for s in wb_xml.iter(f'{{{_XLSX_NS}}}sheet'):
            if s.get('name') == sheet_name:
                rid = s.get(f'{{{_REL_NS}}}id')
                # Trouver le fichier correspondant au rId
                for r in rels_xml.iter(f'{{{_PKG_NS}}}Relationship'):
                    if r.get('Id') == rid:
                        return r.get('Target')
    return None


def _zip_replace_sheet(xlsm_path: str, sheet_target: str, new_sheet_xml: bytes) -> None:
    """Remplace le XML d'une feuille dans le .xlsm sans toucher aux autres.

    Copie toutes les entrées du ZIP sauf sheet_target, puis ajoute la nouvelle.
    """
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsm')
    os.close(tmp_fd)

    # sheet_target peut être absolu ("/xl/worksheets/sheet1.xml") ou
    # relatif ("worksheets/sheet1.xml"). Normaliser vers l'entrée ZIP réelle.
    st = sheet_target.lstrip('/')
    if st.startswith('xl/'):
        full_target = st
    else:
        full_target = f'xl/{st}'

    try:
        with zipfile.ZipFile(xlsm_path, 'r') as src, \
             zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as dst:
            for item in src.infolist():
                if item.filename == full_target:
                    continue  # On va le remplacer
                dst.writestr(item, src.read(item.filename))
            dst.writestr(full_target, new_sheet_xml)
        shutil.move(tmp_path, xlsm_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise


def _zip_replace_file(xlsm_path: str, target_path: str, new_content: bytes) -> None:
    """Remplace un fichier arbitraire dans le ZIP (.xlsm).

    Utilisé pour mettre à jour xl/styles.xml sans toucher aux autres entrées.
    """
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsm')
    os.close(tmp_fd)
    try:
        with zipfile.ZipFile(xlsm_path, 'r') as src, \
             zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as dst:
            for item in src.infolist():
                if item.filename == target_path:
                    continue
                dst.writestr(item, src.read(item.filename))
            dst.writestr(target_path, new_content)
        shutil.move(tmp_path, xlsm_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise


def _zip_add_new_sheet(xlsm_path: str, sheet_name: str,
                       new_sheet_xml: bytes, position: int = 0) -> None:
    """Ajoute une nouvelle feuille dans le .xlsm (première fois seulement).

    Modifie workbook.xml, workbook.xml.rels et [Content_Types].xml.
    """
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsm')
    os.close(tmp_fd)

    try:
        with zipfile.ZipFile(xlsm_path, 'r') as src:
            wb_xml = ET.fromstring(src.read('xl/workbook.xml'))
            rels_xml = ET.fromstring(src.read('xl/_rels/workbook.xml.rels'))
            ct_xml = ET.fromstring(src.read('[Content_Types].xml'))

            # Déterminer le prochain numéro de sheet
            existing_sheets = [n for n in src.namelist()
                               if n.startswith('xl/worksheets/sheet') and n.endswith('.xml')]
            nums = []
            for name in existing_sheets:
                base = name.replace('xl/worksheets/sheet', '').replace('.xml', '')
                if base.isdigit():
                    nums.append(int(base))
            next_num = max(nums) + 1 if nums else 1
            sheet_file = f'worksheets/sheet{next_num}.xml'
            full_path = f'xl/{sheet_file}'

            # Déterminer le prochain rId
            all_rids = [r.get('Id') for r in rels_xml.iter(f'{{{_PKG_NS}}}Relationship')]
            rid_nums = [int(r.replace('rId', '')) for r in all_rids if r.startswith('rId') and r[3:].isdigit()]
            next_rid = f'rId{max(rid_nums) + 1}' if rid_nums else 'rId1'

            # Déterminer le prochain sheetId
            all_sids = [int(s.get('sheetId', 0)) for s in wb_xml.iter(f'{{{_XLSX_NS}}}sheet')]
            next_sid = str(max(all_sids) + 1) if all_sids else '1'

            # Ajouter dans workbook.xml
            sheets_elem = wb_xml.find(f'.//{{{_XLSX_NS}}}sheets')
            new_sheet = ET.SubElement(sheets_elem, f'{{{_XLSX_NS}}}sheet')
            new_sheet.set('name', sheet_name)
            new_sheet.set('sheetId', next_sid)
            new_sheet.set(f'{{{_REL_NS}}}id', next_rid)
            # Déplacer en première position
            if position == 0:
                sheets_elem.remove(new_sheet)
                sheets_elem.insert(0, new_sheet)

            # Ajouter dans workbook.xml.rels
            new_rel = ET.SubElement(rels_xml, f'{{{_PKG_NS}}}Relationship')
            new_rel.set('Id', next_rid)
            new_rel.set('Type', _WS_TYPE)
            new_rel.set('Target', sheet_file)

            # Ajouter dans [Content_Types].xml
            new_ct = ET.SubElement(ct_xml, 'Override')
            new_ct.set('PartName', f'/{full_path}')
            new_ct.set('ContentType', _WS_CT)

            # Écrire le nouveau ZIP
            with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as dst:
                for item in src.infolist():
                    if item.filename == 'xl/workbook.xml':
                        ET.register_namespace('', _XLSX_NS)
                        ET.register_namespace('r', _REL_NS)
                        dst.writestr(item, ET.tostring(wb_xml, xml_declaration=True, encoding='UTF-8'))
                    elif item.filename == 'xl/_rels/workbook.xml.rels':
                        dst.writestr(item, ET.tostring(rels_xml, xml_declaration=True, encoding='UTF-8'))
                    elif item.filename == '[Content_Types].xml':
                        dst.writestr(item, ET.tostring(ct_xml, xml_declaration=True, encoding='UTF-8'))
                    else:
                        dst.writestr(item, src.read(item.filename))
                dst.writestr(full_path, new_sheet_xml)

        shutil.move(tmp_path, xlsm_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise


# ── Chart XML post-processing (cache injection) ──────────────────────────────

_CHART_NS = 'http://schemas.openxmlformats.org/drawingml/2006/chart'


def _build_raw_chart_assets(chart_cache: dict, xlsm_path: str = '') -> dict[str, bytes]:
    """Clone chart2.xml du TABLEAU DE BORD comme base pour le graphique DV2.

    Stratégie V4 : copier chart2.xml 100 % intacte (4 séries, 2 barChart,
    4 axes) → structure OOXML identique à Excel.  On ne modifie QUE les
    valeurs numériques/labels/couleurs dans les caches.  La 4e série
    (Dépenses) reçoit des zéros + remplissage transparent → invisible dans
    le graphique mais la structure reste valide.

    Returns dict {zip_path: bytes} prêt à injecter dans le ZIP final.
    """
    months = chart_cache['months']
    series_data = chart_cache['series']  # [(label, color, [values])]
    anchor_row = chart_cache.get('anchor_row', 34)

    import re as _re

    def _read_chart2(path: str) -> str | None:
        """Lit chart2.xml depuis un fichier xlsm et vérifie sa structure."""
        try:
            import zipfile as _zf
            with _zf.ZipFile(path, 'r') as _z:
                if 'xl/charts/chart2.xml' in _z.namelist():
                    xml = _z.read('xl/charts/chart2.xml').decode('utf-8')
                    # Valider : doit contenir 2 barChart et 4 <ser>
                    if (xml.count('<barChart>') >= 2
                            and xml.count('<ser>') >= 4):
                        return xml
        except Exception:
            pass
        return None

    chart_xml_str = None
    if xlsm_path:
        # 1. Essayer le fichier courant
        chart_xml_str = _read_chart2(xlsm_path)

        # 2. Fallback : backup pristine (Excel-natif, jamais touché par openpyxl)
        if chart_xml_str is None:
            bak = xlsm_path + '.bak_pre_v2_fixes'
            if os.path.exists(bak):
                chart_xml_str = _read_chart2(bak)

    if chart_xml_str is None:
        return {}

    # ── Pas de suppression d'éléments structurels ! ──
    # On garde les 2 barChart, 4 axes, 4 séries intactes.
    # On met seulement à jour les caches de données.

    _MOIS_ABREV = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun',
                   'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']

    def _num_cache(values):
        pts = ''.join(f'<pt idx="{i}"><v>{v}</v></pt>' for i, v in enumerate(values))
        for i in range(len(values), 12):
            pts += f'<pt idx="{i}"><v>0</v></pt>'
        return f'<numCache><formatCode>General</formatCode><ptCount val="12" />{pts}</numCache>'

    def _str_cache_months():
        pts = ''.join(f'<pt idx="{i}"><v>{m}</v></pt>' for i, m in enumerate(_MOIS_ABREV))
        return f'<strCache><ptCount val="12" />{pts}</strCache>'

    # Remplacer les strCache des catégories (mois) dans les <cat> blocs
    chart_xml_str = _re.sub(
        r'(<cat><strRef><f>[^<]+</f>)<strCache>.*?</strCache>',
        lambda m: m.group(1) + _str_cache_months(),
        chart_xml_str, flags=_re.DOTALL)

    # Construire la liste complète de 4 séries :
    # séries 0-2 = données DV2, série 3 = Dépenses (zéros, transparente)
    all_series = list(series_data)  # copie des 3 séries DV2
    # Ajouter une 4e série fantôme (Dépenses) — zéros partout
    all_series.append(('Dépenses', '00000000', [0] * 12))  # couleur ignorée

    # Mettre à jour chaque série (en ordre inverse pour stabilité des positions regex)
    ser_blocks = list(_re.finditer(r'<ser>.*?</ser>', chart_xml_str, _re.DOTALL))
    for ser_idx in range(min(len(ser_blocks), len(all_series)) - 1, -1, -1):
        label, color, values = all_series[ser_idx]
        ser_m = ser_blocks[ser_idx]
        ser_text = ser_m.group()

        # Mettre à jour numCache
        new_nc = _num_cache(values)
        ser_text = _re.sub(
            r'(<val><numRef><f>[^<]+</f>)<numCache>.*?</numCache>',
            lambda m, nc=new_nc: m.group(1) + nc,
            ser_text, flags=_re.DOTALL)

        if ser_idx < 3:
            # Séries DV2 : mettre à jour couleur + label
            ser_text = _re.sub(
                r'<a:srgbClr val="[0-9A-Fa-f]{6}"',
                f'<a:srgbClr val="{color}"',
                ser_text)
            ser_text = _re.sub(
                r'(<strCache><ptCount val="1"\s*/><pt idx="0"><v>)[^<]+(</v>)',
                rf'\g<1>{label}\2',
                ser_text)
        else:
            # Série Dépenses (idx=3) : zéros + remplissage transparent
            # pour qu'elle soit invisible mais structurellement présente.
            ser_text = _re.sub(
                r'<a:solidFill><a:srgbClr val="[0-9A-Fa-f]{6}"\s*/></a:solidFill>',
                '<a:noFill />',
                ser_text, count=1)
            # Vider le label pour masquer dans la légende
            ser_text = _re.sub(
                r'(<strCache><ptCount val="1"\s*/><pt idx="0"><v>)[^<]+(</v>)',
                rf'\g<1> \2',
                ser_text)

        chart_xml_str = chart_xml_str[:ser_m.start()] + ser_text + chart_xml_str[ser_m.end():]

    # Masquer l'entrée Dépenses dans la légende via <legendEntry>
    # Insérer juste après </plotArea>
    legend_hide = (
        '<legendEntry><idx val="3" /><delete val="1" /></legendEntry>'
    )
    # Insérer dans <legend> juste après <legendPos .../>
    chart_xml_str = _re.sub(
        r'(<legend><legendPos val="[^"]*"\s*/>)',
        rf'\1{legend_hide}',
        chart_xml_str)

    chart_xml = chart_xml_str.encode('utf-8')

    # ── 2. Drawing XML ───────────────────────────────────────────────────────
    # Modeled on original: default namespace, twoCellAnchor, <xfrm />
    start_row_0 = anchor_row - 1   # Excel drawing uses 0-based rows
    end_row_0 = start_row_0 + 18
    drawing_xml = (
        '<wsDr xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
        ' xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
        ' xmlns="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing">'
        '<twoCellAnchor>'
        f'<from><col>1</col><colOff>0</colOff><row>{start_row_0}</row><rowOff>0</rowOff></from>'
        f'<to><col>11</col><colOff>0</colOff><row>{end_row_0}</row><rowOff>0</rowOff></to>'
        '<graphicFrame>'
        '<nvGraphicFramePr>'
        '<cNvPr id="1" name="Chart 1" />'
        '<cNvGraphicFramePr />'
        '</nvGraphicFramePr>'
        '<xfrm />'
        '<a:graphic><a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/chart">'
        '<c:chart r:id="rId1" />'
        '</a:graphicData></a:graphic>'
        '</graphicFrame>'
        '<clientData />'
        '</twoCellAnchor></wsDr>'
    ).encode('utf-8')

    # ── 3. Drawing rels ──────────────────────────────────────────────────────
    # Use chart3/drawing3 to avoid collisions with original TDB's
    # chart1/chart2/drawing1
    drawing_rels = (
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart"'
        ' Target="/xl/charts/chart3.xml" Id="rId1" />'
        '</Relationships>'
    ).encode('utf-8')

    return {
        'xl/charts/chart3.xml': chart_xml,
        'xl/drawings/drawing3.xml': drawing_xml,
        'xl/drawings/_rels/drawing3.xml.rels': drawing_rels,
    }


# ── Point d'entrée ───────────────────────────────────────────────────────────

def write_dashboard_v2(ctx: 'CockpitContext', depenses_path: str) -> None:
    """Écrit (ou réécrit) l'onglet DASHBOARD V2 dans DÉPENSES.xlsm.

    Stratégie ZIP hybride :
      1. Ouvre le VRAI workbook avec openpyxl (keep_vba=True) → les indices de
         style sont ceux du fichier réel.
      2. Crée/réécrit la feuille DASHBOARD V2 avec les styles corrects.
      3. Sauvegarde dans un fichier temporaire.
      4. Extrait le sheet XML et xl/styles.xml du fichier temporaire.
      5. ZIP-injecte ces deux fichiers dans l'ORIGINAL, les autres feuilles
         sont préservées byte-for-byte → cached values intacts.

    Paramètres
    ----------
    ctx            : CockpitContext complet (jamais None)
    depenses_path  : Chemin vers DÉPENSES 2026.xlsm
    """
    # 1. Ouvrir le VRAI workbook (styles en contexte)
    real_wb = openpyxl.load_workbook(depenses_path, keep_vba=True, data_only=False)

    # 2. Supprimer l'ancienne feuille DASHBOARD V2 si elle existe
    if _SHEET_NAME in real_wb.sheetnames:
        del real_wb[_SHEET_NAME]

    # 3. Créer la feuille et la placer en première position
    ws = real_wb.create_sheet(_SHEET_NAME, 0)

    _setup_columns(ws)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 150

    # Nav bar en ligne 1
    from dashboard.nav_bar import write_nav_bar, freeze_below_nav
    write_nav_bar(ws, active_sheet=_SHEET_NAME)

    row = 2
    row = _write_header(ws, ctx, row)
    row = _write_priority(ws, ctx, row)
    row = _write_kpi_strip(ws, ctx, row)
    row = _write_revenus_mois(ws, ctx, row)
    row = _write_chart_revenus_depenses(ws, ctx, row)
    row = _write_depenses_cat(ws, ctx, row)
    row = _write_obligations(ws, ctx, row)
    row = _write_alertes(ws, ctx, row)
    row = _write_footer(ws, row)

    ws.freeze_panes = 'A2'
    ws.print_area = f'A1:L{row}'
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4

    # Récupérer les données de cache du graphique avant de fermer le wb
    chart_cache = getattr(ws, '_chart_data_cache', None)

    # 4. Sauver dans un fichier temporaire
    #    Le temp a la bonne structure MAIS cached values corrompus pour les
    #    autres feuilles. Stratégie : inverser l'injection — partir du temp
    #    et restaurer les sheet XMLs originaux.
    #    Note: pas de ws.add_chart() — le chart sera injecté en raw XML.
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsm')
    os.close(tmp_fd)
    try:
        real_wb.save(tmp_path)
    finally:
        real_wb.close()

    # 4b. Construire les assets raw du graphique (chart XML + drawing + rels)
    chart_assets: dict[str, bytes] = {}
    if chart_cache:
        chart_assets = _build_raw_chart_assets(chart_cache, xlsm_path=depenses_path)

    # 5. Préparer la reconstruction ZIP hybride :
    #    - Temp comme base (structure workbook.xml, rels, styles, chart/drawing DV2)
    #    - Restaurer les sheet XMLs + .rels originaux pour toutes les feuilles
    #      SAUF DASHBOARD V2 (préserve cached values, comments, vml, etc.)
    #    - Ajouter les fichiers originaux absents du temp (comments, vmlDrawings,
    #      printerSettings, media, etc.)
    #    - Fusionner [Content_Types].xml (original + ajouts DV2 du temp)
    out_path = ''
    try:
        sheets_restore: dict[str, bytes] = {}   # {temp_path: original_xml}

        with zipfile.ZipFile(tmp_path, 'r') as tz:
            tmp_wb   = ET.fromstring(tz.read('xl/workbook.xml'))
            tmp_rels = ET.fromstring(tz.read('xl/_rels/workbook.xml.rels'))

        # Pour chaque feuille dans le temp (sauf DV2), mapper vers le XML original.
        for sheet in tmp_wb.iter(f'{{{_XLSX_NS}}}sheet'):
            sname = sheet.get('name')
            if sname == _SHEET_NAME:
                continue
            rid = sheet.get(f'{{{_REL_NS}}}id')
            tp = None
            for rel in tmp_rels.iter(f'{{{_PKG_NS}}}Relationship'):
                if rel.get('Id') == rid:
                    tp = rel.get('Target', '').lstrip('/')
                    if not tp.startswith('xl/'):
                        tp = f'xl/{tp}'
                    break
            if not tp:
                continue
            orig_target = _find_sheet_target(depenses_path, sname)
            if not orig_target:
                continue
            op = orig_target.lstrip('/')
            if not op.startswith('xl/'):
                op = f'xl/{op}'
            with zipfile.ZipFile(depenses_path, 'r') as oz:
                if op in oz.namelist():
                    sheets_restore[tp] = oz.read(op)
                op_dir, op_file = (op.rsplit('/', 1) if '/' in op else ('', op))
                op_rels = f'{op_dir}/_rels/{op_file}.rels'
                if op_rels in oz.namelist():
                    tp_dir, tp_file = (tp.rsplit('/', 1) if '/' in tp else ('', tp))
                    tp_rels = f'{tp_dir}/_rels/{tp_file}.rels'
                    # Pas de remapping nécessaire : openpyxl ne crée plus de
                    # drawings (chart injecté en raw XML avec son propre numéro).
                    sheets_restore[tp_rels] = oz.read(op_rels)

        # Restaurer sharedStrings.xml original + convertir DV2 en inline strings
        orig_shared = None
        with zipfile.ZipFile(depenses_path, 'r') as oz:
            if 'xl/sharedStrings.xml' in oz.namelist():
                orig_shared = oz.read('xl/sharedStrings.xml')

        # Lire la table shared strings du TEMP pour convertir DV2
        tmp_strings: list[str] = []
        with zipfile.ZipFile(tmp_path, 'r') as tz:
            if 'xl/sharedStrings.xml' in tz.namelist():
                _ss = ET.fromstring(tz.read('xl/sharedStrings.xml'))
                for si in _ss.iter(f'{{{_XLSX_NS}}}si'):
                    t_elem = si.find(f'{{{_XLSX_NS}}}t')
                    if t_elem is not None and t_elem.text:
                        tmp_strings.append(t_elem.text)
                    else:
                        parts = []
                        for r_elem in si.findall(f'.//{{{_XLSX_NS}}}t'):
                            if r_elem.text:
                                parts.append(r_elem.text)
                        tmp_strings.append(''.join(parts))

        # Trouver le chemin DV2 dans le temp
        dv2_tmp_path = None
        for sheet in tmp_wb.iter(f'{{{_XLSX_NS}}}sheet'):
            if sheet.get('name') == _SHEET_NAME:
                rid = sheet.get(f'{{{_REL_NS}}}id')
                for rel in tmp_rels.iter(f'{{{_PKG_NS}}}Relationship'):
                    if rel.get('Id') == rid:
                        dv2_tmp_path = rel.get('Target', '').lstrip('/')
                        if not dv2_tmp_path.startswith('xl/'):
                            dv2_tmp_path = f'xl/{dv2_tmp_path}'
                        break
                break

        dv2_inlined: bytes | None = None
        if dv2_tmp_path and tmp_strings:
            with zipfile.ZipFile(tmp_path, 'r') as tz:
                if dv2_tmp_path in tz.namelist():
                    dv2_inlined = _convert_shared_to_inline(
                        tz.read(dv2_tmp_path), tmp_strings)

        # 5b. Si chart, ajouter <drawing r:id="rId1"/> au DV2 sheet XML
        #     et préparer le .rels de la feuille DV2
        dv2_sheet_rels: bytes | None = None
        if chart_assets and dv2_tmp_path:
            # Ajouter la référence drawing dans le sheet XML de DV2
            dv2_xml_src = dv2_inlined or None
            if dv2_xml_src is None:
                with zipfile.ZipFile(tmp_path, 'r') as tz:
                    if dv2_tmp_path in tz.namelist():
                        dv2_xml_src = tz.read(dv2_tmp_path)
            if dv2_xml_src is not None:
                dv2_str = dv2_xml_src.decode('utf-8')
                # Insérer <drawing> juste avant </worksheet>
                if '<drawing ' not in dv2_str and '<drawing>' not in dv2_str:
                    dv2_str = dv2_str.replace(
                        '</worksheet>',
                        '<drawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
                        ' r:id="rId1"/></worksheet>')
                    dv2_inlined = dv2_str.encode('utf-8')

            # .rels pour la feuille DV2 → pointe vers drawing3.xml
            dv2_sheet_rels = (
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
                ' Target="/xl/drawings/drawing3.xml" Id="rId1"/>'
                '</Relationships>'
            ).encode('utf-8')

        # 6. Reconstruire le ZIP : temp + restorations + chart assets + fichiers manquants
        out_fd, out_path = tempfile.mkstemp(suffix='.xlsm')
        os.close(out_fd)

        with zipfile.ZipFile(tmp_path, 'r') as tz, \
             zipfile.ZipFile(depenses_path, 'r') as oz, \
             zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as dst:
            written: set[str] = set()

            # Fusionner Content_Types : original + chart/drawing entries
            orig_ct = oz.read('[Content_Types].xml')
            temp_ct = tz.read('[Content_Types].xml')
            # Start from original, add chart_assets entries
            merged_ct = _merge_content_types(orig_ct, temp_ct)
            # Ensure chart + drawing entries exist
            merged_ct_str = merged_ct.decode('utf-8')
            _chart_ct_additions = []
            if chart_assets:
                if '/xl/charts/chart3.xml' not in merged_ct_str:
                    _chart_ct_additions.append(
                        '<Override PartName="/xl/charts/chart3.xml"'
                        ' ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml" />')
                if '/xl/drawings/drawing3.xml' not in merged_ct_str:
                    _chart_ct_additions.append(
                        '<Override PartName="/xl/drawings/drawing3.xml"'
                        ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml" />')
            if _chart_ct_additions:
                merged_ct_str = merged_ct_str.replace(
                    '</Types>', ''.join(_chart_ct_additions) + '</Types>')
                merged_ct = merged_ct_str.encode('utf-8')

            # A — Entrées du temp (avec restauration des sheets originaux)
            for item in tz.infolist():
                fn = item.filename
                if fn == '[Content_Types].xml':
                    dst.writestr(item, merged_ct)
                elif fn in sheets_restore:
                    dst.writestr(item, sheets_restore[fn])
                elif fn == 'xl/sharedStrings.xml' and orig_shared:
                    dst.writestr(item, orig_shared)
                elif fn == dv2_tmp_path and dv2_inlined:
                    dst.writestr(item, dv2_inlined)
                elif fn in chart_assets:
                    # Will be written from chart_assets instead
                    dst.writestr(item, chart_assets[fn])
                else:
                    dst.writestr(item, tz.read(fn))
                written.add(fn)

            # A2 — Inject chart assets not already in temp
            for ca_path, ca_data in chart_assets.items():
                if ca_path not in written:
                    dst.writestr(ca_path, ca_data)
                    written.add(ca_path)

            # A3 — DV2 sheet .rels (drawing reference)
            if dv2_sheet_rels and dv2_tmp_path:
                dv2_dir, dv2_file = dv2_tmp_path.rsplit('/', 1)
                dv2_rels_path = f'{dv2_dir}/_rels/{dv2_file}.rels'
                if dv2_rels_path not in written:
                    dst.writestr(dv2_rels_path, dv2_sheet_rels)
                    written.add(dv2_rels_path)

            # B — Fichiers originaux absents du temp (comments, vmlDrawings,
            #     printerSettings, media, etc.)
            for item in oz.infolist():
                if item.filename not in written:
                    dst.writestr(item, oz.read(item.filename))
                    written.add(item.filename)

            # C — sharedStrings fallback
            if orig_shared and 'xl/sharedStrings.xml' not in written:
                dst.writestr('xl/sharedStrings.xml', orig_shared)

        # S'assurer que [Content_Types].xml référence sharedStrings
        if orig_shared:
            _ensure_content_type_for_shared_strings(out_path)

        shutil.move(out_path, depenses_path)

    except Exception:
        if out_path and os.path.exists(out_path):
            os.unlink(out_path)
        raise
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
