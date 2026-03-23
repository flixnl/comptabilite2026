#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""dashboard/apply_nav_bar.py
=========================================================
Applique le bandeau de navigation sur les feuilles statiques
du fichier DÉPENSES 2026.xlsm.

Feuilles traitées :
  - ESTIMATION FISCALE → overlay ligne 1 (vide)
  - ACOMPTES          → insert_rows(1) puis nav
  - TAXES MARS        → insert_rows(1) puis nav

DASHBOARD V2 est géré par write_dashboard_v2.py (nav intégrée).
TABLEAU DE BORD est masqué (supplanté par DASHBOARD V2).
OBLIGATIONS est une feuille backend → pas de nav.

Aussi :
  - Masque l'onglet TABLEAU DE BORD
  - Applique zoom 150 % sur toutes les feuilles
  - Gèle la ligne d'en-tête sur les feuilles de dépenses mensuelles

Usage :
    python3 -m dashboard.apply_nav_bar <depenses_path>
"""
from __future__ import annotations

import os
import sys
import shutil
import tempfile
import zipfile
from xml.etree import ElementTree as ET

import openpyxl
from dashboard.nav_bar import write_nav_bar, NAV_ROW


# ── Feuilles à traiter ──────────────────────────────────────────────────────

# (sheet_name, method, total_cols)
#   method = 'overlay'  → écrire sur la ligne 1 existante
#   method = 'insert'   → insert_rows(1) d'abord
_TARGETS: list[tuple[str, str, int]] = [
    ('ESTIMATION FISCALE', 'overlay', 8),
    ('ACOMPTES',           'insert',  7),
    ('TAXES MARS',         'insert',  4),
]

# Feuilles de dépenses mensuelles → freeze row 1 (en-tête)
_MOIS_NOMS = ['JANVIER', 'FÉVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN',
              'JUILLET', 'AOÛT', 'SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DÉCEMBRE']
_MONTHLY_SHEETS = [f'{m} 2026' for m in _MOIS_NOMS]
# Feuilles taxes mensuelles aussi
_TAX_SHEETS = [f'TAXES {m}' for m in _MOIS_NOMS]


def apply_nav_bars(depenses_path: str, *, verbose: bool = True) -> list[str]:
    """Applique la nav bar sur les feuilles statiques + ajustements globaux.

    Ajustements globaux :
    - Masque TABLEAU DE BORD
    - Zoom 150 % sur toutes les feuilles
    - Freeze row 1 sur les feuilles de dépenses mensuelles

    Utilise la stratégie « reverse injection » identique à write_dashboard_v2 :
    1. Ouvrir le vrai workbook (keep_vba=True)
    2. Modifier les feuilles cibles (nav bar + ajustements)
    3. Sauver dans un temp
    4. Restaurer les sheet XMLs des feuilles NON modifiées depuis l'original
    5. Restaurer les assets DASHBOARD V2 (drawings/charts) depuis l'original

    Returns
    -------
    list[str] : noms des feuilles modifiées avec succès
    """
    _XLSX_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    _REL_NS  = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    _PKG_NS  = 'http://schemas.openxmlformats.org/package/2006/relationships'

    wb = openpyxl.load_workbook(depenses_path, keep_vba=True, data_only=False)

    modified: list[str] = []

    # ── Nav bar sur les feuilles cibles ──────────────────────────────────────
    _NAV_MARKERS = {'Dashboard', 'Obligations', 'Acomptes', 'Estimation', 'Taxes',
                    '📊', '💰', '📅', '🧮', '🧾'}

    def _is_nav_content(val) -> bool:
        """Détecte si une valeur de cellule est du contenu nav bar."""
        if val is None:
            return False
        s = str(val)
        return any(m in s for m in _NAV_MARKERS)

    for sheet_name, method, total_cols in _TARGETS:
        if sheet_name not in wb.sheetnames:
            if verbose:
                print(f'  ⚠ {sheet_name} introuvable — skip')
            continue

        ws = wb[sheet_name]

        if method == 'insert':
            # Vérifier si la nav bar existe déjà (runs précédents)
            b1_val = ws.cell(row=1, column=2).value
            nav_already_exists = _is_nav_content(b1_val)

            if nav_already_exists:
                # Trouver et supprimer les lignes dupliquées (remnants)
                junk_end = 1
                for r in range(2, 20):
                    # Une ligne est « junk » si elle contient du texte nav
                    row_has_nav = False
                    for c in range(1, max(total_cols + 2, 8)):
                        if _is_nav_content(ws.cell(row=r, column=c).value):
                            row_has_nav = True
                            break
                    if row_has_nav:
                        junk_end = r
                    else:
                        break
                if junk_end > 1:
                    ws.delete_rows(2, junk_end - 1)
                    if verbose:
                        print(f'  🧹 {sheet_name} — {junk_end - 1} ligne(s) dupliquée(s) supprimée(s)')
                # Pas d'insert_rows, juste overlay row 1
            else:
                # Premier run : insérer une ligne pour la nav
                from openpyxl.utils import get_column_letter
                row1_merges = [(m.min_col, m.max_col, m.min_row, m.max_row)
                               for m in list(ws.merged_cells.ranges)
                               if m.min_row <= 1]
                for mc_min_col, mc_max_col, mc_min_row, mc_max_row in row1_merges:
                    old_range = (f'{get_column_letter(mc_min_col)}{mc_min_row}:'
                                 f'{get_column_letter(mc_max_col)}{mc_max_row}')
                    ws.unmerge_cells(old_range)

                ws.insert_rows(1, amount=1)

                for mc_min_col, mc_max_col, mc_min_row, mc_max_row in row1_merges:
                    new_range = (f'{get_column_letter(mc_min_col)}{mc_min_row + 1}:'
                                 f'{get_column_letter(mc_max_col)}{mc_max_row + 1}')
                    ws.merge_cells(new_range)

        write_nav_bar(ws, active_sheet=sheet_name, total_cols=total_cols)

        existing_freeze = ws.freeze_panes
        if not existing_freeze or existing_freeze == 'A1':
            ws.freeze_panes = 'A2'

        # Fix TAXES sheets: ensure summary rows are readable (height ≥ 20)
        # and column A is wide enough for section titles
        if sheet_name.startswith('TAXES'):
            for r in range(1, min(ws.max_row + 1, 40)):
                h = ws.row_dimensions[r].height
                if h is not None and h < 10:
                    ws.row_dimensions[r].height = 22
            # Ensure col A is wide enough for TVQ section title
            if ws.column_dimensions['A'].width is not None and ws.column_dimensions['A'].width < 20:
                ws.column_dimensions['A'].width = 32

        modified.append(sheet_name)
        if verbose:
            print(f'  ✅ {sheet_name} ({method})')

    # ── Masquer TABLEAU DE BORD ──────────────────────────────────────────────
    if 'TABLEAU DE BORD' in wb.sheetnames:
        wb['TABLEAU DE BORD'].sheet_state = 'hidden'
        modified.append('TABLEAU DE BORD')
        if verbose:
            print('  ✅ TABLEAU DE BORD masqué')

    # ── Zoom 150 % sur TOUTES les feuilles ──────────────────────────────────
    for sn in wb.sheetnames:
        ws = wb[sn]
        ws.sheet_view.zoomScale = 150
        if sn not in modified:
            modified.append(sn)

    # ── Freeze row 1 sur les feuilles de dépenses mensuelles ─────────────────
    for sn in _MONTHLY_SHEETS + _TAX_SHEETS:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        # Freeze la ligne d'en-tête (row 1)
        ws.freeze_panes = 'A2'
        if sn not in modified:
            modified.append(sn)
        if verbose and sn in _MONTHLY_SHEETS:
            print(f'  ✅ {sn} — freeze row 1')

    if not modified:
        wb.close()
        return modified

    # Sauver dans un temp
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsm')
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
    finally:
        wb.close()

    # ── Collecter les assets DV2 depuis l'original (drawings, charts) ────────
    #    openpyxl peut corrompre les charts lors du re-save, donc on restaure
    #    le sheet XML + assets DV2 byte-for-byte depuis le fichier d'entrée.
    dv2_assets: dict[str, bytes] = {}   # {zip_path: original_bytes}
    dv2_sheet_path: str | None = None
    try:
        dv2_sheet_path = _find_sheet_target_in_zip(
            depenses_path, 'DASHBOARD V2', _XLSX_NS, _REL_NS, _PKG_NS)
        if dv2_sheet_path:
            with zipfile.ZipFile(depenses_path, 'r') as oz:
                # Restaurer le sheet XML de DV2
                if dv2_sheet_path in oz.namelist():
                    dv2_assets[dv2_sheet_path] = oz.read(dv2_sheet_path)
                # Collecter les rels + drawings + charts de DV2
                sheet_dir, sheet_file = (dv2_sheet_path.rsplit('/', 1)
                                          if '/' in dv2_sheet_path else ('', dv2_sheet_path))
                rels_path = f'{sheet_dir}/_rels/{sheet_file}.rels'
                if rels_path in oz.namelist():
                    dv2_assets[rels_path] = oz.read(rels_path)
                    # Récursif : drawings → charts
                    _collect_dv2_assets(oz, rels_path, dv2_sheet_path, dv2_assets)
    except Exception:
        pass  # Non-fatal : on perd juste la restauration DV2

    # Reverse injection : restaurer les feuilles NON modifiées par la nav bar
    try:
        # Les feuilles qu'on a modifiées via openpyxl (nav bar, freeze, zoom)
        # doivent rester dans leur version temp. MAIS DV2 et les feuilles de
        # données non touchées doivent être restaurées depuis l'original.
        nav_modified_set = set(_TARGETS_NAMES())
        nav_modified_set.add('TABLEAU DE BORD')   # on a changé sheet_state
        # Les monthly+tax sheets ont juste zoom+freeze → modifications XML légères
        # qu'openpyxl gère bien. Pas besoin de les restaurer.

        sheets_restore: dict[str, bytes] = {}

        with zipfile.ZipFile(tmp_path, 'r') as tz:
            tmp_wb = ET.fromstring(tz.read('xl/workbook.xml'))
            tmp_rels = ET.fromstring(tz.read('xl/_rels/workbook.xml.rels'))

        for sheet in tmp_wb.iter(f'{{{_XLSX_NS}}}sheet'):
            sname = sheet.get('name')
            # Restaurer depuis l'original : feuilles NON modifiées du tout
            # (pas de nav, pas de zoom seulement, etc.)
            # En pratique : on restaure uniquement les feuilles de données pures
            # qui n'ont reçu aucune modification. Toutes les feuilles ont reçu
            # le zoom, donc on ne restaure que les assets DV2 (ci-dessous).
            # Pour les cached values : on restaure les feuilles qui n'ont reçu
            # QUE le zoom (pas de nav bar, pas de freeze).
            # Le zoom est stocké dans sheetViews, pas dans les données.
            # openpyxl peut corrompre les cached values → restaurer les sheet
            # XMLs puis re-injecter le zoom via XML patching.
            pass

        # Stratégie simplifiée : restaurer TOUTES les sheet XMLs depuis
        # l'original SAUF celles qui ont reçu la nav bar (insert/overlay).
        # Pour le zoom, on le patche directement dans le XML restauré.
        nav_only = {t[0] for t in _TARGETS}
        nav_only.add('TABLEAU DE BORD')

        for sheet in tmp_wb.iter(f'{{{_XLSX_NS}}}sheet'):
            sname = sheet.get('name')
            if sname == 'DASHBOARD V2':
                continue  # DV2 : restauré via dv2_assets ci-dessous
            rid = sheet.get(f'{{{_REL_NS}}}id')
            tp = None
            for rel in tmp_rels.iter(f'{{{_PKG_NS}}}Relationship'):
                if rel.get('Id') == rid:
                    tp = rel.get('Target', '').lstrip('/')
                    if not tp.startswith('xl/'):
                        tp = f'xl/{tp}'
                    break
            if not tp:
                continue
            orig_path = _find_sheet_target_in_zip(depenses_path, sname,
                                                   _XLSX_NS, _REL_NS, _PKG_NS)
            if not orig_path:
                continue
            with zipfile.ZipFile(depenses_path, 'r') as oz:
                if sname not in nav_only:
                    # Feuilles hors nav : restaurer le XML + zoom/freeze
                    if orig_path in oz.namelist():
                        orig_xml = oz.read(orig_path)
                        need_freeze = (sname in _MONTHLY_SHEETS or sname in _TAX_SHEETS)
                        patched = _patch_zoom_in_xml(
                            orig_xml, 150,
                            freeze_row=1 if need_freeze else None)
                        sheets_restore[tp] = patched

                # Restaurer le .rels pour TOUTES les feuilles (sauf DV2)
                # depuis l'input. openpyxl renumérote les drawings lors du
                # re-save → les .rels du temp sont incorrects.
                op_dir, op_file = (orig_path.rsplit('/', 1) if '/' in orig_path
                                   else ('', orig_path))
                op_rels = f'{op_dir}/_rels/{op_file}.rels'
                if op_rels in oz.namelist():
                    tp_dir, tp_file = (tp.rsplit('/', 1) if '/' in tp
                                      else ('', tp))
                    tp_rels = f'{tp_dir}/_rels/{tp_file}.rels'
                    sheets_restore[tp_rels] = oz.read(op_rels)

        # Restaurer sharedStrings
        orig_shared = None
        with zipfile.ZipFile(depenses_path, 'r') as oz:
            if 'xl/sharedStrings.xml' in oz.namelist():
                orig_shared = oz.read('xl/sharedStrings.xml')

        # Reconstruire le ZIP : temp + restorations + fichiers manquants
        out_fd, out_path = tempfile.mkstemp(suffix='.xlsm')
        os.close(out_fd)

        with zipfile.ZipFile(tmp_path, 'r') as tz, \
             zipfile.ZipFile(depenses_path, 'r') as oz, \
             zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as dst:
            written: set[str] = set()

            # Fusionner Content_Types (original = comments/vml, temp = DV2 assets)
            merged_ct = _merge_content_types(
                oz.read('[Content_Types].xml'),
                tz.read('[Content_Types].xml'))

            # A — Entrées du temp (avec restauration des sheets originaux + DV2)
            #     Pour les drawings/charts, IGNORER la version temp (openpyxl
            #     renumérote et corrompt les charts). Ils seront écrits depuis
            #     l'input en étape B ou depuis dv2_assets en étape C.
            def _is_draw_chart(p: str) -> bool:
                return ((p.startswith('xl/drawings/') or p.startswith('xl/charts/'))
                        and not p.endswith('.vml'))

            for item in tz.infolist():
                fn = item.filename
                if fn == '[Content_Types].xml':
                    dst.writestr(item, merged_ct)
                    written.add(fn)
                elif fn in sheets_restore:
                    dst.writestr(item, sheets_restore[fn])
                    written.add(fn)
                elif fn == 'xl/sharedStrings.xml' and orig_shared:
                    dst.writestr(item, orig_shared)
                    written.add(fn)
                elif fn in dv2_assets:
                    dst.writestr(item, dv2_assets[fn])
                    written.add(fn)
                elif _is_draw_chart(fn):
                    # Ne PAS écrire la version temp — sera restauré depuis input
                    # Ne PAS ajouter à written → step B écrira la version input
                    pass
                else:
                    dst.writestr(item, tz.read(fn))
                    written.add(fn)

            # B — Fichiers originaux absents du temp (comments, vmlDrawings,
            #     printerSettings, media, etc.)
            for item in oz.infolist():
                if item.filename not in written:
                    dst.writestr(item, oz.read(item.filename))
                    written.add(item.filename)

            # C — Assets DV2 absents du temp ET de l'original
            for path, data in dv2_assets.items():
                if path not in written:
                    dst.writestr(path, data)
                    written.add(path)

            # D — sharedStrings fallback
            if orig_shared and 'xl/sharedStrings.xml' not in written:
                dst.writestr('xl/sharedStrings.xml', orig_shared)

        # Nettoyer [Content_Types].xml :
        #   1. Ajouter sharedStrings si absent
        #   2. Supprimer les Override pour des fichiers inexistants
        #      (ex: drawing2.xml créé par openpyxl mais non conservé)
        import re as _re_ct
        _ct_needs_rewrite = False
        with zipfile.ZipFile(out_path, 'r') as _zcheck:
            _ct = _zcheck.read('[Content_Types].xml').decode('utf-8')
            _existing = set(_zcheck.namelist())

        if orig_shared and 'sharedStrings' not in _ct:
            _ct = _ct.replace('</Types>',
                '<Override PartName="/xl/sharedStrings.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sharedStrings+xml"/></Types>')
            _ct_needs_rewrite = True

        # Supprimer les Override pointant vers des fichiers absents
        def _ct_filter(m):
            nonlocal _ct_needs_rewrite
            pn = _re_ct.search(r'PartName="(/[^"]+)"', m.group())
            if pn:
                path = pn.group(1).lstrip('/')
                if path not in _existing:
                    _ct_needs_rewrite = True
                    return ''
            return m.group()
        _ct = _re_ct.sub(r'<Override\s+[^>]+/>', _ct_filter, _ct)

        if _ct_needs_rewrite:
            _ct_tmp = out_path + '.ct'
            with zipfile.ZipFile(out_path, 'r') as _src, \
                 zipfile.ZipFile(_ct_tmp, 'w', zipfile.ZIP_DEFLATED) as _dst:
                for _it in _src.infolist():
                    if _it.filename == '[Content_Types].xml':
                        _dst.writestr(_it, _ct.encode('utf-8'))
                    else:
                        _dst.writestr(_it, _src.read(_it.filename))
            os.replace(_ct_tmp, out_path)

        shutil.move(out_path, depenses_path)

    except Exception:
        if 'out_path' in dir() and os.path.exists(out_path):
            os.unlink(out_path)
        raise
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    return modified


def _TARGETS_NAMES() -> list[str]:
    """Noms des feuilles cibles de la nav bar."""
    return [t[0] for t in _TARGETS]


def _remap_drawing_in_rels(orig_rels: bytes, temp_rels: bytes) -> bytes:
    """Utilise les .rels originaux mais remappe les Targets drawing vers le temp.

    openpyxl peut réassigner les numéros de drawing quand il ajoute DV2.
    Préserve les références ancillaires (vmlDrawing, comments, etc.) de l'original.
    """
    import re as _re

    temp_str = temp_rels.decode('utf-8')
    temp_draw = None
    for m in _re.finditer(r'<Relationship\s+[^>]*/>', temp_str):
        elem = m.group()
        if 'relationships/drawing"' in elem and 'vmlDrawing' not in elem:
            t = _re.search(r'Target="([^"]+)"', elem)
            if t:
                temp_draw = t.group(1)
                break

    if not temp_draw:
        return orig_rels

    orig_str = orig_rels.decode('utf-8')

    def _replace(m):
        elem = m.group()
        if 'relationships/drawing"' in elem and 'vmlDrawing' not in elem:
            return _re.sub(r'Target="[^"]+"', f'Target="{temp_draw}"', elem)
        return elem

    result = _re.sub(r'<Relationship\s+[^>]*/>', _replace, orig_str)
    return result.encode('utf-8')


def _merge_content_types(orig_ct: bytes, temp_ct: bytes) -> bytes:
    """Fusionne [Content_Types].xml : original comme base + ajouts du temp.

    L'original contient les entrées pour comments, vmlDrawings, etc.
    Le temp contient les entrées pour chart, drawing (DV2).
    Le résultat a les deux.
    """
    import re as _re
    orig = orig_ct.decode('utf-8')
    temp = temp_ct.decode('utf-8')

    orig_parts = set(_re.findall(r'PartName="([^"]+)"', orig))
    orig_exts = set(_re.findall(r'<Default[^>]+Extension="([^"]+)"', orig))

    additions = []
    for m in _re.finditer(r'<Override\s+[^>]+/>', temp):
        part_m = _re.search(r'PartName="([^"]+)"', m.group())
        if part_m and part_m.group(1) not in orig_parts:
            additions.append(m.group())
    for m in _re.finditer(r'<Default\s+[^>]+/>', temp):
        ext_m = _re.search(r'Extension="([^"]+)"', m.group())
        if ext_m and ext_m.group(1) not in orig_exts:
            additions.append(m.group())

    if additions:
        orig = orig.replace('</Types>', ''.join(additions) + '</Types>')

    return orig.encode('utf-8')


def _collect_dv2_assets(z: zipfile.ZipFile, rels_path: str,
                         parent_path: str, assets: dict) -> None:
    """Collecte récursivement drawings/charts référencés par un .rels."""
    rels_xml = ET.fromstring(z.read(rels_path))
    for rel in rels_xml:
        tag = rel.tag
        if not (tag.endswith('}Relationship') or tag == 'Relationship'):
            continue
        target = rel.get('Target', '')
        if not target or target.startswith('http'):
            continue
        # Résoudre le chemin relatif
        if target.startswith('/'):
            resolved = target.lstrip('/')
        else:
            parent_dir = parent_path.rsplit('/', 1)[0] if '/' in parent_path else ''
            if target.startswith('../'):
                resolved = os.path.normpath(os.path.join(parent_dir, target)).replace('\\', '/')
            else:
                resolved = f'{parent_dir}/{target}' if parent_dir else target
        if resolved not in z.namelist() or resolved in assets:
            continue
        assets[resolved] = z.read(resolved)
        # Sous-rels (ex: drawing → chart)
        r_dir, r_file = (resolved.rsplit('/', 1) if '/' in resolved else ('', resolved))
        sub_rels = f'{r_dir}/_rels/{r_file}.rels'
        if sub_rels in z.namelist() and sub_rels not in assets:
            assets[sub_rels] = z.read(sub_rels)
            _collect_dv2_assets(z, sub_rels, resolved, assets)


def _patch_zoom_in_xml(sheet_xml: bytes, zoom: int,
                       freeze_row: int | None = None) -> bytes:
    """Patche le zoomScale (et optionnellement un freeze pane) dans un sheet XML.

    Parameters
    ----------
    sheet_xml   : bytes du sheet XML original
    zoom        : valeur du zoom (ex: 150)
    freeze_row  : si fourni, ajoute un freeze pane horizontal après cette ligne
                  (ex: 1 → freeze_panes = 'A2')
    """
    _SS_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    try:
        tree = ET.fromstring(sheet_xml)
        for sv in tree.iter(f'{{{_SS_NS}}}sheetView'):
            sv.set('zoomScale', str(zoom))
            sv.set('zoomScaleNormal', str(zoom))

            if freeze_row is not None:
                # Supprimer tout <pane> existant
                for old_pane in list(sv.findall(f'{{{_SS_NS}}}pane')):
                    sv.remove(old_pane)
                # Ajouter le freeze pane
                pane = ET.SubElement(sv, f'{{{_SS_NS}}}pane')
                pane.set('ySplit', str(freeze_row))
                pane.set('topLeftCell', f'A{freeze_row + 1}')
                pane.set('activePane', 'bottomLeft')
                pane.set('state', 'frozen')
                # Placer <pane> AVANT les <selection> (requis par le schéma)
                children = list(sv)
                sv.remove(pane)
                # Insérer avant le premier <selection> ou à la fin
                inserted = False
                for i, child in enumerate(children):
                    if child.tag.endswith('}selection') or child.tag == 'selection':
                        sv.insert(i, pane)
                        inserted = True
                        break
                if not inserted:
                    sv.append(pane)
                # Ajouter/modifier la selection pour le bottomLeft pane
                has_bl = False
                for sel in sv.findall(f'{{{_SS_NS}}}selection'):
                    if sel.get('pane') == 'bottomLeft':
                        has_bl = True
                        break
                if not has_bl:
                    sel = ET.SubElement(sv, f'{{{_SS_NS}}}selection')
                    sel.set('pane', 'bottomLeft')

        # Re-sérialiser en préservant les namespaces
        ET.register_namespace('', _SS_NS)
        ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
        ET.register_namespace('mc', 'http://schemas.openxmlformats.org/markup-compatibility/2006')
        ET.register_namespace('x14ac', 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac')
        return ET.tostring(tree, xml_declaration=True, encoding='UTF-8')
    except Exception:
        return sheet_xml  # fallback : retourner l'original si le patch échoue


def _find_sheet_target_in_zip(xlsm_path, sheet_name, xlsx_ns, rel_ns, pkg_ns):
    """Trouve le chemin XML d'une feuille dans le .xlsm."""
    with zipfile.ZipFile(xlsm_path, 'r') as z:
        wb_xml = ET.fromstring(z.read('xl/workbook.xml'))
        rels_xml = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
        for s in wb_xml.iter(f'{{{xlsx_ns}}}sheet'):
            if s.get('name') == sheet_name:
                rid = s.get(f'{{{rel_ns}}}id')
                for r in rels_xml.iter(f'{{{pkg_ns}}}Relationship'):
                    if r.get('Id') == rid:
                        t = r.get('Target', '').lstrip('/')
                        if not t.startswith('xl/'):
                            t = f'xl/{t}'
                        return t
    return None


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python3 -m dashboard.apply_nav_bar <depenses_path>')
        sys.exit(1)
    path = sys.argv[1]
    print(f'Applying nav bars to {os.path.basename(path)}...')
    done = apply_nav_bars(path)
    print(f'Done — {len(done)} sheets updated: {", ".join(done)}')
