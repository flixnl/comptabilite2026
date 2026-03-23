#!/usr/bin/env python3
"""
dashboard/system_snapshot.py — Snapshot complet du système comptable (v1.1).

Capture l'état de toutes les composantes critiques à un instant donné
et écrit un fichier JSON lisible, indenté et stable.

Usage
─────
    from dashboard.system_snapshot import build_snapshot, write_snapshot
    snap = build_snapshot(ctx, outils_dir)
    write_snapshot(snap, outils_dir)

Intégration
───────────
Appelé automatiquement à la fin de scan_pipeline.run(), après les health
checks et avant le git push.

Garanties
─────────
- Aucun crash : chaque section est isolée dans un try/except
- Fallback systématique : valeurs par défaut si données manquantes
- JSON sérialisable : dates → ISO 8601, floats arrondis à 2 décimales
- Validation pré-écriture : valeurs invalides → null + warning

Historique
──────────
v1.0  2026-03-22  Snapshot initial (10 sections)
v1.1  2026-03-22  +meta, +data_freshness, +internal_consistency,
                  +pipeline_stats, git amélioré, validation pré-écriture,
                  obligations résumé dérivé de all_obligations
"""
from __future__ import annotations

import json
import os
import platform
import socket
import sys
import time
from datetime import date, datetime
from typing import Any, Optional, TYPE_CHECKING

if TYPE_CHECKING:
    from dashboard.context import CockpitContext

_SNAPSHOT_VERSION = '1.1'
_SCHEMA_VERSION   = 2


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _safe(fn, default=None):
    """Exécute fn() et retourne default si exception."""
    try:
        return fn()
    except Exception:
        return default


def _round2(val) -> float:
    """Arrondit à 2 décimales, retourne 0.0 si non numérique."""
    try:
        return round(float(val), 2)
    except (TypeError, ValueError):
        return 0.0


def _isoformat(d) -> str:
    """Convertit une date/datetime en ISO 8601, '' si None."""
    if d is None:
        return ''
    if hasattr(d, 'isoformat'):
        return d.isoformat()
    return str(d)


# ─── Section builders ────────────────────────────────────────────────────────

def _build_meta_section(start_time: float) -> dict:
    """Métadonnées système et diagnostic (Section 2)."""
    elapsed_ms = round((time.time() - start_time) * 1000)
    return {
        'snapshot_version': _SNAPSHOT_VERSION,
        'schema_version': _SCHEMA_VERSION,
        'python_version': platform.python_version(),
        'host': socket.gethostname(),
        'os': f'{platform.system()} {platform.release()}',
        'pipeline_duration_ms': elapsed_ms,
    }


def _build_pipeline_section(
    ctx: Optional['CockpitContext'],
    scan_ok: bool,
    scan_time: str,
) -> dict:
    """Pipeline metadata."""
    return {
        'scan_status': 'OK' if scan_ok else 'ERROR',
        'last_scan_time': scan_time,
        'imports_count': 0,  # rempli par l'appelant si disponible
    }


def _build_data_freshness_section(scan_time: str) -> dict:
    """Fraîcheur des données (Section 3)."""
    now = datetime.now()
    section = {
        'last_scan_time': scan_time,
        'last_dashboard_build': '',
        'last_snapshot_build': now.strftime('%Y-%m-%d %H:%M:%S'),
        'minutes_since_last_scan': 0,
    }

    # Lire le timestamp du dernier dashboard HTML
    def _read_dashboard_mtime():
        from core.paths import resolve_main_paths
        paths = resolve_main_paths()
        outils = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        html_path = os.path.join(outils, 'Tableau de bord comptable.html')
        if os.path.exists(html_path):
            mtime = os.path.getmtime(html_path)
            section['last_dashboard_build'] = datetime.fromtimestamp(mtime).strftime(
                '%Y-%m-%d %H:%M:%S')

    _safe(_read_dashboard_mtime)

    # Calculer minutes depuis le dernier scan
    try:
        scan_dt = datetime.strptime(scan_time, '%Y-%m-%d %H:%M:%S')
        delta = (now - scan_dt).total_seconds() / 60
        section['minutes_since_last_scan'] = round(max(0, delta), 1)
    except (ValueError, TypeError):
        pass

    return section


def _build_taxes_section(ctx: Optional['CockpitContext']) -> dict:
    """Taxes : montant mensuel, cellules Excel, position Python."""
    section = {
        'current_month_label': '',
        'current_month_amount': 0.0,
        'excel_cell_C23': 0.0,
        'excel_cell_H10': 0.0,
        'python_tax_position': 0.0,
    }
    if ctx is None:
        return section

    section['current_month_label'] = ctx.taxes_current_label or ''
    section['current_month_amount'] = _round2(ctx.taxes_current_month)

    # TaxesPosition (cumul annuel — source de vérité analytics)
    ac = getattr(ctx, 'alert_ctx', None)
    tp = getattr(ac, 'taxes_position', None) if ac else None
    if tp is not None:
        section['python_tax_position'] = _round2(tp.total_restantes)
        section['tps_restantes'] = _round2(tp.tps_restantes)
        section['tvq_restantes'] = _round2(tp.tvq_restantes)
        section['tps_collectees'] = _round2(tp.tps_collectees)
        section['tvq_collectees'] = _round2(tp.tvq_collectees)
        section['cti'] = _round2(tp.cti)
        section['rti'] = _round2(tp.rti)

    # Lire C23 et H10 depuis le fichier Excel (data_only)
    def _read_excel_cells():
        from core.paths import resolve_main_paths
        import openpyxl
        paths = resolve_main_paths()
        wb = openpyxl.load_workbook(paths.depenses, data_only=True, read_only=True)
        try:
            mois_names = [
                '', 'TAXES JANVIER', 'TAXES FÉVRIER', 'TAXES MARS',
                'TAXES AVRIL', 'TAXES MAI', 'TAXES JUIN',
                'TAXES JUILLET', 'TAXES AOÛT', 'TAXES SEPTEMBRE',
                'TAXES OCTOBRE', 'TAXES NOVEMBRE', 'TAXES DÉCEMBRE',
            ]
            ref = ctx.reference_date or date.today()
            mois_idx = ref.month
            sheet_name = mois_names[mois_idx] if mois_idx < len(mois_names) else ''
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                val = ws.cell(row=23, column=3).value
                if val is not None:
                    section['excel_cell_C23'] = _round2(val)

            if 'TABLEAU DE BORD' in wb.sheetnames:
                ws_tdb = wb['TABLEAU DE BORD']
                h10 = ws_tdb.cell(row=10, column=8).value
                if h10 is not None:
                    section['excel_cell_H10'] = _round2(h10)
        finally:
            wb.close()

    _safe(_read_excel_cells)
    return section


def _build_revenue_section(ctx: Optional['CockpitContext']) -> dict:
    """Revenus : mois courant et YTD."""
    section = {'current_month': 0.0, 'ytd_total': 0.0}
    if ctx is None or not ctx.monthly_revenue:
        return section

    ref = ctx.reference_date or date.today()
    ytd = 0.0
    current = 0.0
    for m in ctx.monthly_revenue:
        rev = (m.get('freelance', 0) or 0) + (m.get('salaire', 0) or 0)
        ytd += rev
        if m.get('mois') == ref.month:
            current = rev
    section['current_month'] = _round2(current)
    section['ytd_total'] = _round2(ytd)
    return section


def _build_expenses_section(ctx: Optional['CockpitContext']) -> dict:
    """Dépenses : mois courant et YTD."""
    section = {'current_month': 0.0, 'ytd_total': 0.0}
    if ctx is None or not ctx.expenses_by_month:
        return section

    ref = ctx.reference_date or date.today()
    section['ytd_total'] = _round2(sum(ctx.expenses_by_month.values()))
    section['current_month'] = _round2(ctx.expenses_by_month.get(ref.month, 0))
    return section


def _build_ledger_section() -> dict:
    """Cash ledger : total, entrées, par obligation."""
    section = {
        'cash_ledger_total': 0.0,
        'ledger_entries': [],
        'per_obligation': {},
    }

    def _load():
        from fiscal.cash_ledger import load_ledger
        entries = [e for e in load_ledger() if e.is_active]
        section['cash_ledger_total'] = _round2(sum(e.amount for e in entries))
        section['ledger_entries'] = [
            {
                'id': e.id,
                'date': _isoformat(e.date),
                'amount': _round2(e.amount),
                'type': e.type,
                'obligation_id': e.obligation_id,
                'authority': e.authority,
                'status': e.status,
            }
            for e in entries
        ]
        by_obl: dict[str, float] = {}
        for e in entries:
            oid = e.obligation_id or '_sans'
            by_obl[oid] = by_obl.get(oid, 0.0) + e.amount
        section['per_obligation'] = {k: _round2(v) for k, v in by_obl.items()}

    _safe(_load)
    return section


def _build_obligations_section(ctx: Optional['CockpitContext']) -> dict:
    """Obligations : paid/total par obligation connue.

    Section 1 fix : le résumé (nantel_paid, tax_can_paid, etc.) est
    dérivé APRÈS construction de all_obligations, garantissant la cohérence.
    """
    section: dict = {
        'nantel_paid': 0.0, 'nantel_total': 0.0,
        'tax_can_paid': 0.0, 'tax_can_total': 0.0,
        'tax_qc_paid': 0.0, 'tax_qc_total': 0.0,
        'grand_total': 0.0, 'grand_paid': 0.0, 'grand_solde': 0.0,
        'all_obligations': [],
    }
    if ctx is None or not ctx.obligations:
        return section

    # 1) Construire all_obligations
    all_obligations = []
    for o in ctx.obligations:
        oid = (o.get('id', '') or o.get('label', '')).lower()
        paid = _round2(o.get('paid', 0))
        total = _round2(o.get('total', 0))
        label = o.get('label', oid)

        all_obligations.append({
            'id': o.get('id', ''),
            'label': label,
            'paid': paid,
            'total': total,
            'solde': _round2(total - paid),
            'pct': _round2(o.get('pct', 0)),
            'source': o.get('source', ''),
            'due_date': _isoformat(o.get('due_date')),
        })

    section['all_obligations'] = all_obligations

    # 2) Dériver le résumé depuis all_obligations (source unique)
    _map = {
        'nantel': ('nantel_paid', 'nantel_total'),
        'impot_2025_can': ('tax_can_paid', 'tax_can_total'),
        'impot_2026_can': ('tax_can_paid', 'tax_can_total'),
        'impot_2025_qc': ('tax_qc_paid', 'tax_qc_total'),
        'impot_2026_qc': ('tax_qc_paid', 'tax_qc_total'),
    }
    for obl in all_obligations:
        oid = obl['id'].lower()
        for key, (paid_field, total_field) in _map.items():
            if key in oid:
                section[paid_field] = _round2(section[paid_field] + obl['paid'])
                section[total_field] = _round2(section[total_field] + obl['total'])
                break

    # 3) Grand total (toutes obligations)
    section['grand_total'] = _round2(sum(o['total'] for o in all_obligations))
    section['grand_paid'] = _round2(sum(o['paid'] for o in all_obligations))
    section['grand_solde'] = _round2(section['grand_total'] - section['grand_paid'])

    return section


def _build_dashboard_section(ctx: Optional['CockpitContext']) -> dict:
    """Valeurs du dashboard : Excel + Python."""
    section: dict = {'excel_values': {}, 'python_values': {}}
    if ctx is None:
        return section

    section['python_values'] = {
        'score_label': ctx.score_label,
        'score_emoji': ctx.score_emoji,
        'taxes_current_month': _round2(ctx.taxes_current_month),
        'taxes_current_label': ctx.taxes_current_label or '',
        'current_treasury': _round2(ctx.current_treasury),
        'treasury_after_tax': _round2(ctx.treasury_after_tax),
        'error': ctx.error or '',
    }

    def _read_excel_dashboard():
        from core.paths import resolve_main_paths
        import openpyxl
        paths = resolve_main_paths()
        wb = openpyxl.load_workbook(paths.depenses, data_only=True, read_only=True)
        try:
            if 'TABLEAU DE BORD' not in wb.sheetnames:
                return
            ws = wb['TABLEAU DE BORD']
            cells = {
                'B10': (10, 2), 'E10': (10, 5), 'H10': (10, 8), 'K9': (9, 11),
                'H7': (7, 8), 'B7': (7, 2), 'E7': (7, 5), 'K7': (7, 11),
            }
            for label, (r, c) in cells.items():
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    try:
                        section['excel_values'][label] = _round2(val)
                    except (TypeError, ValueError):
                        section['excel_values'][label] = str(val)
                else:
                    section['excel_values'][label] = None
        finally:
            wb.close()

    _safe(_read_excel_dashboard)
    return section


def _build_git_section(outils_dir: str) -> dict:
    """État Git enrichi (Section 6) : branch, commit, ahead/behind."""
    section = {
        'branch': '',
        'last_commit_ok': True,
        'last_commit_hash': '',
        'last_commit_time': '',
        'last_commit_message': '',
        'last_push_ok': True,
        'ahead_of_origin': 0,
        'behind_origin': 0,
        'working_tree_dirty': False,
        'modified_files': [],
    }

    def _check_git():
        from dashboard.git_utils import run_git_command, find_git_repo
        repo = find_git_repo(outils_dir)
        if not repo:
            section['last_commit_ok'] = False
            section['last_push_ok'] = False
            return

        # Branche courante
        ok_br, br_out = run_git_command(
            ['rev-parse', '--abbrev-ref', 'HEAD'], repo_dir=repo, quiet=True)
        if ok_br:
            section['branch'] = br_out.strip()

        # Working tree status
        ok, out = run_git_command(
            ['status', '--porcelain'], repo_dir=repo, quiet=True)
        if ok and out.strip():
            section['working_tree_dirty'] = True
            section['modified_files'] = [
                line.strip() for line in out.strip().split('\n')
                if line.strip()
            ][:20]

        # Dernier commit : hash + message + date
        ok_log, log_out = run_git_command(
            ['log', '-1', '--format=%H|%s|%aI'], repo_dir=repo, quiet=True)
        if ok_log and log_out.strip():
            parts = log_out.strip().split('|', 2)
            if len(parts) >= 1:
                section['last_commit_hash'] = parts[0]
            if len(parts) >= 2:
                section['last_commit_message'] = parts[1]
            if len(parts) >= 3:
                section['last_commit_time'] = parts[2]

        # Ahead / behind origin
        ok_rev, rev_out = run_git_command(
            ['rev-list', '--left-right', '--count', 'HEAD...@{upstream}'],
            repo_dir=repo, quiet=True)
        if ok_rev and rev_out.strip():
            lr = rev_out.strip().split()
            if len(lr) == 2:
                section['ahead_of_origin'] = int(lr[0])
                section['behind_origin'] = int(lr[1])

        # Push status
        section['last_push_ok'] = (section['ahead_of_origin'] == 0)

    _safe(_check_git)
    return section


def _build_health_section(ctx: Optional['CockpitContext']) -> dict:
    """Résultats des health checks."""
    section = {
        'system_health_score': '',
        'checks': [],
        'warnings': [],
        'errors': [],
    }

    def _run_health():
        from dashboard.system_health import check_system_health

        tp = None
        ac = getattr(ctx, 'alert_ctx', None) if ctx else None
        if ac:
            tp = getattr(ac, 'taxes_position', None)

        report = check_system_health(taxes_position=tp)
        section['system_health_score'] = report.score

        for r in report.results:
            entry = {
                'label': r.label,
                'level': r.level.name if hasattr(r.level, 'name') else str(r.level),
                'detail': r.detail,
            }
            section['checks'].append(entry)

            level_str = entry['level'].upper()
            if level_str == 'WARNING':
                section['warnings'].append(f'{r.label}: {r.detail}')
            elif level_str == 'ERROR':
                section['errors'].append(f'{r.label}: {r.detail}')

    _safe(_run_health)
    return section


def _build_autopilot_section(ctx: Optional['CockpitContext']) -> dict:
    """Alertes autopilot depuis le CockpitContext."""
    section = {
        'risk_level': '',
        'alerts': [],
    }
    if ctx is None or not ctx.alerts:
        return section

    section['risk_level'] = ctx.score_label or ''
    for a in ctx.alerts:
        section['alerts'].append({
            'type': a.type,
            'severity': a.severity,
            'message': a.message,
            'value': _round2(a.value) if a.value is not None else None,
            'ref': a.ref or '',
        })
    return section


def _build_pipeline_stats_section() -> dict:
    """Statistiques pipeline depuis les logs (Section 5)."""
    section = {
        'gmail_emails_scanned': 0,
        'pdf_imported': 0,
        'rows_fixed': 0,
        'warnings_count': 0,
    }

    def _parse_logs():
        outils = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        log_path = os.path.join(outils, 'compta_menubar.log')
        if not os.path.exists(log_path):
            return

        today_str = date.today().strftime('%Y-%m-%d')
        with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
            for line in f:
                if today_str not in line:
                    continue
                ll = line.lower()
                if 'gmail' in ll and ('scan' in ll or 'email' in ll):
                    section['gmail_emails_scanned'] += 1
                if 'pdf' in ll and 'import' in ll:
                    section['pdf_imported'] += 1
                if 'fix' in ll or 'corrig' in ll:
                    section['rows_fixed'] += 1
                if '⚠' in line or 'warning' in ll:
                    section['warnings_count'] += 1

    _safe(_parse_logs)
    return section


def _build_internal_consistency(snapshot: dict) -> dict:
    """Validation croisée des sections (Section 4).

    Compare les valeurs entre sources pour détecter les divergences.
    """
    section = {
        'taxes_excel_vs_python_delta': None,
        'ledger_vs_obligations_delta': None,
        'dashboard_excel_vs_python_delta': None,
    }

    # 1) Taxes : Excel C23 vs Python current_month_amount
    taxes = snapshot.get('taxes', {})
    xl_c23 = taxes.get('excel_cell_C23', 0)
    py_month = taxes.get('current_month_amount', 0)
    if xl_c23 is not None and py_month is not None:
        try:
            section['taxes_excel_vs_python_delta'] = _round2(
                float(xl_c23) - float(py_month))
        except (TypeError, ValueError):
            pass

    # 2) Ledger vs Obligations : comparer les totaux payés
    ledger = snapshot.get('ledger', {})
    obligations = snapshot.get('obligations', {})
    ledger_total = ledger.get('cash_ledger_total', 0)
    obl_paid = obligations.get('grand_paid', 0)
    if ledger_total is not None and obl_paid is not None:
        try:
            section['ledger_vs_obligations_delta'] = _round2(
                float(ledger_total) - float(obl_paid))
        except (TypeError, ValueError):
            pass

    # 3) Dashboard Excel H7 vs Python taxes_current_month
    dashboard = snapshot.get('dashboard', {})
    xl_h7 = dashboard.get('excel_values', {}).get('H7')
    py_tax = dashboard.get('python_values', {}).get('taxes_current_month', 0)
    if xl_h7 is not None and py_tax is not None:
        try:
            section['dashboard_excel_vs_python_delta'] = _round2(
                float(xl_h7) - float(py_tax))
        except (TypeError, ValueError):
            pass

    return section


# ─── Validation pré-écriture (Section 8) ─────────────────────────────────────

def _validate_snapshot(snapshot: dict) -> list[str]:
    """Valide les invariants du snapshot. Retourne une liste de warnings.

    Corrige les valeurs invalides en les remplaçant par None.
    """
    warnings: list[str] = []

    # taxes.current_month_amount doit être un float >= 0
    taxes = snapshot.get('taxes', {})
    cma = taxes.get('current_month_amount')
    if cma is not None:
        try:
            v = float(cma)
            if v < 0:
                warnings.append(f'taxes.current_month_amount négatif: {v}')
                taxes['current_month_amount'] = None
        except (TypeError, ValueError):
            warnings.append(f'taxes.current_month_amount invalide: {cma!r}')
            taxes['current_month_amount'] = None

    # ledger.cash_ledger_total doit être >= 0
    ledger = snapshot.get('ledger', {})
    clt = ledger.get('cash_ledger_total')
    if clt is not None:
        try:
            v = float(clt)
            if v < 0:
                warnings.append(f'ledger.cash_ledger_total négatif: {v}')
                ledger['cash_ledger_total'] = None
        except (TypeError, ValueError):
            warnings.append(f'ledger.cash_ledger_total invalide: {clt!r}')
            ledger['cash_ledger_total'] = None

    # obligations : grand_total >= grand_paid
    obligations = snapshot.get('obligations', {})
    gt = obligations.get('grand_total', 0)
    gp = obligations.get('grand_paid', 0)
    if gt is not None and gp is not None:
        try:
            if float(gp) > float(gt) + 0.01:  # tolérance 1 cent
                warnings.append(
                    f'obligations.grand_paid ({gp}) > grand_total ({gt})')
        except (TypeError, ValueError):
            pass

    # résumé obligations vs all_obligations cohérence
    all_obls = obligations.get('all_obligations', [])
    if all_obls:
        sum_total = _round2(sum(o.get('total', 0) for o in all_obls))
        sum_paid = _round2(sum(o.get('paid', 0) for o in all_obls))
        if abs(sum_total - obligations.get('grand_total', 0)) > 0.02:
            warnings.append(
                f'obligations.grand_total ({obligations.get("grand_total")}) '
                f'!= sum(all_obligations.total) ({sum_total})')
        if abs(sum_paid - obligations.get('grand_paid', 0)) > 0.02:
            warnings.append(
                f'obligations.grand_paid ({obligations.get("grand_paid")}) '
                f'!= sum(all_obligations.paid) ({sum_paid})')

    return warnings


# ─── API publique ────────────────────────────────────────────────────────────

def build_snapshot(
    ctx: Optional['CockpitContext'] = None,
    outils_dir: str = '',
    scan_ok: bool = True,
    pipeline_stats: Optional[dict] = None,
) -> dict:
    """Construit le snapshot complet du système comptable.

    Chaque section est isolée : une erreur dans une section n'empêche
    pas les autres de se remplir.

    Parameters
    ----------
    ctx            : CockpitContext frais (ou None si indisponible)
    outils_dir     : chemin vers _OUTILS/
    scan_ok        : résultat du scan pipeline
    pipeline_stats : dict optionnel {gmail_emails_scanned, pdf_imported, ...}

    Returns
    -------
    dict — snapshot JSON-sérialisable, validé.
    """
    t0 = time.time()
    now = datetime.now()
    scan_time = now.strftime('%Y-%m-%d %H:%M:%S')

    snapshot = {
        'timestamp': now.isoformat(),

        'meta': _safe(
            lambda: _build_meta_section(t0),
            {'snapshot_version': _SNAPSHOT_VERSION, 'schema_version': _SCHEMA_VERSION}),

        'pipeline': _safe(
            lambda: _build_pipeline_section(ctx, scan_ok, scan_time),
            {'scan_status': 'ERROR', 'last_scan_time': scan_time, 'imports_count': 0}),

        'data_freshness': _safe(
            lambda: _build_data_freshness_section(scan_time),
            {'last_scan_time': scan_time, 'minutes_since_last_scan': 0}),

        'taxes': _safe(
            lambda: _build_taxes_section(ctx),
            {'current_month_label': '', 'current_month_amount': 0.0}),

        'revenue': _safe(
            lambda: _build_revenue_section(ctx),
            {'current_month': 0.0, 'ytd_total': 0.0}),

        'expenses': _safe(
            lambda: _build_expenses_section(ctx),
            {'current_month': 0.0, 'ytd_total': 0.0}),

        'ledger': _safe(
            lambda: _build_ledger_section(),
            {'cash_ledger_total': 0.0, 'ledger_entries': [], 'per_obligation': {}}),

        'obligations': _safe(
            lambda: _build_obligations_section(ctx),
            {'nantel_paid': 0.0, 'nantel_total': 0.0, 'all_obligations': []}),

        'dashboard': _safe(
            lambda: _build_dashboard_section(ctx),
            {'excel_values': {}, 'python_values': {}}),

        'git': _safe(
            lambda: _build_git_section(outils_dir),
            {'branch': '', 'last_commit_ok': False, 'working_tree_dirty': False}),

        'health_checks': _safe(
            lambda: _build_health_section(ctx),
            {'system_health_score': '', 'warnings': [], 'errors': []}),

        'autopilot': _safe(
            lambda: _build_autopilot_section(ctx),
            {'risk_level': '', 'alerts': []}),

        'pipeline_stats': _safe(
            lambda: pipeline_stats if pipeline_stats else _build_pipeline_stats_section(),
            {'gmail_emails_scanned': 0, 'pdf_imported': 0, 'rows_fixed': 0,
             'warnings_count': 0}),
    }

    # Section 4 : cohérence interne (calculée APRÈS les autres sections)
    snapshot['internal_consistency'] = _safe(
        lambda: _build_internal_consistency(snapshot),
        {'taxes_excel_vs_python_delta': None, 'ledger_vs_obligations_delta': None,
         'dashboard_excel_vs_python_delta': None})

    # Section 8 : validation pré-écriture
    validation_warnings = _safe(lambda: _validate_snapshot(snapshot), [])
    if validation_warnings:
        hc = snapshot.get('health_checks', {})
        existing = hc.get('warnings', [])
        existing.extend(
            [f'[SNAPSHOT_VALIDATION] {w}' for w in validation_warnings])
        hc['warnings'] = existing

    # Finaliser meta.pipeline_duration_ms
    try:
        snapshot['meta']['pipeline_duration_ms'] = round(
            (time.time() - t0) * 1000)
    except (KeyError, TypeError):
        pass

    return snapshot


def write_snapshot(
    snapshot: dict,
    outils_dir: str,
    filename: str = 'system_snapshot.json',
) -> str:
    """Écrit le snapshot en JSON indenté dans outils_dir/.

    Returns
    -------
    str — chemin absolu du fichier écrit.
    """
    path = os.path.join(outils_dir, filename)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(snapshot, f, indent=2, ensure_ascii=False, default=str)
    return path


# ─── CLI ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    _dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _dir not in sys.path:
        sys.path.insert(0, _dir)

    # Construire un contexte frais pour le snapshot
    try:
        from dashboard.context import build_cockpit_context
        from core.paths import resolve_main_paths
        paths = resolve_main_paths()
        ctx = build_cockpit_context(paths.ventes, paths.depenses)
    except Exception as e:
        print(f'⚠️  Contexte indisponible: {e}')
        ctx = None

    snap = build_snapshot(ctx=ctx, outils_dir=_dir)
    out  = write_snapshot(snap, _dir)
    print(f'✅ Snapshot v{_SNAPSHOT_VERSION} écrit: {out}')
    print(json.dumps(snap, indent=2, ensure_ascii=False, default=str)[:3000])
