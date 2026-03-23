#!/usr/bin/env python3
"""
dashboard/write_acomptes_tab.py — Synchronisation du registre JSON vers l'onglet
                                   ACOMPTES de DÉPENSES.xlsm.

RÔLE
────
Ce module est le pont entre la source de vérité Python (fiscal/installments_ledger.json)
et la vue Excel (onglet ACOMPTES dans DÉPENSES.xlsm).

Après chaque ajout d'acompte via le menubar, cette fonction est appelée pour :
  1. Lire fiscal/installments_ledger.json (load_installment_payments)
  2. Remplir les montants dans l'onglet ACOMPTES de DÉPENSES.xlsm
  3. Mettre à jour D14 (TOTAL VERSÉ) — cellule référencée par ESTIMATION FISCALE
     et TABLEAU DE BORD via les formules ACOMPTES!D14

⚠️  RÈGLE ABSOLUE
─────────────────
Ce module LIT le registre JSON (source de vérité) et ÉCRIT les montants
dans l'onglet ACOMPTES UNIQUEMENT.
Il ne modifie AUCUN onglet TAXES *, JANVIER, FÉVRIER, etc.
Il ne crée AUCUNE entrée dans DÉPENSES.

Structure de l'onglet ACOMPTES (DÉPENSES.xlsm)
───────────────────────────────────────────────
  R1  : titre
  R2  : sous-titre
  R4  : en-têtes  (B=Date, C=Type, D=Montant, E=Mode, F=Notes)
  R5  : Q1 ARC  (Fédéral T2)      2026-03-15
  R6  : Q1 RQ   (Provincial TP-1) 2026-03-15
  R7  : Q2 ARC                    2026-06-15
  R8  : Q2 RQ                     2026-06-15
  R9  : Q3 ARC                    2026-09-15
  R10 : Q3 RQ                     2026-09-15
  R11 : Q4 ARC                    2026-12-15
  R12 : Q4 RQ                     2026-12-15
  R14 : TOTAL VERSÉ — D14 est lue par ESTIMATION FISCALE R52 et TABLEAU DE BORD

Règle d'affectation trimestrielle
──────────────────────────────────
  Q1 : date ≤ 15 mars
  Q2 : 16 mars ≤ date ≤ 15 juin
  Q3 : 16 juin ≤ date ≤ 15 septembre
  Q4 : date > 15 septembre

Usage
─────
    from dashboard.write_acomptes_tab import write_acomptes_tab
    ok, msg = write_acomptes_tab('/path/to/DÉPENSES 2026.xlsm')
"""
from __future__ import annotations

import os
import shutil
from datetime import date, datetime
from typing import Optional

# Import à la racine du module pour permettre le mocking dans les tests
try:
    from fiscal.installments_ledger import load_installment_payments
except ImportError:
    # Fallback si le module n'est pas encore dans sys.path (import tardif dans la fn)
    load_installment_payments = None  # type: ignore


# ─── Structure fixe de l'onglet ACOMPTES ──────────────────────────────────────

#: Colonne D = Montant (1-indexed pour openpyxl)
_COL_MONTANT  = 4
_COL_DATE     = 2
_COL_TYPE     = 3
_COL_MODE     = 5
_COL_NOTES    = 6

#: Ligne TOTAL VERSÉ (D14 = lue par ESTIMATION FISCALE + TABLEAU DE BORD)
_ROW_TOTAL = 14

#: Lignes de données (Q1-Q4, ARC + RQ)
#  (quarter, authority) → row
_SLOT_ROW: dict[tuple[int, str], int] = {
    (1, 'ARC'): 5,
    (1, 'RQ'):  6,
    (2, 'ARC'): 7,
    (2, 'RQ'):  8,
    (3, 'ARC'): 9,
    (3, 'RQ'): 10,
    (4, 'ARC'): 11,
    (4, 'RQ'): 12,
}

#: Lignes supplémentaires pour paiements hors calendrier fixe (row 5-12 déjà pris)
_EXTRA_ROWS_START = 15

#: Nom de l'onglet
_SHEET_NAME = 'ACOMPTES'


# ─── Affectation trimestrielle ────────────────────────────────────────────────

def _quarter_for(payment_date: date, year: int) -> int:
    """Retourne le numéro de trimestre (1-4) pour une date de paiement.

    Calendrier des acomptes (Revenu Québec / ARC) :
      Q1 : ≤ 15 mars     Q2 : ≤ 15 juin
      Q3 : ≤ 15 sept     Q4 : ≤ 31 déc
    """
    if payment_date <= date(year, 3, 15):
        return 1
    elif payment_date <= date(year, 6, 15):
        return 2
    elif payment_date <= date(year, 9, 15):
        return 3
    else:
        return 4


# ─── Lecture du ledger ────────────────────────────────────────────────────────

def _load_payments_for_year(year: int):
    """Charge les acomptes de l'année fiscale depuis le registre JSON.

    Filtre par effective_tax_year (= tax_year si renseigné, sinon date.year).
    Ainsi un acompte payé en 2026 mais ciblant l'année fiscale 2025
    alimentera le fichier DÉPENSES 2025.xlsm, pas 2026.

    Returns list[InstallmentPayment] triée par date.
    """
    # Utiliser l'import module-level si disponible, sinon import tardif
    _loader = load_installment_payments
    if _loader is None:
        from fiscal.installments_ledger import load_installment_payments as _loader
    all_payments = _loader()
    return [p for p in all_payments if p.effective_tax_year == year]


# ─── Écriture dans l'onglet ACOMPTES ─────────────────────────────────────────

def write_acomptes_tab(
    depenses_path: str,
    year:          Optional[int] = None,
    backup:        bool          = True,
) -> tuple[bool, str]:
    """Synchronise le registre JSON vers l'onglet ACOMPTES de DÉPENSES.xlsm.

    ⚠️  Ne modifie AUCUN onglet en dehors de 'ACOMPTES'.
        Ne crée AUCUNE entrée dans les onglets mensuels (JANVIER 2026, etc.).

    Parameters
    ----------
    depenses_path : chemin absolu vers DÉPENSES 2026.xlsm
    year          : année civile (défaut : année en cours)
    backup        : si True, crée une copie avant écriture (défaut : True)

    Returns
    -------
    (success: bool, message: str)
    """
    year = year or date.today().year

    # ── Vérifications préliminaires ───────────────────────────────────────────
    if not depenses_path:
        return False, 'Chemin DÉPENSES absent'
    if not os.path.isfile(depenses_path):
        return False, f'DÉPENSES.xlsm introuvable : {depenses_path}'

    # ── Charger les paiements ─────────────────────────────────────────────────
    try:
        payments = _load_payments_for_year(year)
    except Exception as exc:
        return False, f'Lecture ledger échouée : {exc}'

    # ── Backup avant écriture ─────────────────────────────────────────────────
    if backup:
        try:
            ts  = datetime.now().strftime('%Y%m%d_%H%M%S')
            bck = f'{os.path.splitext(depenses_path)[0]}_backup_{ts}_ACOMPTES.xlsm'
            shutil.copy2(depenses_path, bck)
        except OSError:
            pass   # backup non-fatal

    # ── Ouvrir le classeur ────────────────────────────────────────────────────
    try:
        import openpyxl
        wb = openpyxl.load_workbook(depenses_path, keep_vba=True)
    except Exception as exc:
        return False, f'Ouverture DÉPENSES échouée : {exc}'

    if _SHEET_NAME not in wb.sheetnames:
        return False, f"Onglet '{_SHEET_NAME}' introuvable dans {os.path.basename(depenses_path)}"

    ws = wb[_SHEET_NAME]

    # ── Agréger les paiements par (trimestre, autorité) ───────────────────────
    totals: dict[tuple[int, str], float] = {k: 0.0 for k in _SLOT_ROW}
    extras: list[tuple] = []   # paiements hors calendrier (ne devraient pas exister)

    for p in payments:
        q   = _quarter_for(p.date, year)
        key = (q, p.authority)
        if key in totals:
            totals[key] = round(totals[key] + p.amount, 2)
        else:
            extras.append(p)

    # ── Écrire les montants dans les lignes de l'onglet ───────────────────────
    for (q, auth), amount in totals.items():
        row = _SLOT_ROW[(q, auth)]
        cell = ws.cell(row=row, column=_COL_MONTANT)
        cell.value = amount if amount > 0 else None   # None = cellule vide (plus propre)

    # ── Écrire les paiements "extra" à partir de la ligne _EXTRA_ROWS_START ──
    # (cas théorique — le ledger couvre toujours un des 8 trimestres)
    if extras:
        r = _EXTRA_ROWS_START
        # S'assurer qu'il y a de la place (ne pas écraser du contenu existant)
        while ws.cell(row=r, column=_COL_DATE).value is not None:
            r += 1
        for p in extras:
            ws.cell(row=r, column=_COL_DATE).value    = str(p.date)
            ws.cell(row=r, column=_COL_TYPE).value    = (
                'Fédéral (T2)' if p.authority == 'ARC' else 'Provincial (TP-1)')
            ws.cell(row=r, column=_COL_MONTANT).value = p.amount
            ws.cell(row=r, column=_COL_NOTES).value   = p.note or ''
            r += 1

    # ── Mettre à jour D14 — TOTAL VERSÉ ──────────────────────────────────────
    # D14 est la cellule lue par ESTIMATION FISCALE (R52) et TABLEAU DE BORD.
    # On écrit une formule SUM pour que la mise à jour reste cohérente.
    total_all = round(sum(totals.values()), 2)
    ws.cell(row=_ROW_TOTAL, column=_COL_MONTANT).value = (
        f'=SUM(D5:D12)'   # formule Excel — recalcule si on modifie manuellement
    )
    # Forcer aussi la valeur numérique dans une cellule adjacente pour fiabilité
    # (D14 aura la formule, E14 aura la valeur Python pour inspection)
    ws.cell(row=_ROW_TOTAL, column=_COL_MONTANT + 1).value = total_all or None

    # ── Forcer recalcul à l'ouverture (H10 TABLEAU DE BORD et formules liées) ─
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass   # non-fatal

    # ── Sauvegarder ──────────────────────────────────────────────────────────
    try:
        wb.save(depenses_path)
    except Exception as exc:
        return False, f'Sauvegarde DÉPENSES échouée : {exc}'

    # ── Message de retour ─────────────────────────────────────────────────────
    n  = sum(1 for v in totals.values() if v > 0)
    msg = (
        f'Onglet ACOMPTES mis à jour — {n} versement(s) — '
        f'total {year} : {total_all:,.2f} $'
    )
    return True, msg


# ─── CLI ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import sys, glob as _g
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from core.paths import resolve_main_paths

    paths = resolve_main_paths()
    dep   = paths.depenses

    # Chercher le .xlsm si c'est un dossier
    if os.path.isdir(dep):
        hits = _g.glob(os.path.join(dep, '*.xlsm'))
        dep  = hits[0] if hits else dep

    print(f'Fichier : {dep}')
    ok, msg = write_acomptes_tab(dep)
    print(f'{"✅" if ok else "❌"}  {msg}')
