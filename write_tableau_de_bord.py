#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""dashboard/write_tableau_de_bord.py
=========================================================
Écrit le BANDEAU (rows 5–10) et le COCKPIT (rows 120–143)
dans DÉPENSES 2026.xlsm.

Look ultra-premium inspiré Notion / Linear / Stripe :
  • Fond blanc · texte near-black · gris hiérarchisé
  • KPI strip horizontal (4 grands chiffres + labels discrets)
  • Bloc PRIORITÉ coloré (rouge / jaune / vert selon urgence)
  • Cockpit minimal : SITUATION · ALERTES · ACTIONS
  • Pas de bordures lourdes · spacing généreux · couleurs douces

Garanties de sécurité
─────────────────────
  • K9  (freelance_facture_total → ESTIMATION FISCALE!C7)  : JAMAIS touché
  • B10 / E10 / H10 (sync_ventes Python writes)            : JAMAIS touchés
  • Rows 11–119 (section mensuelle + tout le reste)         : JAMAIS touchées
  • Rows 137+  (graphiques, etc.)                           : JAMAIS touchées

Cellules Python injectées par ce module
────────────────────────────────────────
  Bandeau (rows 5–10) :
    H7   TaxesPosition.total_restantes        (float, pas de formule Excel)
    L9   Score emoji + label  ex: '🟢 OK'     (str, tiny — row spacer)
    K10  Action prioritaire 1                  (str, col 11 — hors merges)
    L10  Action prioritaire 2                  (str, col 12 — hors merges)

  Cockpit (rows 120–143) :
    Row 121 : En-tête SITUATION
    Rows 122–125 : 4 lignes SITUATION (formules + valeurs Python)
    Row 127 : En-tête OBLIGATIONS & PROGRESSION
    Rows 128–131 : 4 lignes obligations avec barres Unicode
    Row 133 : En-tête ALERTES
    Rows 134–136 : 3 alertes Python (colorées par sévérité)
    Row 138 : En-tête ACTIONS
    Rows 139–141 : 3 actions Python
    Row 143 : Footer mise à jour
"""
from __future__ import annotations

import os
import sys
from datetime import date
from typing import Optional, TYPE_CHECKING

# _read_obligations est la source canonique dans dashboard.context
from dashboard.context import _read_obligations  # noqa: F401 (ré-exportée pour compatibilité tests)

if TYPE_CHECKING:
    from dashboard.context import CockpitContext

# ── Constantes ────────────────────────────────────────────────────────────────
_SHEET = 'TABLEAU DE BORD'

# Cellules écrites par sync_ventes_to_depenses.py — NE JAMAIS EFFACER
_SYNC_VENTES_CELLS: set = {
    (9, 11),   # K9  — freelance_facture_total → ESTIMATION FISCALE!C7
    (10, 2),   # B10 — AR à recevoir
    (10, 5),   # E10 — revenu_total
    (10, 8),   # H10 — taxes formula string
}

# ── Palette Notion / Stripe ───────────────────────────────────────────────────
_WHITE    = 'FFFFFF'
_NEAR_BLK = '111827'   # near-black — grands chiffres KPI
_G700     = '374151'   # gris foncé — texte action
_G500     = '6B7280'   # gris moyen — labels section, libellés KV
_G400     = '9CA3AF'   # gris clair — labels KPI sous les chiffres
_G300     = 'D1D5DB'   # gris doux — barre progression future/neutre
_G200     = 'E5E7EB'   # gris très clair — séparateurs
_G100     = 'F3F4F6'   # gris quasi-blanc — alternance légère
_RED50    = 'FEF2F2'   # fond rouge pâle
_RED800   = '991B1B'   # texte rouge foncé
_YEL50    = 'FFFBEB'   # fond jaune pâle
_YEL800   = '92400E'   # texte jaune-brun foncé
_GRN50    = 'F0FDF4'   # fond vert pâle
_GRN700   = '15803D'   # texte vert foncé
_BLU50    = 'EFF6FF'   # fond bleu pâle
_BLU700   = '1D4ED8'   # texte bleu foncé

# Alias rétro-compatibilité (référencés dans tests ou scripts externes)
_NAVY    = '1F3864'
_GRAY_HD = 'D6DCE4'


# ── Helpers texte ──────────────────────────────────────────────────────────────

def _fmt(amount: float) -> str:
    """Montant lisible : 9 000 $  (espace fine insécable · style CAD)"""
    return f'{amount:,.0f}\u202f$'.replace(',', '\u202f')


def _s(ws, row: int, col: int, value) -> None:
    """Setter court."""
    ws.cell(row=row, column=col).value = value


def _top_actions(ctx: 'CockpitContext', n: int = 3) -> list[str]:
    """Retourne n actions métier priorisées (texte sobre, sans emoji)."""
    actions: list[str] = []
    for severity in ('HIGH', 'MEDIUM', 'LOW'):
        for a in (x for x in ctx.alerts if x.severity == severity):
            if len(actions) >= n:
                break
            actions.append(a.message)
        if len(actions) >= n:
            break
    while len(actions) < n:
        actions.append('\u2014')
    return actions


def _top_alerts(ctx: 'CockpitContext', n: int = 3) -> list[tuple[str, str]]:
    """Retourne n alertes : (message, severity) triées HIGH→MEDIUM→LOW."""
    result: list[tuple[str, str]] = []
    for severity in ('HIGH', 'MEDIUM', 'LOW'):
        for a in (x for x in ctx.alerts if x.severity == severity):
            if len(result) >= n:
                break
            result.append((a.message, severity))
        if len(result) >= n:
            break
    while len(result) < n:
        result.append(('\u2014', 'NONE'))
    return result


def _priority(ctx: 'CockpitContext') -> tuple[str, str, str]:
    """Retourne (texte_priorité, pale_bg_hex, dark_fg_hex) pour le bloc PRIORITÉ.

    Logique d'urgence décroissante :
      1. Acompte en retard (ERR)
      2. Alerte HIGH (ERR)
      3. Acompte imminent (WARN)
      4. Alerte MEDIUM (WARN)
      5. Tout va bien (OK)
    """
    ac = ctx.alert_ctx

    # 1. Acompte en retard — urgence maximale
    if (ac and ac.next_installment and ac.installments
            and ac.installments.required):
        nxt = ac.next_installment
        if nxt.status == 'EN RETARD':
            days = abs(nxt.days_left) if nxt.days_left is not None else '?'
            amt  = _fmt(nxt.period.amount) if nxt.period else '?'
            due  = str(nxt.period.due_date) if nxt.period else '?'
            return (
                f'Acompte en retard de {days}j \u2014 {amt} d\u00fb le {due}',
                _RED50, _RED800,
            )

    # 2. Alertes HIGH
    high = [a for a in ctx.alerts if a.severity == 'HIGH']
    if high:
        return (high[0].message, _RED50, _RED800)

    # 3. Acompte imminent
    if (ac and ac.next_installment and ac.installments
            and ac.installments.required):
        nxt = ac.next_installment
        if nxt.status == 'IMMINENT':
            days = nxt.days_left if nxt.days_left is not None else '?'
            amt  = _fmt(nxt.period.amount) if nxt.period else '?'
            due  = str(nxt.period.due_date) if nxt.period else '?'
            return (
                f'Acompte dans {days}j \u2014 {amt} le {due}',
                _YEL50, _YEL800,
            )

    # 4. Alertes MEDIUM
    medium = [a for a in ctx.alerts if a.severity == 'MEDIUM']
    if medium:
        return (medium[0].message, _YEL50, _YEL800)

    # 5. Tout va bien
    return ('Aucune alerte critique \u00b7 situation sous contr\u00f4le', _GRN50, _GRN700)


# ── Helpers styles ─────────────────────────────────────────────────────────────

def _fill(color: str):
    from openpyxl.styles import PatternFill
    return PatternFill('solid', fgColor=color)


def _font(color: str = _NEAR_BLK, bold: bool = False, size: int = 10,
          italic: bool = False):
    from openpyxl.styles import Font
    return Font(color=color, bold=bold, size=size, italic=italic, name='Calibri')


def _align(horiz: str = 'left', vert: str = 'center', wrap: bool = False):
    from openpyxl.styles import Alignment
    return Alignment(horizontal=horiz, vertical=vert, wrap_text=wrap)


def _section_header(ws, row: int, text: str, height: int = 18) -> None:
    """Header de section minimal : texte gris 8pt bold · séparateur bottom."""
    from openpyxl.styles import Border, Side
    ws.row_dimensions[row].height = height
    sep = Border(bottom=Side(style='thin', color=_G200))
    for c in range(1, 13):
        cell = ws.cell(row=row, column=c)
        cell.fill  = _fill(_WHITE)
        cell.border = sep
    lbl = ws.cell(row=row, column=2)
    lbl.value     = text.upper()
    lbl.font      = _font(_G500, bold=True, size=8)
    lbl.alignment = _align('left', 'center')


def _kv_row(
    ws,
    row:              int,
    label:            str,
    value,
    value_is_formula: bool = False,
    status_level:     str  = 'NONE',
    height:           int  = 26,
    alt:              bool = False,
) -> None:
    """Ligne KV : label B (gris 10pt) · valeur H (bold 13pt, bg conditionnel).

    Fond de la zone valeur (cols 8–12) coloré selon status_level :
      'OK'   → vert pâle  · 'WARN' → jaune pâle
      'ERR'  → rouge pâle · 'NONE' → blanc
    """
    _bg_map = {'OK': _GRN50, 'WARN': _YEL50, 'ERR': _RED50, 'NONE': _WHITE}
    _fg_map = {'OK': _GRN700, 'WARN': _YEL800, 'ERR': _RED800, 'NONE': _NEAR_BLK}
    row_bg = _G100 if alt else _WHITE
    val_bg = _bg_map.get(status_level, _WHITE)
    val_fg = _fg_map.get(status_level, _NEAR_BLK)

    ws.row_dimensions[row].height = height

    # Fond ligne complète (label zone)
    for c in range(1, 8):
        ws.cell(row=row, column=c).fill = _fill(row_bg)
    # Fond zone valeur (cols 8–12)
    for c in range(8, 13):
        ws.cell(row=row, column=c).fill = _fill(val_bg)

    # Label (col B = 2)
    cell_lbl = ws.cell(row=row, column=2)
    cell_lbl.value     = '  ' + label
    cell_lbl.font      = _font(_G500, bold=False, size=10)
    cell_lbl.alignment = _align('left', 'center')
    cell_lbl.fill      = _fill(row_bg)

    # Valeur (col H = 8) — maître visuel de la zone H:L
    cell_val = ws.cell(row=row, column=8)
    if value_is_formula:
        cell_val.value         = value
        cell_val.number_format = '#,##0\u202f[$\u202f$-C0C]'
    elif isinstance(value, float):
        cell_val.value         = value
        cell_val.number_format = '#,##0\u202f[$\u202f$-C0C]'
    else:
        cell_val.value = str(value) if value is not None else '\u2014'
    cell_val.font      = _font(val_fg, bold=True, size=13)
    cell_val.alignment = _align('right', 'center')
    cell_val.fill      = _fill(val_bg)


def _spacer(ws, row: int, height: int = 6) -> None:
    """Ligne d'espacement vide (blanc, hauteur réduite)."""
    ws.row_dimensions[row].height = height
    for c in range(1, 13):
        if (row, c) not in _SYNC_VENTES_CELLS:
            cell = ws.cell(row=row, column=c)
            cell.value = None
            cell.fill  = _fill(_WHITE)


# ── Helpers obligations ────────────────────────────────────────────────────────

def _bar(pct: float, n: int = 12) -> str:
    """Barre de progression Unicode : ████████░░░░ (12 chars)."""
    filled = max(0, min(n, round(pct * n)))
    return '\u2588' * filled + '\u2591' * (n - filled)


def _obl_color(pct: float, days_to_due: int) -> tuple[str, str]:
    """Retourne (fg_hex, bg_hex) selon la progression + l'échéance.

    Logique :
      pct >= 0.70            → vert (bon)
      pct >  0 et < 0.70     → ambre (en cours)
      pct == 0 et dû < 60j   → rouge (urgent)
      pct == 0 et futur       → gris (pas commencé)
    """
    if pct >= 0.70:
        return (_GRN700, _GRN50)
    if pct > 0:
        return (_YEL800, _YEL50)
    if days_to_due < 60:
        return (_RED800, _RED50)
    return (_G400, _WHITE)


# _read_obligations — déplacée dans dashboard/context.py (STEP 1.1 migration V2)
# Importée en tête de fichier : from dashboard.context import _read_obligations


def _obligation_row(
    ws,
    row:    int,
    label:  str,
    paid:   float,
    total:  float,
    pct:    float,
    fg:     str,
    bg:     str,
    height: int = 30,
) -> None:
    """Écrit une ligne obligation avec barre de progression Unicode.

    Layout : label (col B) · payé/total (col G) · ████░░░░ (col H–K) · % (col L)
    """
    ws.row_dimensions[row].height = height

    # Fond zone label : blanc
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = _fill(_WHITE)
    # Fond zone valeur : bg conditionnel
    for c in range(7, 13):
        ws.cell(row=row, column=c).fill = _fill(bg)

    # Label (col B)
    c_lbl = ws.cell(row=row, column=2)
    c_lbl.value     = '  ' + label
    c_lbl.font      = _font(_G700, bold=True, size=10)
    c_lbl.alignment = _align('left', 'center')

    # Payé / Total (col G)
    c_amt = ws.cell(row=row, column=7)
    c_amt.value     = f'{_fmt(paid)} / {_fmt(total)}'
    c_amt.font      = _font(_G400, bold=False, size=9)
    c_amt.alignment = _align('right', 'center')
    c_amt.fill      = _fill(bg)

    # Barre Unicode (col H–K = col 8)
    c_bar = ws.cell(row=row, column=8)
    c_bar.value     = '  ' + _bar(pct, n=12)
    c_bar.font      = _font(fg, bold=False, size=11)
    c_bar.alignment = _align('left', 'center')
    c_bar.fill      = _fill(bg)

    # Pourcentage (col L = col 12)
    c_pct = ws.cell(row=row, column=12)
    c_pct.value     = f'{pct * 100:.0f}\u202f%'
    c_pct.font      = _font(fg, bold=True, size=10)
    c_pct.alignment = _align('right', 'center')
    c_pct.fill      = _fill(bg)


# ── BANDEAU (rows 5–10) ───────────────────────────────────────────────────────

def _write_bandeau(ws, ctx: 'CockpitContext') -> None:
    """Bandeau ultra-premium Notion/Stripe (rows 5–10).

    Layout :
      Row 5  (h=20) : Header minimal — 'TABLEAU DE BORD · 2026' + date
      Row 6  (h=36) : PRIORITÉ — accent bar A6 + texte dynamique B6:L6
      Row 7  (h=46) : KPI strip — 4 grands chiffres (18pt bold)
      Row 8  (h=16) : KPI labels — 4 labels discrets (8pt gris)
      Row 9  (h=8)  : Spacer — K9 JAMAIS touché · L9=score
      Row 10 (h=1)  : Données cachées — B10/E10/H10 protégés (K10/L10 = vides)

    Cellules préservées : K9=(9,11) · B10=(10,2) · E10=(10,5) · H10=(10,8)
    """
    from openpyxl.styles import Border, Side, Font, Alignment

    ac = ctx.alert_ctx

    # ── Nettoyage rows 1–4 : ancien header pré-refactor ──────────────────────
    # Défusion AVANT l'effacement pour éviter AttributeError sur MergedCell
    to_unmerge_14 = [str(mr) for mr in list(ws.merged_cells.ranges)
                     if mr.min_row <= 4]
    for rng in to_unmerge_14:
        ws.unmerge_cells(rng)
    for r in range(1, 5):
        ws.row_dimensions[r].height = 3  # quasi-invisible
        for c in range(1, 30):
            cell = ws.cell(row=r, column=c)
            cell.value     = None
            cell.fill      = _fill(_WHITE)
            cell.font      = Font(name='Calibri', size=10)
            cell.alignment = Alignment()
            cell.border    = Border()

    # ── Défusion de toutes les plages dans rows 5–10 ──────────────────────────
    to_unmerge = [str(mr) for mr in list(ws.merged_cells.ranges)
                  if 5 <= mr.min_row <= 10 or 5 <= mr.max_row <= 10]
    for rng in to_unmerge:
        ws.unmerge_cells(rng)

    # ── Effacement ciblé rows 5–9 (sauf K9) ──────────────────────────────────
    for r in range(5, 10):
        for c in range(1, 17):
            if (r, c) not in _SYNC_VENTES_CELLS:
                cell = ws.cell(row=r, column=c)
                cell.value     = None
                cell.fill      = _fill(_WHITE)
                cell.font      = Font(name='Calibri', size=10)
                cell.alignment = Alignment()
                cell.border    = Border()

    # ── Effacement row 10 — uniquement cols non-protégées ─────────────────────
    for c in range(1, 17):
        if (10, c) not in _SYNC_VENTES_CELLS:
            ws.cell(row=10, column=c).value = None

    # ── Row 5 (h=20) : Header minimal ─────────────────────────────────────────
    ws.row_dimensions[5].height = 20
    bot = Border(bottom=Side(style='thin', color=_G200))
    for c in range(1, 13):
        cell = ws.cell(row=5, column=c)
        cell.fill   = _fill(_WHITE)
        cell.border = bot

    c5t = ws.cell(row=5, column=2)
    c5t.value     = '  TABLEAU DE BORD \u00b7 2026'
    c5t.font      = _font(_G500, bold=True, size=9)
    c5t.alignment = _align('left', 'center')

    c5d = ws.cell(row=5, column=12)
    c5d.value     = '=TEXT(TODAY(),"DD MMM YYYY")'
    c5d.font      = _font(_G400, bold=False, size=9)
    c5d.alignment = _align('right', 'center')

    # ── Row 6 (h=36) : PRIORITÉ ───────────────────────────────────────────────
    ws.row_dimensions[6].height = 36
    prio_text, prio_pale, prio_dark = _priority(ctx)

    # A6 = barre d'accent (couleur statut pleine)
    a6 = ws.cell(row=6, column=1)
    a6.fill   = _fill(prio_dark)
    a6.border = Border()

    # B6:L6 = fond pâle + texte priorité (B6 = master après merge)
    for c in range(2, 13):
        ws.cell(row=6, column=c).fill   = _fill(prio_pale)
        ws.cell(row=6, column=c).border = Border()

    c6 = ws.cell(row=6, column=2)
    c6.value     = '  ' + prio_text
    c6.font      = _font(prio_dark, bold=True, size=11)
    c6.alignment = _align('left', 'center')

    # ── Row 7 (h=46) : KPI values ─────────────────────────────────────────────
    ws.row_dimensions[7].height = 46
    sep_left = Border(left=Side(style='thin', color=_G200))

    for c in range(1, 13):
        ws.cell(row=7, column=c).fill = _fill(_WHITE)

    # KPI 1 : Cash — B7 (master B7:D7)
    c7_cash = ws.cell(row=7, column=2)
    c7_cash.value         = "='STRAT\u00c9GIES FISCALES'!B12"
    c7_cash.font          = _font(_NEAR_BLK, bold=True, size=18)
    c7_cash.alignment     = _align('center', 'center')
    c7_cash.number_format = '#,##0\u202f[$\u202f$-C0C]'

    # KPI 2 : AR — E7 (master E7:G7)
    c7_ar = ws.cell(row=7, column=5)
    c7_ar.value         = '=B10'
    c7_ar.font          = _font(_NEAR_BLK, bold=True, size=18)
    c7_ar.alignment     = _align('center', 'center')
    c7_ar.number_format = '#,##0\u202f[$\u202f$-C0C]'
    c7_ar.border        = sep_left

    # KPI 3 : TPS/TVQ — H7 (master H7:J7) — Python float, jamais formule
    # FIX 6 : Préférer taxes_current_month (montant mensuel à remettre)
    # quand disponible, au lieu du cumul annuel total_restantes.
    _tcm = getattr(ctx, 'taxes_current_month', 0.0) or 0.0
    if _tcm > 0:
        taxes_val = round(_tcm, 2)
    elif ac is not None and ac.taxes_position is not None:
        taxes_val = round(ac.taxes_position.total_restantes, 2)
    else:
        taxes_val = 0.0
    c7_tax = ws.cell(row=7, column=8)
    c7_tax.value         = taxes_val
    c7_tax.font          = _font(_NEAR_BLK, bold=True, size=18)
    c7_tax.alignment     = _align('center', 'center')
    c7_tax.number_format = '#,##0\u202f[$\u202f$-C0C]'
    c7_tax.border        = sep_left

    # KPI 4 : Impôt 2025 — K7 (master K7:L7)
    c7_imp = ws.cell(row=7, column=11)
    c7_imp.value         = '=OBLIGATIONS!L3+OBLIGATIONS!L4'
    c7_imp.font          = _font(_NEAR_BLK, bold=True, size=18)
    c7_imp.alignment     = _align('center', 'center')
    c7_imp.number_format = '#,##0\u202f[$\u202f$-C0C]'
    c7_imp.border        = sep_left

    # ── Row 8 (h=16) : KPI labels ─────────────────────────────────────────────
    ws.row_dimensions[8].height = 16
    bot8 = Border(bottom=Side(style='thin', color=_G200))
    for c in range(1, 13):
        cell = ws.cell(row=8, column=c)
        cell.fill   = _fill(_WHITE)
        cell.border = bot8

    for col, lbl in [
        (2,  'CASH DISPONIBLE'),
        (5,  'AR \u00c0 RECEVOIR'),
        (8,  'TPS / TVQ'),
        (11, 'IMP\u00d4T 2025'),
    ]:
        cell = ws.cell(row=8, column=col)
        cell.value     = lbl
        cell.font      = _font(_G400, bold=False, size=8)
        cell.alignment = _align('center', 'center')
        cell.fill      = _fill(_WHITE)

    # ── Row 9 (h=8) : Spacer fin — K9 JAMAIS touché ──────────────────────────
    ws.row_dimensions[9].height = 8
    for c in range(1, 13):
        if (9, c) not in _SYNC_VENTES_CELLS:
            cell = ws.cell(row=9, column=c)
            cell.fill   = _fill(_WHITE)
            cell.border = Border()

    # L9 (col 12) = score — discret dans le spacer (7pt gris)
    c9s = ws.cell(row=9, column=12)
    c9s.value     = f'{ctx.score_emoji} {ctx.score_label}'
    c9s.font      = _font(_G400, bold=False, size=7)
    c9s.alignment = _align('right', 'center')
    c9s.fill      = _fill(_WHITE)

    # ── Row 10 (h=1) : données cachées — K10/L10 volontairement vides ────────
    # Les actions sont affichées dans le cockpit (rows 133–135).
    # K10/L10 ne reçoivent plus de texte pour éviter tout débordement visuel.
    ws.row_dimensions[10].height = 1

    # ── Fusions ───────────────────────────────────────────────────────────────
    ws.merge_cells('B6:L6')    # PRIORITÉ texte
    ws.merge_cells('B7:D7')    # KPI Cash
    ws.merge_cells('E7:G7')    # KPI AR
    ws.merge_cells('H7:J7')    # KPI TPS/TVQ (H7 = master Python float)
    ws.merge_cells('K7:L7')    # KPI Impôt 2025
    ws.merge_cells('B8:D8')    # Label Cash
    ws.merge_cells('E8:G8')    # Label AR
    ws.merge_cells('H8:J8')    # Label TPS/TVQ
    ws.merge_cells('K8:L8')    # Label Impôt
    ws.merge_cells('B10:D10')  # B10 master protégé
    ws.merge_cells('E10:G10')  # E10 master protégé
    ws.merge_cells('H10:J10')  # H10 master protégé


# ── COCKPIT (rows 120–136) ────────────────────────────────────────────────────

def _write_cockpit(
    ws,
    ctx: 'CockpitContext',
    ytd_2025: float,
    ytd_2026: float,
    today: date,
    obligations: list[dict] | None = None,
) -> None:
    """Cockpit Notion/Stripe avec section OBLIGATIONS à partir de la row 120.

    Layout :
      120     : Spacer
      121     : SITUATION — section header
      122–125 : 4 lignes KV (TPS/TVQ · Impôt 2025 · Acomptes · Prochain)
      126     : Spacer
      127     : OBLIGATIONS & PROGRESSION — section header
      128–131 : 4 lignes obligations avec barres Unicode
      132     : Spacer
      133     : ALERTES — section header
      134–136 : 3 alertes Python (fond coloré par sévérité)
      137     : Spacer
      138     : ACTIONS — section header
      139–141 : 3 actions Python
      142     : Spacer
      143     : Footer mise à jour
    """
    from openpyxl.styles import Font, Alignment, Border, Side

    ac  = ctx.alert_ctx
    obl = obligations or []

    # ── Effacement rows 120–165 ───────────────────────────────────────────────
    for r in range(120, 166):
        for c in range(1, 17):
            cell = ws.cell(row=r, column=c)
            cell.value     = None
            cell.fill      = _fill(_WHITE)
            cell.font      = Font(name='Calibri', size=10)
            cell.alignment = Alignment()
            cell.border    = Border()
        ws.row_dimensions[r].height = None

    # ── Row 120 : Spacer ──────────────────────────────────────────────────────
    _spacer(ws, 120, height=12)

    # ══════════════════════════════════════════════════════════════════════════
    # SITUATION (rows 121–125)
    # ══════════════════════════════════════════════════════════════════════════
    _section_header(ws, 121, 'Situation financi\u00e8re')

    # Row 122 : TPS / TVQ restantes
    if ac is not None and ac.taxes_position is not None:
        tp = ac.taxes_position
        tv = tp.total_restantes
        taxes_lvl = 'ERR' if tv > 5000 else ('WARN' if tv > 0 else 'OK')
    else:
        taxes_lvl = 'NONE'

    _kv_row(
        ws, 122,
        label='TPS / TVQ \u00e0 remettre',
        value='=H7',
        value_is_formula=True,
        status_level=taxes_lvl,
    )

    # Row 123 : Impôt 2025 restant
    _kv_row(
        ws, 123,
        label='Imp\u00f4t 2025 restant',
        value='=OBLIGATIONS!L3+OBLIGATIONS!L4',
        value_is_formula=True,
        status_level='NONE',
        alt=True,
    )

    # Row 124 : Acomptes versés AF 2025
    ytd25_lvl = 'OK' if ytd_2025 > 0 else 'NONE'
    _kv_row(
        ws, 124,
        label='Acomptes vers\u00e9s AF 2025',
        value=round(ytd_2025, 2),
        status_level=ytd25_lvl,
    )

    # Row 125 : Prochain acompte
    nxt = ac.next_installment if (ac is not None) else None
    inst_required = (
        ac is not None
        and ac.installments is not None
        and ac.installments.required
    )
    if nxt and nxt.period and inst_required:
        p = nxt.period
        prio_str = f'{str(p.due_date)} \u00b7 {_fmt(p.amount)}'
        nxt_lvl  = (
            'ERR'  if nxt.status == 'EN RETARD'  else
            'WARN' if nxt.status == 'IMMINENT'    else
            'NONE'
        )
    elif ac is not None and ac.installments and not ac.installments.required:
        prio_str = 'Non requis cette ann\u00e9e'
        nxt_lvl  = 'OK'
    else:
        prio_str = '\u2014'
        nxt_lvl  = 'NONE'

    _kv_row(
        ws, 125,
        label='Prochain acompte',
        value=prio_str,
        status_level=nxt_lvl,
        alt=True,
    )

    # ── Row 126 : Spacer ──────────────────────────────────────────────────────
    _spacer(ws, 126, height=8)

    # ══════════════════════════════════════════════════════════════════════════
    # OBLIGATIONS & PROGRESSION (rows 127–131)
    # ══════════════════════════════════════════════════════════════════════════
    _section_header(ws, 127, 'Obligations & progression')

    # Afficher jusqu'à 4 obligations actives
    for i, ob in enumerate(obl[:4]):
        r   = 128 + i
        pct = ob['pct']
        due = ob.get('due_date')
        days_to_due = (due - today).days if due else 9999
        fg, bg = _obl_color(pct, days_to_due)

        _obligation_row(
            ws, r,
            label=ob['label'],
            paid=ob['paid'],
            total=ob['total'],
            pct=pct,
            fg=fg,
            bg=bg,
        )

    # Si < 4 obligations, lignes vides
    for i in range(len(obl[:4]), 4):
        r = 128 + i
        ws.row_dimensions[r].height = 30
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = _fill(_WHITE)
        ws.cell(row=r, column=2).value = '  \u2014'
        ws.cell(row=r, column=2).font  = _font(_G400, bold=False, size=10)

    # ── Row 132 : Spacer ──────────────────────────────────────────────────────
    _spacer(ws, 132, height=8)

    # ══════════════════════════════════════════════════════════════════════════
    # ALERTES (rows 133–136)
    # ══════════════════════════════════════════════════════════════════════════
    _section_header(ws, 133, 'Alertes actives')

    _sev_bg = {'HIGH': _RED50,  'MEDIUM': _YEL50,  'LOW': _BLU50,  'NONE': _WHITE}
    _sev_fg = {'HIGH': _RED800, 'MEDIUM': _YEL800, 'LOW': _BLU700, 'NONE': _G400}

    for i, (msg, sev) in enumerate(_top_alerts(ctx, n=3)):
        r  = 134 + i
        bg = _sev_bg.get(sev, _WHITE)
        fg = _sev_fg.get(sev, _G400)
        ws.row_dimensions[r].height = 24
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = _fill(bg)
        cell = ws.cell(row=r, column=2)
        cell.value     = ('  \u2022 ' + msg) if msg != '\u2014' else '  \u2014'
        cell.font      = _font(fg, bold=(sev in ('HIGH', 'MEDIUM')), size=10)
        cell.alignment = _align('left', 'center')
        cell.fill      = _fill(bg)

    # ── Row 137 : Spacer ──────────────────────────────────────────────────────
    _spacer(ws, 137, height=8)

    # ══════════════════════════════════════════════════════════════════════════
    # ACTIONS (rows 138–141)
    # ══════════════════════════════════════════════════════════════════════════
    _section_header(ws, 138, 'Actions prioritaires')

    for i, action_text in enumerate(_top_actions(ctx, n=3)):
        r      = 139 + i
        row_bg = _G100 if i % 2 == 1 else _WHITE
        ws.row_dimensions[r].height = 24
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = _fill(row_bg)
        cell = ws.cell(row=r, column=2)
        cell.value     = f'  \u2192  {action_text}'
        cell.font      = _font(_G700, bold=False, size=10)
        cell.alignment = _align('left', 'center')
        cell.fill      = _fill(row_bg)

    # ── Row 142–143 : Footer ──────────────────────────────────────────────────
    _spacer(ws, 142, height=4)
    ws.row_dimensions[143].height = 14
    top_sep = Border(top=Side(style='thin', color=_G200))
    for c in range(1, 13):
        cell = ws.cell(row=143, column=c)
        cell.fill   = _fill(_WHITE)
        cell.border = top_sep
    c143 = ws.cell(row=143, column=2)
    c143.value     = '="Mis \u00e0 jour : "&TEXT(NOW(),"DD MMM YYYY \u00e0 HH:MM")'
    c143.font      = _font(_G400, bold=False, size=8, italic=True)
    c143.alignment = _align('left', 'center')


# ── API publique ──────────────────────────────────────────────────────────────

def write_tableau_de_bord(
    depenses_path: str,
    *,
    ctx: 'CockpitContext',
    ytd_2025: Optional[float] = None,
    ytd_2026: float = 0.0,
    today: Optional[date] = None,
    save: bool = True,
):
    """Écrit le TABLEAU DE BORD (bandeau KPI + cockpit Notion/Stripe).

    Parameters
    ----------
    depenses_path : Chemin absolu vers DÉPENSES 2026.xlsm
    ctx           : CockpitContext depuis build_cockpit_context()
    ytd_2025      : Total acomptes versés AF 2025 — si None, utilise ctx.ytd_2025 (STEP 1.3)
    ytd_2026      : Total acomptes versés AF 2026 — ytd_total(2026)
    today         : Date de référence (défaut : date.today())
    save          : Si True, sauvegarde (False en mode test)

    Returns
    -------
    openpyxl.Workbook (utile pour les assertions de test)
    """
    import openpyxl
    from pathlib import Path

    today = today or date.today()
    path  = Path(depenses_path)

    if not path.exists():
        raise FileNotFoundError(f'Fichier introuvable : {path}')

    try:
        _sd = os.path.dirname(os.path.abspath(__file__))
        sys.path.insert(0, os.path.dirname(_sd))
        from core.paths import assert_not_icloud
        assert_not_icloud(str(path), context='write_tableau_de_bord')
    except ImportError:
        pass

    wb = openpyxl.load_workbook(str(path), keep_vba=True)

    if _SHEET not in wb.sheetnames:
        raise ValueError(f"Onglet '{_SHEET}' introuvable dans {path.name}")

    ws = wb[_SHEET]

    # Lire les obligations depuis la feuille OBLIGATIONS (si elle existe)
    obligations = _read_obligations(wb)

    # Résoudre ytd_2025 : fallback sur ctx.ytd_2025 si non passé explicitement (STEP 1.3)
    _ytd_2025 = ytd_2025 if ytd_2025 is not None else getattr(ctx, 'ytd_2025', 0.0)

    _write_bandeau(ws, ctx)
    _write_cockpit(ws, ctx, _ytd_2025, ytd_2026, today, obligations=obligations)

    if save:
        try:
            _sd = os.path.dirname(os.path.abspath(__file__))
            sys.path.insert(0, os.path.dirname(_sd))
            from core.paths import safe_excel_write
            safe_excel_write(str(path), wb)
        except ImportError:
            wb.save(str(path))

    return wb


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import argparse

    ap = argparse.ArgumentParser(description='Écriture TABLEAU DE BORD — style Notion/Stripe')
    ap.add_argument('--depenses', help='Chemin vers DÉPENSES 2026.xlsm')
    ap.add_argument('--dry-run', action='store_true', help='Ne pas sauvegarder')
    args = ap.parse_args()

    _script_dir = os.path.dirname(os.path.abspath(__file__))
    sys.path.insert(0, os.path.dirname(_script_dir))

    from core.paths import resolve_main_paths
    from dashboard.context import build_cockpit_context
    from fiscal.installments_ledger import ytd_total

    _paths   = resolve_main_paths(2026)
    dep_path = args.depenses or _paths.depenses

    print(f'Fichier : {dep_path}')
    ctx_obj = build_cockpit_context(_paths.ventes_dir, _paths.depenses_dir)
    # ytd_2025 est maintenant dans ctx_obj.ytd_2025 (STEP 1.3)
    y26 = ytd_total(2026)

    write_tableau_de_bord(
        dep_path,
        ctx=ctx_obj,
        ytd_2026=y26,
        save=not args.dry_run,
    )
    print('✅ Dry-run terminé.' if args.dry_run else '✅ TABLEAU DE BORD mis à jour.')
